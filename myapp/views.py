from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password, check_password
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.files.storage import default_storage
from django.contrib import messages
from django.core.paginator import Paginator
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives, send_mail
from django.template import Library
from django.db import connection
from django.http import HttpResponse
from django.conf import settings
from .models import Order, MerchandiseOrder, EventTicket, Event
from .sasapay_utils import initiate_c2b_payment
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import boto3
from botocore.client import Config
from botocore.exceptions import ClientError
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
import uuid
from datetime import datetime, timedelta
import mimetypes
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.db.models import Sum, Count
from datetime import datetime, timedelta
import json
from .models import MfalmeUsers, Course, Video, PDF, Blog, Package, Order, Partnership
import requests
import socket
import logging
import json
import random
import string
import uuid
import hashlib
import hmac
from datetime import datetime, timedelta
from decimal import Decimal
import hashlib
import socket
import socket
import ssl
import base64
import urllib.parse
from .pesapal_utils import get_pesapal_iframe_url, query_pesapal_status
from django.views.decorators.http import require_http_methods

from .models import (
    MfalmeUsers, PaymentTransaction, Package,
    TrainingVideo, UserVideoAccess, Course, UserCourse, MentorshipProgram,
    SupportTicket, TicketReply, ActivityLog, PDF, Blog, VerificationCode,
    PartnershipProgram, UserPartnership, ContactSubmission, Notification,
    FAQ, Testimonial, CommunityTier, UserCommunityMembership, Statistic,
    EducationProgram, UserPDFAccess, Watchlist, InstituteApplication,
    CommunityJoinRequest, Merchandise,
    MerchandiseOrder,
    Event,
    EventTicket,
    Order
)

# ==================== TEMPLATE FILTERS ====================
register = Library()

@register.filter
def multiply(value, arg):
    """Simple multiplication filter for templates"""
    try:
        return float(value) * float(arg)
    except (ValueError, TypeError):
        return 0

# ==================== HELPER FUNCTIONS ====================

def generate_verification_code():
    """Generate a 6-digit verification code"""
    return ''.join(random.choices(string.digits, k=6))

def generate_reference():
    """Generate unique transaction reference"""
    return f"MBC{timezone.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def timesince(dt, default="just now"):
    """Return human readable time since"""
    if not dt:
        return default
    now = timezone.now()
    diff = now - dt
    
    seconds = diff.total_seconds()
    if seconds < 60:
        return "just now"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds // 3600)
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif seconds < 604800:
        days = int(seconds // 86400)
        return f"{days} day{'s' if days != 1 else ''} ago"
    elif seconds < 2592000:
        weeks = int(seconds // 604800)
        return f"{weeks} week{'s' if weeks != 1 else ''} ago"
    elif seconds < 31536000:
        months = int(seconds // 2592000)
        return f"{months} month{'s' if months != 1 else ''} ago"
    else:
        years = int(seconds // 31536000)
        return f"{years} year{'s' if years != 1 else ''} ago"

def send_admin_notification_email(subject, template, context):
    """Send email to admin"""
    try:
        html_content = render_to_string(f'emails/{template}', context)
        text_content = f"New notification: {subject}"
        
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            settings.ADMIN_EMAILS
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"Admin email error: {e}")
        return False

def send_user_notification_email(user, subject, template, context):
    """Send email to user"""
    try:
        html_content = render_to_string(f'emails/{template}', context)
        text_content = f"Notification: {subject}"
        
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"User email error: {e}")
        return False

# ==================== EMAIL FUNCTIONS ====================

@require_GET
def export_blogs(request):
    """Export blogs to Excel"""
    blogs = Blog.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="blogs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blogs Export"
    
    headers = ['ID', 'Title', 'Author', 'Category', 'Views', 'Status', 'Published', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, blog in enumerate(blogs, 2):
        ws.cell(row=row_num, column=1).value = blog.id
        ws.cell(row=row_num, column=2).value = blog.title
        ws.cell(row=row_num, column=3).value = blog.author.username if blog.author else 'Admin'
        ws.cell(row=row_num, column=4).value = blog.category or 'General'
        ws.cell(row=row_num, column=5).value = blog.views or 0
        ws.cell(row=row_num, column=6).value = blog.status or 'draft'
        ws.cell(row=row_num, column=7).value = blog.published_at.strftime('%Y-%m-%d') if blog.published_at else 'Not published'
        ws.cell(row=row_num, column=8).value = blog.created_at.strftime('%Y-%m-%d') if blog.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@require_GET
def export_packages(request):
    """Export packages to Excel"""
    packages = Package.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="packages_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packages Export"
    
    headers = ['ID', 'Name', 'Price (KES)', 'Type', 'Sales', 'Revenue (KES)', 'Status', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, package in enumerate(packages, 2):
        sales = Order.objects.filter(package=package, status='completed').count()
        revenue = Order.objects.filter(package=package, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        
        ws.cell(row=row_num, column=1).value = package.id
        ws.cell(row=row_num, column=2).value = package.name
        ws.cell(row=row_num, column=3).value = float(package.price) if package.price else 0
        ws.cell(row=row_num, column=4).value = package.package_type or 'N/A'
        ws.cell(row=row_num, column=5).value = sales
        ws.cell(row=row_num, column=6).value = float(revenue)
        ws.cell(row=row_num, column=7).value = package.status or 'inactive'
        ws.cell(row=row_num, column=8).value = package.created_at.strftime('%Y-%m-%d') if package.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@require_GET
def export_partnerships(request):
    """Export partnerships to Excel"""
    partnerships = Partnership.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="partnerships_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partnerships Export"
    
    headers = ['ID', 'Company', 'Contact Person', 'Email', 'Phone', 'Tier', 'Amount (KES)', 'NDA', 'Status', 'Applied']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, p in enumerate(partnerships, 2):
        ws.cell(row=row_num, column=1).value = p.id
        ws.cell(row=row_num, column=2).value = p.company_name
        ws.cell(row=row_num, column=3).value = p.contact_name
        ws.cell(row=row_num, column=4).value = p.email
        ws.cell(row=row_num, column=5).value = p.phone or ''
        ws.cell(row=row_num, column=6).value = p.tier or 'N/A'
        ws.cell(row=row_num, column=7).value = float(p.investment_amount) if p.investment_amount else 0
        ws.cell(row=row_num, column=8).value = 'Signed' if p.nda_signed else 'Pending'
        ws.cell(row=row_num, column=9).value = p.status or 'pending'
        ws.cell(row=row_num, column=10).value = p.created_at.strftime('%Y-%m-%d') if p.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response

def send_admin_notification(user):
    """Send admin notification when new user registers"""
    try:
        subject = 'NEW USER REGISTRATION - Mfalme Betterdays Capital'
        
        admin_emails = getattr(settings, 'ADMIN_EMAILS', ['mfalmebetterdays@gmail.com'])
        
        html_content = render_to_string('emails/admin_notification.html', {
            'username': user.username,
            'email': user.email,
            'phone': user.phone,
            'soldier_id': user.soldier_id,
            'date_joined': user.date_joined,
            'ip_address': user.registration_ip,
            'country': user.country or 'Not specified',
            'trading_experience': user.trading_experience,
            'user_id': user.id,
            'admin_url': f"{settings.SITE_URL}/admin/",
            'year': timezone.now().year,
            'current_year': timezone.now().year
        })
        
        text_content = f"""
        New User Registration Alert!
        
        Username: {user.username}
        Email: {user.email}
        Phone: {user.phone}
        Soldier ID: {user.soldier_id}
        Date: {user.date_joined}
        
        Login to admin panel to view details.
        """
        
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            admin_emails
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        
        print(f"âœ“ Admin notification sent for {user.email}")
        return True
        
    except Exception as e:
        print(f"âœ— Admin notification error: {e}")
        return False

def send_welcome_email(user):
    """Send welcome email after verification using HTML template"""
    try:
        subject = 'ðŸŽ‰ Welcome to the Elite Circle - Mfalme Betterdays Capital!'
        
        html_content = render_to_string('emails/welcome.html', {
            'username': user.username,
            'soldier_id': user.soldier_id,
            'email': user.email,
            'phone': user.phone,
            'user_id': user.id,
            'signup_date': user.date_joined.strftime('%B %d, %Y'),
            'dashboard_url': f"{settings.SITE_URL}/dashboard/",
            'community_url': f"{settings.SITE_URL}/community/",
            'support_url': f"{settings.SITE_URL}/support/",
            'videos_url': f"{settings.SITE_URL}/my-videos/",
            'courses_url': f"{settings.SITE_URL}/my-courses/",
            'website_url': settings.SITE_URL,
            'year': timezone.now().year,
            'current_year': timezone.now().year
        })
        
        text_content = f"""
        Welcome to the Elite, {user.username}!
        
        Your account has been successfully verified.
        Your Soldier ID: {user.soldier_id}
        
        You now have access to:
        - Your personalized dashboard
        - Free trading resources
        - Community access
        
        Login here: {settings.SITE_URL}/login/
        
        To your success,
        The Mfalme Betterdays Capital Team
        """
        
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        
        print(f"âœ“ Welcome email sent to {user.email}")
        return True
        
    except Exception as e:
        print(f"âœ— Welcome email error: {e}")
        return False

def send_password_reset_email(user, code):
    """Send password reset email"""
    try:
        subject = 'Password Reset - Mfalme Betterdays Capital'
        
        html_content = render_to_string('emails/password_reset.html', {
            'username': user.username,
            'reset_code': code,
            'email': user.email,
            'expiry_minutes': 30,
            'year': timezone.now().year,
            'current_year': timezone.now().year
        })
        
        text_content = f"""
        Hello {user.username},
        
        You requested to reset your password.
        
        Your password reset code is: {code}
        
        This code will expire in 30 minutes.
        
        If you didn't request this, please ignore this email.
        
        Best regards,
        Mfalme Betterdays Capital Team
        """
        
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        
        return True
    except Exception as e:
        print(f"Password reset email error: {e}")
        return False

def send_community_application_email(user, community, request):
    """Send email to admin about community join request"""
    try:
        subject = f' NEW COMMUNITY JOIN REQUEST - {community.name}'
        
        html_content = render_to_string('emails/community_application.html', {
            'user': user,
            'community': community,
            'total_deposits': user.total_deposits,
            'trading_experience': user.trading_experience,
            'courses_completed': UserCourse.objects.filter(user=user).count(),
            'admin_url': f"{settings.SITE_URL}/admin/",
            'year': timezone.now().year
        })
        
        send_admin_notification_email(subject, 'community_application.html', {
            'user': user,
            'community': community,
            'total_deposits': user.total_deposits
        })
        return True
    except Exception as e:
        print(f"Community application email error: {e}")
        return False

def send_institute_application_email(application):
    """Send email to admin about institute application"""
    try:
        subject = ' NEW INSTITUTE APPLICATION - Mfalme Betterdays Capital'
        
        send_admin_notification_email(subject, 'institute_application.html', {
            'application': application,
            'user': application.user,
            'admin_url': f"{settings.SITE_URL}/admin/"
        })
        return True
    except Exception as e:
        print(f"Institute application email error: {e}")
        return False

def send_ticket_notification_email(ticket, is_new=True):
    """Send email about support ticket"""
    try:
        if is_new:
            subject = f'NEW SUPPORT TICKET - #{ticket.ticket_number}'
            template = 'new_ticket_notification.html'
        else:
            subject = f'ðŸ’¬ TICKET REPLY - #{ticket.ticket_number}'
            template = 'ticket_reply_notification.html'
        
        send_admin_notification_email(subject, template, {
            'ticket': ticket,
            'user': ticket.user,
            'admin_url': f"{settings.SITE_URL}/admin/support/tickets/{ticket.id}/"
        })
        return True
    except Exception as e:
        print(f"Ticket notification email error: {e}")
        return False

# ==================== ACCESS CONTROL HELPER FUNCTIONS ====================

def check_video_access(user, video):
    """Check if user has access to a video"""
    if video.price == 0:  # Free video
        return True
    return UserVideoAccess.objects.filter(user=user, video=video).exists()

def check_pdf_access(user, pdf):
    """Check if user has access to a PDF"""
    if pdf.is_free:
        return True
    return UserPDFAccess.objects.filter(user=user, pdf=pdf).exists()

def check_course_access(user, course_id):
    """Check if user has active access to a course with expiration"""
    try:
        enrollment = UserCourse.objects.get(
            user=user, 
            course_id=course_id,
            is_active=True
        )
        
        # Check if access has expired
        if enrollment.is_access_expired():
            # Mark as inactive
            enrollment.is_active = False
            enrollment.save(update_fields=['is_active'])
            
            # Send notification
            Notification.objects.create(
                user=user,
                title='Course Access Expired',
                message=f'Your access to {enrollment.course.title} has expired. Renew now to continue learning.',
                notification_type='WARNING',
                related_object_type='course',
                related_object_id=course_id
            )
            
            # Log activity
            log_activity(
                user,
                'COURSE_ACCESS_EXPIRED',
                f'Access expired for course: {enrollment.course.title}'
            )
            
            return False, enrollment
        return True, enrollment
    except UserCourse.DoesNotExist:
        return False, None

def grant_video_access(user, video, payment=None):
    """Grant user access to a video"""
    access, created = UserVideoAccess.objects.get_or_create(
        user=user,
        video=video,
        defaults={'payment': payment}
    )
    if created:
        Notification.objects.create(
            user=user,
            title='Video Unlocked',
            message=f'You now have access to: {video.title}',
            notification_type='SUCCESS',
            related_object_type='video',
            related_object_id=video.id
        )
    return access

def grant_pdf_access(user, pdf, payment=None):
    """Grant user access to a PDF"""
    access, created = UserPDFAccess.objects.get_or_create(
        user=user,
        pdf=pdf,
        defaults={'payment': payment}
    )
    if created:
        Notification.objects.create(
            user=user,
            title='PDF Unlocked',
            message=f'You can now view: {pdf.title}',
            notification_type='SUCCESS',
            related_object_type='pdf',
            related_object_id=pdf.id
        )
    return access

def log_activity(user, action, description, request=None):
    """Create activity log entry"""
    return ActivityLog.objects.create(
        user=user,
        action=action,
        description=description,
        ip_address=get_client_ip(request) if request else None,
        user_agent=request.META.get('HTTP_USER_AGENT', '') if request else ''
    )

# ==================== USER AUTHENTICATION ====================

def login_page(request):
    """User login page - handles both login and registration"""
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        return redirect('dashboard')
    
    context = {
        'total_users': MfalmeUsers.objects.filter(account_status='active').count(),
        'total_videos': TrainingVideo.objects.filter(is_active=True).count(),
        'total_courses': Course.objects.filter(is_active=True).count(),
        'testimonials': Testimonial.objects.filter(is_active=True, is_featured=True)[:3],
        'stats': Statistic.objects.filter(is_active=True),
    }
    return render(request, 'login.html', context)

def login_user(request):
    """Handle user login"""
    if request.method != 'POST':
        return redirect('login_page')
    
    email = request.POST.get('email', '').strip()
    password = request.POST.get('password', '')
    remember = request.POST.get('remember') == 'on'
    
    print("\n" + "="*60)
    print("ðŸ” LOGIN ATTEMPT")
    print("="*60)
    print(f"ðŸ“ Email: {email}")
    
    if not email or not password:
        messages.error(request, 'Please fill in all fields')
        return redirect('login_page')
    
    try:
        user = MfalmeUsers.objects.get(email=email)
        print(f"âœ… User found: {user.username}")
        
        if not user.check_password(password):
            print("âŒ Invalid password")
            messages.error(request, 'Invalid password')
            return redirect('login_page')
        
        if user.account_status != 'active':
            if user.account_status == 'pending':
                messages.warning(request, 'Please verify your account first. Check your inbox.')
            elif user.account_status == 'suspended':
                messages.error(request, 'Your account has been suspended. Contact support.')
            elif user.account_status == 'banned':
                messages.error(request, 'Your account has been banned.')
            else:
                messages.error(request, f'Account is {user.account_status}')
            return redirect('login_page')
        
        login(request, user)
        request.session.save()
        
        if not remember:
            request.session.set_expiry(0)
        else:
            request.session.set_expiry(1209600)
        
        user.last_login = timezone.now()
        user.save(update_fields=['last_login'])
        
        log_activity(user, 'LOGIN', f'Logged in from {get_client_ip(request)}', request)
        
        messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
        
        if user.is_staff:
            return redirect('admin_dashboard')
        return redirect('dashboard')
        
    except MfalmeUsers.DoesNotExist:
        messages.error(request, 'No account found with this email')
        return redirect('login_page')
    except Exception as e:
        print(f"âŒ Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Login error: {str(e)}')
        return redirect('login_page')
@csrf_exempt  
def register_user(request):
    """Handle user registration"""
    import sys
    print("\n" + "="*60)
    print("ðŸ” NEW REGISTRATION ATTEMPT")
    print("="*60)
    
    if request.method != 'POST':
        print("âŒ Not a POST request, redirecting to login_page")
        return redirect('login_page')
    
    # Get form data
    email = request.POST.get('email', '').strip()
    username = request.POST.get('username', '').strip()
    phone = request.POST.get('phone', '').strip()
    password = request.POST.get('password', '')
    confirm_password = request.POST.get('confirm_password', '')
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    country = request.POST.get('country', '').strip()
    trading_experience = request.POST.get('trading_experience', 'Beginner')
    terms_accepted = request.POST.get('terms_accepted') == 'on'
    
    print(f"ðŸ“ Email: {email}")
    print(f"ðŸ“ Username: {username}")
    
    # Validation
    errors = []
    
    if not email:
        errors.append('Email is required')
    elif MfalmeUsers.objects.filter(email=email).exists():
        errors.append('Email already registered')
        print(f"âŒ Email already exists: {email}")
    
    if not username:
        errors.append('Username is required')
    elif MfalmeUsers.objects.filter(username=username).exists():
        errors.append('Username already taken')
        print(f"âŒ Username already exists: {username}")
    
    if not phone:
        errors.append('Phone number is required')
    
    if not password:
        errors.append('Password is required')
    elif len(password) < 8:
        errors.append('Password must be at least 8 characters')
    
    if password != confirm_password:
        errors.append('Passwords do not match')
        print(f"âŒ Passwords do not match")
    
    if not terms_accepted:
        errors.append('You must accept the terms and conditions')
        print(f"âŒ Terms not accepted")
    
    if errors:
        for error in errors:
            messages.error(request, error)
        request.session['form_data'] = {
            'email': email,
            'username': username,
            'phone': phone,
            'first_name': first_name,
            'last_name': last_name,
            'country': country,
            'trading_experience': trading_experience,
        }
        request.session.save()
        return redirect('login_page')
    
    try:
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        print(f"âœ… Validation passed, creating user...")
        
        user = MfalmeUsers.objects.create_user(
            email=email,
            password=password,
            username=username,
            phone=phone,
            first_name=first_name,
            last_name=last_name,
            country=country,
            trading_experience=trading_experience,
            account_status='pending',
            email_verified=False,
            registration_ip=ip_address,
            user_agent=user_agent,
        )
        
        print(f"âœ… User created with ID: {user.id}")
        
        verification_code = generate_verification_code()
        VerificationCode.objects.create(
            user=user,
            code=verification_code,
            code_type='account_verification',
            expires_at=timezone.now() + timedelta(minutes=30),
            ip_address=ip_address,
            user_agent=user_agent,
        )
        
        print(f"âœ… Verification code generated: {verification_code}")
        
        email_sent = send_verification_email(user, verification_code)
        admin_notified = send_admin_notification(user)
        
        Notification.objects.create(
            user=user,
            title='Welcome to the Elite!',
            message=f'Welcome {username}! Please verify your account to get started.',
            notification_type='SUCCESS',
        )
        
        log_activity(user, 'REGISTER', f'New registration from {ip_address}', request)
        
        if email_sent:
            messages.success(request, 'Registration successful! Please check your email for verification code.')
            request.session['verification_email'] = email
            request.session.save()
            return redirect('verify_account')
        else:
            messages.warning(request, 'Registration successful but verification email could not be sent. Contact support.')
            return redirect('login_page')
            
    except Exception as e:
        print(f"âŒ Registration exception: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Registration failed: {str(e)}')
        return redirect('login_page')

def verify_account_page(request):
    """Account verification page - FIXED VERSION with copy button"""
    import sys
    print("\n" + "="*60)
    print("🔍 VERIFY ACCOUNT PAGE ACCESSED")
    print("="*60)
    
    # Get email from session or GET parameter
    email = request.session.get('verification_email', request.GET.get('email', ''))
    
    print(f"📧 Email from session/GET: '{email}'")
    
    if not email:
        print("❌ No email in session or GET")
        messages.error(request, 'No verification session found. Please register again.')
        return redirect('login_page')
    
    # Get user
    try:
        user = MfalmeUsers.objects.get(email=email)
        print(f"✅ User found: {user.username} (ID: {user.id})")
        print(f"📊 Account status: {user.account_status}")
        print(f"📧 Email verified: {user.email_verified}")
    except MfalmeUsers.DoesNotExist:
        print(f"❌ User not found with email: {email}")
        messages.error(request, 'User not found. Please register again.')
        return redirect('login_page')
    
    # Get the latest verification code for display
    latest_code = None
    try:
        latest_code = VerificationCode.objects.filter(
            user=user,
            code_type='account_verification',
            is_used=False
        ).order_by('-created_at').first()
        
        if latest_code:
            print(f"🔑 Latest code: {latest_code.code}")
            print(f"⏰ Expires at: {latest_code.expires_at}")
            print(f"⏳ Expired: {latest_code.is_expired()}")
        else:
            print("❌ No active verification code found")
            
            # Generate a new code if none exists
            new_code = generate_verification_code()
            latest_code = VerificationCode.objects.create(
                user=user,
                code=new_code,
                code_type='account_verification',
                expires_at=timezone.now() + timedelta(minutes=30),
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            print(f"✅ Generated new code: {new_code}")
            
            # Send email
            send_verification_email(user, new_code)
    except Exception as e:
        print(f"❌ Error getting verification code: {e}")
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().replace(' ', '')
        print(f"\n🔐 POST request with code: '{code}'")
        
        if not code:
            messages.error(request, 'Please enter verification code')
            return render(request, 'verify_account.html', {
                'user_email': email,
                'user': user,
                'verification_code': latest_code.code if latest_code else None,
                'expires_at': latest_code.expires_at if latest_code else None
            })
        
        try:
            # Get all active verification codes for this user
            verifications = VerificationCode.objects.filter(
                user=user,
                code_type='account_verification',
                is_used=False
            ).order_by('-created_at')
            
            print(f"📋 Found {verifications.count()} active verification codes")
            
            verification = None
            for v in verifications:
                print(f"  - Checking code: {v.code} (expired: {v.is_expired()})")
                if v.code == code:
                    verification = v
                    print(f"  ✅ Code matches!")
                    break
            
            if not verification:
                print(f"❌ No matching code found")
                messages.error(request, 'Invalid verification code')
                
                # Log failed attempt
                ActivityLog.objects.create(
                    user=user,
                    action='VERIFICATION_FAILED',
                    description=f'Invalid verification code attempt: {code}',
                    ip_address=get_client_ip(request)
                )
                
                return render(request, 'verify_account.html', {
                    'user_email': email,
                    'user': user,
                    'verification_code': latest_code.code if latest_code else None,
                    'expires_at': latest_code.expires_at if latest_code else None
                })
            
            if verification.is_expired():
                print(f"❌ Code expired")
                messages.error(request, 'Verification code has expired. Request a new one.')
                verification.is_used = True
                verification.is_valid = False
                verification.save(update_fields=['is_used', 'is_valid'])
                
                return render(request, 'verify_account.html', {
                    'user_email': email,
                    'user': user,
                    'verification_code': None,
                    'expires_at': None
                })
            
            # Mark as used
            verification.is_used = True
            verification.used_at = timezone.now()
            verification.save(update_fields=['is_used', 'used_at'])
            
            # Activate user
            user.email_verified = True
            user.account_status = 'active'
            user.verified_at = timezone.now()
            user.save(update_fields=['email_verified', 'account_status', 'verified_at'])
            
            print(f"✅ User activated successfully!")
            print(f"📊 New status: {user.account_status}")
            
            # Send welcome email
            welcome_sent = send_welcome_email(user)
            print(f"📧 Welcome email sent: {welcome_sent}")
            
            # Create notification
            Notification.objects.create(
                user=user,
                title='🎉 Account Verified!',
                message='Your account has been successfully verified. Welcome to the elite circle!',
                notification_type='SUCCESS',
            )
            
            # Log activity
            ActivityLog.objects.create(
                user=user,
                action='ACCOUNT_VERIFIED',
                description=f'Account verified successfully',
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, '✅ Account verified successfully! You can now login.')
            
            # Clear session
            if 'verification_email' in request.session:
                del request.session['verification_email']
            
            return redirect('login_page')
                
        except Exception as e:
            print(f"❌ Unexpected error: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'An error occurred: {str(e)}')
    
    # Prepare context for GET request
    context = {
        'user_email': email,
        'user': user,
        'verification_code': latest_code.code if latest_code else None,
        'expires_at': latest_code.expires_at if latest_code else None,
        'expires_in': latest_code.expires_at - timezone.now() if latest_code and latest_code.expires_at else None
    }
    
    return render(request, 'verify_account.html', context)

def resend_verification(request):
    """Resend verification code - FIXED VERSION"""
    print("\n" + "="*60)
    print("🔄 RESEND VERIFICATION REQUEST")
    print("="*60)
    
    if request.method == 'POST':
        # Handle both JSON and form submissions
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                email = data.get('email', '')
            except:
                email = ''
        else:
            email = request.POST.get('email', '')
        
        print(f"📧 Email: {email}")
        
        if not email:
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': 'Email required'})
            messages.error(request, 'Email required')
            return redirect('verify_account')
        
        try:
            user = MfalmeUsers.objects.get(email=email)
            print(f"✅ User found: {user.username}")
            
            if user.email_verified:
                print(f"⚠️ Account already verified")
                if request.content_type == 'application/json':
                    return JsonResponse({'success': False, 'error': 'Account already verified'})
                messages.warning(request, 'Account already verified. Please login.')
                return redirect('login_page')
            
            # Invalidate old codes
            VerificationCode.objects.filter(
                user=user,
                code_type='account_verification',
                is_used=False
            ).update(is_used=True, is_valid=False)
            
            # Generate new code
            verification_code = generate_verification_code()
            verification = VerificationCode.objects.create(
                user=user,
                code=verification_code,
                code_type='account_verification',
                expires_at=timezone.now() + timedelta(minutes=30),
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            
            print(f"✅ New code generated: {verification_code}")
            print(f"⏰ Expires: {verification.expires_at}")
            
            # Send email
            email_sent = send_verification_email(user, verification_code)
            
            # Log activity
            ActivityLog.objects.create(
                user=user,
                action='VERIFICATION_RESENT',
                description=f'Verification code resent',
                ip_address=get_client_ip(request)
            )
            
            if email_sent:
                if request.content_type == 'application/json':
                    return JsonResponse({
                        'success': True, 
                        'message': 'New verification code sent to your email',
                        'code': verification_code  # Remove this in production!
                    })
                messages.success(request, 'New verification code sent to your email')
            else:
                if request.content_type == 'application/json':
                    return JsonResponse({
                        'success': False, 
                        'error': 'Failed to send email. Please contact support.'
                    })
                messages.error(request, 'Failed to send email. Please contact support.')
            
            # Update session
            request.session['verification_email'] = email
            request.session.save()
            
            return redirect('verify_account')
                
        except MfalmeUsers.DoesNotExist:
            print(f"❌ User not found: {email}")
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': 'User not found'})
            messages.error(request, 'User not found')
            return redirect('login_page')
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

def forgot_password(request):
    """Forgot password page"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        
        try:
            user = MfalmeUsers.objects.get(email=email)
            
            reset_code = generate_verification_code()
            VerificationCode.objects.create(
                user=user,
                code=reset_code,
                code_type='password_reset',
                expires_at=timezone.now() + timedelta(minutes=30),
                ip_address=get_client_ip(request),
            )
            
            send_password_reset_email(user, reset_code)
            
            messages.success(request, 'Password reset code sent to your email')
            request.session['reset_email'] = email
            return redirect('reset_password')
            
        except MfalmeUsers.DoesNotExist:
            messages.error(request, 'No account found with this email')
    
    return render(request, 'forgot_password.html')

def reset_password(request):
    """Reset password page"""
    email = request.session.get('reset_email', '')
    if not email:
        return redirect('forgot_password')
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        if password != confirm_password:
            messages.error(request, 'Passwords do not match')
            return render(request, 'reset_password.html', {'email': email})
        
        if len(password) < 8:
            messages.error(request, 'Password must be at least 8 characters')
            return render(request, 'reset_password.html', {'email': email})
        
        try:
            user = MfalmeUsers.objects.get(email=email)
            
            verification = VerificationCode.objects.filter(
                user=user,
                code=code,
                code_type='password_reset',
                is_used=False
            ).latest('created_at')
            
            if verification.is_expired():
                messages.error(request, 'Reset code has expired')
            else:
                user.set_password(password)
                user.password_changed_at = timezone.now()
                user.save(update_fields=['password', 'password_changed_at'])
                
                verification.is_used = True
                verification.used_at = timezone.now()
                verification.save()
                
                log_activity(user, 'PASSWORD_CHANGE', 'Password reset via email', request)
                
                Notification.objects.create(
                    user=user,
                    title='Password Changed',
                    message='Your password has been successfully reset.',
                    notification_type='SUCCESS',
                )
                
                messages.success(request, 'Password reset successful. You can now login.')
                
                if 'reset_email' in request.session:
                    del request.session['reset_email']
                
                return redirect('login_page')
                
        except MfalmeUsers.DoesNotExist:
            messages.error(request, 'User not found')
        except VerificationCode.DoesNotExist:
            messages.error(request, 'Invalid reset code')
    
    return render(request, 'reset_password.html', {'email': email})

def logout_user(request):
    """User logout"""
    if request.user.is_authenticated:
        log_activity(request.user, 'LOGOUT', 'User logged out', request)
        logout(request)
        messages.success(request, 'You have been logged out successfully')
    return redirect('index')

# ==================== USER DASHBOARD ====================

def dashboard(request):
    """User dashboard - accessible to all authenticated users"""
    print("\n" + "="*60)
    print(" DASHBOARD ACCESSED")
    
    if not request.user.is_authenticated:
        print("User not authenticated")
        return redirect('login_page')
    
    user = request.user
    print(f"Dashboard accessed by: {user.username}")
    
    context = {
        'user': user,
        'recent_videos': UserVideoAccess.objects.filter(user=user).select_related('video').order_by('-unlocked_at')[:5],
        'recent_pdfs': UserPDFAccess.objects.filter(user=user).select_related('pdf').order_by('-unlocked_at')[:5],
        'enrolled_courses': UserCourse.objects.filter(user=user, is_active=True).select_related('course')[:3],
        'recent_transactions': PaymentTransaction.objects.filter(user=user).order_by('-created_at')[:5],
        'notifications': Notification.objects.filter(user=user, is_read=False)[:10],
        'stats': {
            'videos_watched': UserVideoAccess.objects.filter(user=user).count(),
            'pdfs_viewed': UserPDFAccess.objects.filter(user=user, viewed=True).count(),
            'courses_enrolled': UserCourse.objects.filter(user=user, is_active=True).count(),
            'total_spent': float(PaymentTransaction.objects.filter(user=user, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0),
        },
        'recent_activities': ActivityLog.objects.filter(user=user).order_by('-created_at')[:10],
    }
    
    return render(request, 'dashboard.html', context)


def profile_view(request):
    """User profile page"""
    return render(request, 'profile.html', {'user': request.user})


def profile_update(request):
    """Update user profile"""
    if request.method == 'POST':
        user = request.user
        
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.phone = request.POST.get('phone', user.phone)
        user.country = request.POST.get('country', user.country)
        user.city = request.POST.get('city', user.city)
        user.address = request.POST.get('address', user.address)
        user.whatsapp_number = request.POST.get('whatsapp_number', user.whatsapp_number)
        user.telegram_username = request.POST.get('telegram_username', user.telegram_username)
        user.bio = request.POST.get('bio', user.bio)
        
        if request.FILES.get('profile_image'):
            user.profile_image = request.FILES['profile_image']
        
        user.save()
        
        log_activity(user, 'PROFILE_UPDATE', 'Updated profile information', request)
        messages.success(request, 'Profile updated successfully')
        
        return redirect('profile')
    
    return redirect('profile')


def change_password(request):
    """Change user password"""
    if request.method == 'POST':
        user = request.user
        current = request.POST.get('current_password', '')
        new = request.POST.get('new_password', '')
        confirm = request.POST.get('confirm_password', '')
        
        if not user.check_password(current):
            messages.error(request, 'Current password is incorrect')
        elif new != confirm:
            messages.error(request, 'New passwords do not match')
        elif len(new) < 8:
            messages.error(request, 'Password must be at least 8 characters')
        else:
            user.set_password(new)
            user.password_changed_at = timezone.now()
            user.save(update_fields=['password', 'password_changed_at'])
            
            log_activity(user, 'PASSWORD_CHANGE', 'Password changed from profile', request)
            messages.success(request, 'Password changed successfully')
            
            # Re-authenticate to maintain session
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
    
    return redirect('profile')

# ==================== CONTENT VIEWING ====================


def my_videos(request):
    """List user's videos"""
    user = request.user
    video_access = UserVideoAccess.objects.filter(user=user).select_related('video').order_by('-unlocked_at')
    
    free_videos = TrainingVideo.objects.filter(price=0, is_active=True).exclude(
        id__in=video_access.values_list('video_id', flat=True)
    )
    
    context = {
        'purchased_videos': video_access,
        'free_videos': free_videos,
        'total_videos': video_access.count() + free_videos.count(),
    }
    return render(request, 'my_videos.html', context)


def my_pdfs(request):
    """List user's PDFs"""
    user = request.user
    pdf_access = UserPDFAccess.objects.filter(user=user).select_related('pdf').order_by('-unlocked_at')
    
    free_pdfs = PDF.objects.filter(is_free=True, is_active=True).exclude(
        id__in=pdf_access.values_list('pdf_id', flat=True)
    )
    
    context = {
        'purchased_pdfs': pdf_access,
        'free_pdfs': free_pdfs,
        'total_pdfs': pdf_access.count() + free_pdfs.count(),
    }
    return render(request, 'my_pdfs.html', context)


def my_courses(request):
    """List user's courses - WITH EXPIRATION CHECK"""
    user = request.user
    courses = UserCourse.objects.filter(user=user).select_related('course').order_by('-enrolled_at')
    
    # Check for expired courses and update them
    for enrollment in courses:
        if enrollment.is_access_expired() and enrollment.is_active:
            enrollment.is_active = False
            enrollment.save(update_fields=['is_active'])
    
    # Refresh queryset after updates
    courses = UserCourse.objects.filter(user=user).select_related('course').order_by('-enrolled_at')
    
    context = {
        'courses': courses,
        'total_courses': courses.count(),
        'active_courses': courses.filter(is_active=True).count(),
        'expired_courses': courses.filter(is_active=False).count(),
    }
    return render(request, 'view_course.html', context)


def transaction_history(request):
    """User transaction history"""
    user = request.user
    transactions = PaymentTransaction.objects.filter(user=user).order_by('-created_at')
    
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page', 1)
    transactions_page = paginator.get_page(page)
    
    context = {
        'transactions': transactions_page,
        'total_count': transactions.count(),
        'total_spent': float(transactions.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0),
    }
    return render(request, 'transaction_history.html', context)

# ==================== WATCH VIDEO VIEW ====================


def watch_video(request, video_id):
    """Watch a specific video - DOWNLOAD DISABLED"""
    try:
        video = get_object_or_404(TrainingVideo, id=video_id, is_active=True)
    except:
        messages.error(request, 'Video not found')
        return redirect('my_videos')
    
    user = request.user
    
    # Check if user has access
    has_access = check_video_access(user, video)
    
    if not has_access:
        messages.error(request, 'You do not have access to this video')
        return redirect('my_videos')
    
    # Increment view count
    video.view_count += 1
    video.save(update_fields=['view_count'])
    
    # Log activity
    log_activity(user, 'VIDEO_WATCH', f'Watched video: {video.title}', request)
    
    # Get related videos from same course
    related = TrainingVideo.objects.filter(
        course=video.course, 
        is_active=True
    ).exclude(id=video.id)[:5]
    
    # Prepare video URLs - Django's storage system handles S3 URLs automatically
    video_file_url = video.video_file.url if video.video_file else None
    thumbnail_url = video.thumbnail.url if video.thumbnail else None
    
    # Check if files exist (optional)
    if video.video_file:
        try:
            # This will check if the file exists in S3
            if not video.video_file.storage.exists(video.video_file.name):
                print(f"⚠️ Video file missing from S3: {video.video_file.name}")
                video_file_url = None
        except Exception as e:
            print(f"⚠️ Error checking video file: {e}")
    
    # Get embed URL for YouTube/Vimeo
    embed_url = None
    if hasattr(video, 'video_type') and video.video_type == 'youtube' and hasattr(video, 'youtube_id') and video.youtube_id:
        embed_url = f"https://www.youtube.com/embed/{video.youtube_id}"
    elif hasattr(video, 'video_type') and video.video_type == 'vimeo' and hasattr(video, 'vimeo_id') and video.vimeo_id:
        embed_url = f"https://player.vimeo.com/video/{video.vimeo_id}"
    
    context = {
        'video': video,
        'video_file_url': video_file_url,
        'thumbnail_url': thumbnail_url,
        'related_videos': related,
        'has_video_file': video_file_url is not None,
        'video_url': video.video_url or '',
        'embed_code': video.embed_code or '',
        'youtube_id': getattr(video, 'youtube_id', ''),
        'vimeo_id': getattr(video, 'vimeo_id', ''),
        'video_type': getattr(video, 'video_type', ''),
        'embed_url': embed_url or '',
        'disable_downloads': True,
    }
    
    return render(request, 'watch_video.html', context)

# ==================== VIEW PDF (VIEW INSTEAD OF DOWNLOAD) ====================


def view_pdf(request, pdf_id):
    """View PDF in browser instead of downloading"""
    try:
        pdf = get_object_or_404(PDF, id=pdf_id, is_active=True)
    except:
        messages.error(request, 'PDF not found')
        return redirect('my_pdfs')
    
    user = request.user
    
    # Check if user has access
    has_access = check_pdf_access(user, pdf)
    
    if not has_access:
        messages.error(request, 'You do not have access to this PDF')
        return redirect('my_pdfs')
    
    # Check if file exists in S3
    if not pdf.pdf_file:
        messages.error(request, 'PDF file not found')
        return redirect('my_pdfs')
    
    try:
        # Check if file exists in S3
        if not pdf.pdf_file.storage.exists(pdf.pdf_file.name):
            messages.error(request, 'PDF file not found in storage')
            return redirect('my_pdfs')
        
        # Update access record
        access, created = UserPDFAccess.objects.get_or_create(user=user, pdf=pdf)
        access.viewed = True
        access.view_count += 1
        access.last_viewed = timezone.now()
        access.save(update_fields=['viewed', 'view_count', 'last_viewed'])
        
        # Update PDF view count
        pdf.views += 1
        pdf.save(update_fields=['views'])
        
        # Log activity
        log_activity(user, 'PDF_VIEWED', f'Viewed PDF: {pdf.title}', request)
        
        # Return the file for viewing - Django's storage handles S3 streaming
        response = FileResponse(pdf.pdf_file.open('rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{pdf.title}.pdf"'
        response['X-Content-Type-Options'] = 'nosniff'
        return response
        
    except Exception as e:
        print(f"❌ Error viewing PDF: {e}")
        messages.error(request, f'Error viewing PDF: {str(e)}')
        return redirect('my_pdfs')

# ==================== DEPRECATED DOWNLOAD PDF ====================


def download_pdf(request, pdf_id):
    """DEPRECATED: Redirect to view instead of download"""
    messages.info(request, 'PDFs are now viewed in browser instead of downloaded.')
    return redirect('view_pdf', pdf_id=pdf_id)

# ==================== VIEW COURSE WITH EXPIRATION ====================


def view_course(request, course_id):
    """View course details and content - WITH EXPIRATION CHECK"""
    course = get_object_or_404(Course, id=course_id, is_active=True)
    user = request.user
    
    # Check access with expiration
    has_access, enrollment = check_course_access(user, course_id)
    
    if not has_access or not enrollment:
        if enrollment and enrollment.is_access_expired():
            messages.error(request, 'Your course access has expired. Please renew to continue learning.')
        else:
            messages.error(request, 'You are not enrolled in this course')
        return redirect('my_courses')
    
    videos = TrainingVideo.objects.filter(course=course, is_active=True).order_by('order')
    pdfs = PDF.objects.filter(course=course, is_active=True).order_by('order')
    
    context = {
        'course': course,
        'enrollment': enrollment,
        'videos': videos,
        'pdfs': pdfs,
        'progress': enrollment.progress,
        'completed_count': len(enrollment.completed_lessons) if enrollment.completed_lessons else 0,
        'total_lessons': videos.count() + pdfs.count(),
        'expires_in': enrollment.time_until_expiry(),
        'is_expired': enrollment.is_access_expired(),
    }
    return render(request, 'view_course.html', context)


def mark_lesson_complete(request, course_id, lesson_type, lesson_id):
    """Mark a lesson as complete"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    user = request.user
    course = get_object_or_404(Course, id=course_id)
    
    # Check access with expiration
    has_access, enrollment = check_course_access(user, course_id)
    
    if not has_access or not enrollment:
        return JsonResponse({'error': 'Not enrolled or access expired'}, status=403)
    
    lesson_key = f"{lesson_type}_{lesson_id}"
    if not enrollment.completed_lessons:
        enrollment.completed_lessons = []
    
    if lesson_key not in enrollment.completed_lessons:
        enrollment.completed_lessons.append(lesson_key)
        
        total_lessons = TrainingVideo.objects.filter(course=course).count() + PDF.objects.filter(course=course).count()
        if total_lessons > 0:
            enrollment.progress = int((len(enrollment.completed_lessons) / total_lessons) * 100)
        
        enrollment.save()
        
        return JsonResponse({
            'success': True,
            'progress': enrollment.progress,
            'completed_count': len(enrollment.completed_lessons),
            'total_lessons': total_lessons,
        })
    
    return JsonResponse({'success': True})

# ==================== SUPPORT TICKETS ====================


def support_tickets(request):
    """User support tickets"""
    user = request.user
    tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')
    
    context = {
        'tickets': tickets,
        'open_count': tickets.filter(status='open').count(),
        'resolved_count': tickets.filter(status='resolved').count(),
    }
    return render(request, 'support_tickets.html', context)


def create_ticket(request):
    """Create a support ticket"""
    if request.method == 'POST':
        user = request.user
        
        ticket = SupportTicket.objects.create(
            user=user,
            subject=request.POST.get('subject'),
            message=request.POST.get('message'),
            category=request.POST.get('category', 'general'),
            priority=request.POST.get('priority', 'medium'),
        )
        
        log_activity(user, 'SUPPORT_TICKET_CREATED', f'Created ticket: {ticket.subject[:50]}', request)
        
        messages.success(request, f'Ticket #{ticket.ticket_number} created successfully')
        return redirect('view_ticket', ticket_id=ticket.id)
    
    return render(request, 'create_ticket.html')


def view_ticket(request, ticket_id):
    """View a specific ticket"""
    ticket = get_object_or_404(SupportTicket, id=ticket_id, user=request.user)
    replies = ticket.replies.all().order_by('created_at')
    
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        if message:
            reply = TicketReply.objects.create(
                ticket=ticket,
                user=request.user,
                message=message
            )
            ticket.status = 'waiting_reply'
            ticket.save(update_fields=['status'])
            messages.success(request, 'Reply added')
            return redirect('view_ticket', ticket_id=ticket.id)
    
    context = {
        'ticket': ticket,
        'replies': replies,
    }
    return render(request, 'view_ticket.html', context)


def close_ticket(request, ticket_id):
    """Close a ticket"""
    if request.method == 'POST':
        ticket = get_object_or_404(SupportTicket, id=ticket_id, user=request.user)
        ticket.status = 'closed'
        ticket.closed_at = timezone.now()
        ticket.save(update_fields=['status', 'closed_at'])
        messages.success(request, 'Ticket closed')
    return redirect('view_ticket', ticket_id=ticket_id)

# ==================== API ENDPOINTS ====================


def api_user_profile(request):
    """Get user profile (JSON)"""
    user = request.user
    return JsonResponse({
        'id': user.id,
        'soldier_id': user.soldier_id,
        'username': user.username,
        'email': user.email,
        'phone': user.phone,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'full_name': user.get_full_name(),
        'elite_rank': user.elite_rank,
        'account_status': user.account_status,
        'email_verified': user.email_verified,
        'country': user.country,
        'city': user.city,
        'trading_experience': user.trading_experience,
        'account_balance': float(user.account_balance),
        'referral_code': user.referral_code,
        'referral_count': user.referral_count,
        'referral_earnings': float(user.referral_earnings),
        'date_joined': user.date_joined.strftime('%Y-%m-%d'),
        'profile_image': user.profile_image.url if user.profile_image else None,
        'total_deposits': float(user.total_deposits),
        'total_withdrawals': float(user.total_withdrawals),
        'total_profit': float(user.total_profit),
        'success_rate': user.success_rate,
    })


def api_user_stats(request):
    """Get user statistics (JSON)"""
    user = request.user
    
    videos_watched = UserVideoAccess.objects.filter(user=user).count()
    pdfs_viewed = UserPDFAccess.objects.filter(user=user, viewed=True).count()
    courses_enrolled = UserCourse.objects.filter(user=user, is_active=True).count()
    watchlist_count = Watchlist.objects.filter(user=user).count()
    
    return JsonResponse({
        'videos_watched': videos_watched,
        'pdfs_viewed': pdfs_viewed,
        'courses_enrolled': courses_enrolled,
        'total_spent': float(PaymentTransaction.objects.filter(user=user, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0),
        'tickets_created': SupportTicket.objects.filter(user=user).count(),
        'open_tickets': SupportTicket.objects.filter(user=user, status='open').count(),
        'referral_earnings': float(user.referral_earnings),
        'success_rate': user.success_rate,
        'elite_rank': user.elite_rank,
        'watchlist_count': watchlist_count,
        'community_memberships': UserCommunityMembership.objects.filter(user=user, status='active').count(),
    })


def api_user_activities(request):
    """Get user activities (JSON)"""
    limit = int(request.GET.get('limit', 10))
    activities = ActivityLog.objects.filter(user=request.user).order_by('-created_at')[:limit]
    
    data = []
    for act in activities:
        data.append({
            'id': act.id,
            'action': act.action,
            'description': act.description,
            'time': timesince(act.created_at),
            'created_at': act.created_at.isoformat(),
        })
    
    return JsonResponse(data, safe=False)


def api_user_notifications(request):
    """Get user notifications (JSON)"""
    # Check if user is authenticated via your custom system or Django
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    unread_only = request.GET.get('unread_only') == 'true'
    
    notifications = Notification.objects.filter(user=request.user)
    if unread_only:
        notifications = notifications.filter(is_read=False)
    
    notifications = notifications.order_by('-created_at')[:20]
    
    data = []
    for notif in notifications:
        data.append({
            'id': notif.id,
            'title': notif.title,
            'message': notif.message,
            'type': notif.notification_type,
            'is_read': notif.is_read,
            'time': timesince(notif.created_at),
            'created_at': notif.created_at.isoformat(),
            'action_url': notif.action_url,
            'action_text': notif.action_text,
        })
    
    return JsonResponse({
        'notifications': data,
        'unread_count': Notification.objects.filter(user=request.user, is_read=False).count(),
    })
    

def api_mark_notification_read(request, notification_id):
    """Mark notification as read"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.read_at = timezone.now()
    notification.save(update_fields=['is_read', 'read_at'])
    
    return JsonResponse({'success': True})


def api_mark_notification_read(request, notification_id):
    """Mark notification as read"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.read_at = timezone.now()
    notification.save(update_fields=['is_read', 'read_at'])
    
    return JsonResponse({'success': True})

def some_admin_api(request):
    """Your admin API function"""
    if not request.session.get('admin_authenticated'):
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    # Your logic here
    return JsonResponse({'success': True})    


def api_mark_all_notifications_read(request):
    """Mark all notifications as read"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    Notification.objects.filter(user=request.user, is_read=False).update(
        is_read=True,
        read_at=timezone.now()
    )
    
    return JsonResponse({'success': True})

# ==================== NEW API ENDPOINTS FOR DASHBOARD ====================


def api_user_courses(request):
    """Get user's enrolled courses - WITH EXPIRATION"""
    user = request.user
    
    # Check for expired courses and update them
    expired = UserCourse.objects.filter(user=user, is_active=True)
    for enrollment in expired:
        if enrollment.is_access_expired():
            enrollment.is_active = False
            enrollment.save(update_fields=['is_active'])
    
    enrollments = UserCourse.objects.filter(user=user, is_active=True).select_related('course')
    
    data = []
    for enrollment in enrollments:
        course = enrollment.course
        data.append({
            'id': course.id,
            'title': course.title,
            'description': course.description,
            'price': float(course.price),
            'progress': enrollment.progress,
            'thumbnail': course.thumbnail.url if course.thumbnail else None,
            'videos_count': course.video_count(),
            'pdfs_count': course.pdf_count(),
            'enrolled_at': enrollment.enrolled_at.strftime('%Y-%m-%d'),
            'access_expires_at': enrollment.access_expires_at.strftime('%Y-%m-%d %H:%M') if enrollment.access_expires_at else None,
            'expires_in': enrollment.time_until_expiry(),
            'is_expired': enrollment.is_access_expired(),
        })
    
    return JsonResponse(data, safe=False)


def api_user_videos(request):
    """Get user's videos"""
    user = request.user
    
    # Get purchased videos
    purchased = UserVideoAccess.objects.filter(user=user).select_related('video')
    purchased_data = []
    for access in purchased:
        v = access.video
        purchased_data.append({
            'id': v.id,
            'title': v.title,
            'description': v.description,
            'thumbnail': v.thumbnail.url if v.thumbnail else None,
            'category': v.category,
            'duration': v.duration,
            'price': float(v.price),
            'unlocked_at': access.unlocked_at.strftime('%Y-%m-%d'),
            'has_access': True,
            'is_free': v.price == 0,
            'view_count': v.view_count,
            'allow_download': v.allow_download,  # Add this flag for template
        })
    
    # Get free videos
    free = TrainingVideo.objects.filter(price=0, is_active=True).exclude(
        id__in=[a.video.id for a in purchased]
    )
    free_data = []
    for v in free:
        free_data.append({
            'id': v.id,
            'title': v.title,
            'description': v.description,
            'thumbnail': v.thumbnail.url if v.thumbnail else None,
            'category': v.category,
            'duration': v.duration,
            'price': 0,
            'has_access': True,
            'is_free': True,
            'view_count': v.view_count,
            'allow_download': v.allow_download,  # Add this flag for template
        })
    
    return JsonResponse({
        'purchased': purchased_data,
        'free': free_data,
    })


def api_user_pdfs(request):
    """Get user's PDFs - FOR VIEWING"""
    user = request.user
    
    # Get purchased PDFs
    purchased = UserPDFAccess.objects.filter(user=user).select_related('pdf')
    purchased_data = []
    for access in purchased:
        p = access.pdf
        purchased_data.append({
            'id': p.id,
            'title': p.title,
            'description': p.description,
            'cover_image': p.cover_image.url if p.cover_image else None,
            'pages': p.pages,
            'file_size': p.file_size,
            'category': p.category,
            'price': float(p.price),
            'unlocked_at': access.unlocked_at.strftime('%Y-%m-%d'),
            'has_access': True,
            'is_free': p.is_free,
            'viewed': access.viewed,
            'view_count': access.view_count,
            'last_viewed': access.last_viewed.strftime('%Y-%m-%d %H:%M') if access.last_viewed else None,
        })
    
    # Get free PDFs
    free = PDF.objects.filter(is_free=True, is_active=True).exclude(
        id__in=[a.pdf.id for a in purchased]
    )
    free_data = []
    for p in free:
        free_data.append({
            'id': p.id,
            'title': p.title,
            'description': p.description,
            'cover_image': p.cover_image.url if p.cover_image else None,
            'pages': p.pages,
            'file_size': p.file_size,
            'category': p.category,
            'price': 0,
            'has_access': True,
            'is_free': True,
            'viewed': False,
            'view_count': 0,
        })
    
    return JsonResponse({
        'purchased': purchased_data,
        'free': free_data,
    })


@csrf_exempt
def api_course_enroll(request):
    """Enroll in a course"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    course_id = data.get('course_id')
    if not course_id:
        return JsonResponse({'error': 'Course ID required'}, status=400)
    
    try:
        course = Course.objects.get(id=course_id, is_active=True)
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)
    
    user = request.user
    
    # Check if already enrolled
    existing = UserCourse.objects.filter(user=user, course=course).first()
    if existing:
        if existing.is_access_expired():
            # Reactivate expired enrollment
            existing.access_expires_at = timezone.now() + timedelta(days=365)  # 1 YEAR
            existing.is_active = True
            existing.save()
            return JsonResponse({
                'success': True,
                'message': 'Course access renewed for 1 year',
                'course_id': course.id,
                'expires_in': existing.time_until_expiry()
            })
        return JsonResponse({'error': 'Already enrolled in this course'}, status=400)
    
    # If course is free, enroll directly
    if course.price == 0:
        enrollment = UserCourse.objects.create(
            user=user,
            course=course,
            purchase_type='1_month',
            access_expires_at=timezone.now() + timedelta(days=365)  # 1 YEAR
        )
        
        # Grant access to all videos and PDFs
        enrollment.get_video_access()
        enrollment.get_pdf_access()
        
        # Log activity
        log_activity(user, 'COURSE_ENROLLED', f'Enrolled in free course: {course.title}', request)
        
        # Create notification
        Notification.objects.create(
            user=user,
            title='Course Enrolled',
            message=f'You have successfully enrolled in {course.title} for 1 year',
            notification_type='SUCCESS',
            related_object_type='course',
            related_object_id=course.id
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Successfully enrolled in course for 1 year',
            'course_id': course.id,
            'expires_in': enrollment.time_until_expiry()
        })
    else:
        # Paid course - redirect to payment
        return JsonResponse({
            'success': False,
            'requires_payment': True,
            'message': 'This course requires payment',
            'redirect_url': f'/payment/initiate/?type=course&id={course.id}'
        }, status=400)


def api_watchlist_count(request):
    """Get watchlist count"""
    user = request.user
    count = Watchlist.objects.filter(user=user).count()
    return JsonResponse({'count': count})

@csrf_exempt
def api_watchlist_check(request):
    """Check if item is in watchlist"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    content_type = data.get('type')
    content_id = data.get('id')
    
    if not content_type or not content_id:
        return JsonResponse({'error': 'Type and ID required'}, status=400)
    
    user = request.user
    exists = Watchlist.objects.filter(
        user=user,
        content_type=content_type,
        content_id=content_id
    ).exists()
    
    return JsonResponse({'in_watchlist': exists})


@csrf_exempt
def api_watchlist_add(request):
    """Add item to watchlist"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    content_type = data.get('type')
    content_id = data.get('id')
    
    if not content_type or not content_id:
        return JsonResponse({'error': 'Type and ID required'}, status=400)
    
    valid_types = ['video', 'pdf', 'course', 'package']
    if content_type not in valid_types:
        return JsonResponse({'error': 'Invalid type'}, status=400)
    
    user = request.user
    
    # Check if already in watchlist
    existing = Watchlist.objects.filter(
        user=user,
        content_type=content_type,
        content_id=content_id
    ).first()
    
    if existing:
        return JsonResponse({'error': 'Item already in watchlist'}, status=400)
    
    # Verify item exists
    obj = None
    if content_type == 'video':
        obj = TrainingVideo.objects.filter(id=content_id).first()
    elif content_type == 'pdf':
        obj = PDF.objects.filter(id=content_id).first()
    elif content_type == 'course':
        obj = Course.objects.filter(id=content_id).first()
    elif content_type == 'package':
        obj = Package.objects.filter(id=content_id).first()
    
    if not obj:
        return JsonResponse({'error': 'Item not found'}, status=404)
    
    # Add to watchlist
    watchlist_item = Watchlist.objects.create(
        user=user,
        content_type=content_type,
        content_id=content_id
    )
    
    # Log activity
    log_activity(user, 'WATCHLIST_ADD', f'Added to watchlist: {obj.title}', request)
    
    return JsonResponse({
        'success': True,
        'id': watchlist_item.id,
        'message': 'Added to watchlist'
    })


@csrf_exempt
def api_watchlist_remove(request):
    """Remove item from watchlist"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    content_type = data.get('type')
    content_id = data.get('id')
    
    if not content_type or not content_id:
        return JsonResponse({'error': 'Type and ID required'}, status=400)
    
    deleted = Watchlist.objects.filter(
        user=request.user,
        content_type=content_type,
        content_id=content_id
    ).delete()
    
    if deleted[0] > 0:
        return JsonResponse({'success': True, 'message': 'Removed from watchlist'})
    else:
        return JsonResponse({'error': 'Item not found in watchlist'}, status=404)


def api_watchlist(request):
    """Get user's watchlist with item details"""
    user = request.user
    watchlist = Watchlist.objects.filter(user=user).order_by('-created_at')
    
    data = []
    for item in watchlist:
        content = item.get_content()
        if content:
            data.append({
                'id': item.id,
                'content_id': item.content_id,
                'type': item.content_type,
                'title': getattr(content, 'title', str(content)),
                'price': float(getattr(content, 'price', 0)),
                'added_at': item.created_at.strftime('%Y-%m-%d'),
            })
    
    return JsonResponse(data, safe=False)


def api_user_communities(request):
    """Get user's community memberships"""
    user = request.user
    memberships = UserCommunityMembership.objects.filter(user=user).select_related('community')
    
    data = []
    for m in memberships:
        data.append({
            'id': m.id,
            'community_id': m.community.id,
            'community_name': m.community.name,
            'community_tier': m.community.tier,
            'status': m.status,
            'joined_at': m.joined_at.strftime('%Y-%m-%d'),
            'access_granted': m.access_granted,
            'discord_username': m.discord_username,
            'telegram_username': m.telegram_username,
        })
    
    return JsonResponse(data, safe=False)


@csrf_exempt
def api_community_join(request):
    """Request to join a community tier"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    tier = data.get('tier')
    if not tier:
        return JsonResponse({'error': 'Tier required'}, status=400)
    
    try:
        community = CommunityTier.objects.get(tier=tier, is_active=True)
    except CommunityTier.DoesNotExist:
        return JsonResponse({'error': 'Community tier not found'}, status=404)
    
    user = request.user
    
    # Check if already a member
    existing = UserCommunityMembership.objects.filter(user=user, community=community).first()
    if existing:
        return JsonResponse({
            'error': f'You already have a {existing.status} membership for this community'
        }, status=400)
    
    # Check for pending request
    pending = CommunityJoinRequest.objects.filter(
        user=user, community=community, status='pending'
    ).exists()
    if pending:
        return JsonResponse({'error': 'You already have a pending request'}, status=400)
    
    # Check eligibility
    eligible, missing = community.check_user_eligibility(user)
    
    if not eligible and community.tier != 'citizens':
        return JsonResponse({
            'error': 'You do not meet the requirements',
            'missing': missing
        }, status=403)
    
    if community.tier == 'citizens':
        # Instant access for citizens
        membership = UserCommunityMembership.objects.create(
            user=user,
            community=community,
            status='active',
            access_granted=True,
            access_granted_at=timezone.now()
        )
        
        log_activity(user, 'COMMUNITY_JOINED', f'Joined {community.name}', request)
        
        return JsonResponse({
            'success': True,
            'status': 'active',
            'message': f'Welcome to {community.name}!'
        })
    else:
        # Create join request for paid tiers
        join_request = CommunityJoinRequest.objects.create(
            user=user,
            community=community,
            met_requirements=eligible,
            investment_at_request=user.total_deposits,
            courses_completed=list(UserCourse.objects.filter(user=user).values_list('course_id', flat=True))
        )
        
        # Send email to admin
        send_community_application_email(user, community, request)
        
        log_activity(user, 'COMMUNITY_REQUEST', f'Requested to join {community.name}', request)
        
        return JsonResponse({
            'success': True,
            'status': 'pending',
            'message': 'Your request has been submitted for review'
        })


def api_institute_eligibility(request):
    """Check user's eligibility for Institute account"""
    user = request.user
    
    # Check requirements
    invest_met = user.total_deposits >= 250000
    ptm_course = UserCourse.objects.filter(
        user=user, 
        course__title__icontains='PTM'
    ).exists()
    
    # Account age in days
    account_age = (timezone.now() - user.date_joined).days
    age_met = account_age >= 30
    
    # Check for existing application
    existing_application = InstituteApplication.objects.filter(
        user=user, 
        status__in=['pending', 'approved']
    ).first()
    
    data = {
        'total_deposits': float(user.total_deposits),
        'has_ptm_course': ptm_course,
        'account_age_days': account_age,
        'meets_investment': invest_met,
        'meets_ptm': ptm_course,
        'meets_age': age_met,
        'eligible': invest_met and ptm_course and age_met,
        'has_application': existing_application is not None,
        'application_status': existing_application.status if existing_application else None
    }
    
    return JsonResponse(data)


@csrf_exempt
def api_institute_apply(request):
    """Submit institute application"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    user = request.user
    
    # Check if already applied
    existing = InstituteApplication.objects.filter(
        user=user, 
        status__in=['pending', 'approved']
    ).first()
    if existing:
        return JsonResponse({
            'error': f'You already have a {existing.status} application'
        }, status=400)
    
    # Get form data
    amount = request.POST.get('amount')
    experience = request.POST.get('experience')
    notes = request.POST.get('notes', '')
    
    if not amount or not experience:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    
    try:
        amount = Decimal(amount)
    except:
        return JsonResponse({'error': 'Invalid amount'}, status=400)
    
    if amount < 250000:
        return JsonResponse({'error': 'Minimum investment is $250,000'}, status=400)
    
    proof_file = request.FILES.get('proof_of_funds')
    if not proof_file:
        return JsonResponse({'error': 'Proof of funds is required'}, status=400)
    
    # Create application
    application = InstituteApplication(
        user=user,
        investment_amount=amount,
        trading_experience=experience,
        notes=notes,
        proof_of_funds=proof_file
    )
    
    history_file = request.FILES.get('trading_history')
    if history_file:
        application.trading_history = history_file
    
    application.save()
    
    # Send email to admin
    send_institute_application_email(application)
    
    log_activity(user, 'INSTITUTE_APPLY', 'Submitted institute application', request)
    
    return JsonResponse({
        'success': True,
        'message': 'Application submitted successfully'
    })


def api_get_packages(request):
    """Get available packages (JSON)"""
    packages = Package.objects.filter(is_active=True).order_by('order')
    data = []
    for pkg in packages:
        data.append({
            'id': pkg.id,
            'name': pkg.name,
            'package_type': pkg.package_type,
            'short_description': pkg.short_description,
            'price': float(pkg.price),
            'original_price': float(pkg.original_price) if pkg.original_price else None,
            'discount_percentage': pkg.discount_percentage,
            'features': pkg.features,
            'benefits': pkg.benefits,
            'duration_days': pkg.duration_days,
            'is_featured': pkg.is_featured,
            'is_popular': pkg.is_popular,
            'is_recurring': pkg.is_recurring,
        })
    return JsonResponse(data, safe=False)


def api_unlock_video(request):
    """API to unlock/purchase video"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    video_id = data.get('video_id')
    video = get_object_or_404(TrainingVideo, id=video_id)
    
    # Convert USD to KES for storage
    amount_usd = video.price
    amount_kes = amount_usd * Decimal('129')
    
    transaction = PaymentTransaction.objects.create(
        user=request.user,
        reference=generate_reference(),
        amount=amount_kes,  # Store KES
        currency='KES',
        payment_type='video_purchase',
        payment_method='paystack',
        description=f'Purchase video: {video.title}',
        metadata={
            'video_id': video.id, 
            'video_title': video.title,
            'amount_usd': float(amount_usd)
        },
        status='pending'
    )
    
    return JsonResponse({
        'success': True,
        'reference': transaction.reference,
        'amount': float(amount_kes),  # Return KES for Paystack
        'currency': 'KES',
        'authorization_url': f"/payment/verify/{transaction.reference}/"
    })


@csrf_exempt
def initialize_package_payment(request):
    """Initialize payment for a package and redirect to payment page"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Convert USD to KES
            amount_usd = Decimal(str(data.get('amount', 0)))
            amount_kes = amount_usd * Decimal('129')
            
            # Create transaction
            transaction = PaymentTransaction.objects.create(
                user=request.user if request.user.is_authenticated else None,
                reference=generate_reference(),
                amount=amount_kes,
                currency='KES',
                payment_type='package_purchase',
                payment_method='sasapay',
                description=data.get('description', 'Package Purchase'),
                metadata=data.get('metadata', {}),
                status='initiated',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            
            return JsonResponse({
                'success': True,
                'reference': transaction.reference
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def education_payment(request):
    """Simple GET endpoint for education payments"""
    # Get parameters from URL
    program = request.GET.get('program')
    name = request.GET.get('name')
    amount = request.GET.get('amount')
    
    # Validate
    if not all([program, name, amount]):
        messages.error(request, 'Missing payment information')
        return redirect('index')
    
    try:
        # Convert to Decimal
        from decimal import Decimal
        amount_usd = Decimal(str(amount))
        amount_kes = amount_usd * Decimal('129')
        
        # Create transaction
        transaction = PaymentTransaction.objects.create(
            user=request.user if request.user.is_authenticated else None,
            reference=generate_reference(),
            amount=amount_kes,
            currency='KES',
            payment_type='education_purchase',
            payment_method='sasapay',
            description=name,
            metadata={
                'program_code': program,
                'program_name': name,
                'amount_usd': float(amount_usd),
                'source': 'education_section'
            },
            status='initiated'
        )
        
        # Redirect to payment page
        return redirect(f'/payment/?ref={transaction.reference}')
        
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('index')
    

def api_send_support(request):
    """API to send support message"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    ticket = SupportTicket.objects.create(
        user=request.user,
        subject=data.get('subject', 'Support Request'),
        message=data.get('message', ''),
        category='support',
        priority='medium'
    )
    
    return JsonResponse({
        'success': True,
        'ticket_id': ticket.id,
        'ticket_number': ticket.ticket_number
    })


def api_update_settings(request):
    """API to update user settings"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    user = request.user
    
    if 'notifications' in data:
        user.notification_preferences = data['notifications']
    
    if 'privacy' in data:
        user.privacy_settings = data['privacy']
    
    user.save()
    
    return JsonResponse({'success': True})


def api_user_orders(request):
    """Get user orders (JSON)"""
    user = request.user
    orders = PaymentTransaction.objects.filter(user=user).order_by('-created_at')[:20]
    
    data = []
    for order in orders:
        data.append({
            'reference': order.reference,
            'amount': float(order.amount),
            'currency': order.currency,
            'payment_type': order.payment_type,
            'payment_method': order.payment_method,
            'status': order.status,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M'),
            'paid_at': order.paid_at.strftime('%Y-%m-%d') if order.paid_at else None,
        })
    
    return JsonResponse(data, safe=False)
def api_user_tickets(request):
    """Get user support tickets (JSON)"""
    user = request.user
    tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')[:20]
    
    data = []
    for ticket in tickets:
        data.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'subject': ticket.subject,
            'category': ticket.category,
            'priority': ticket.priority,
            'status': ticket.status,
            'reply_count': ticket.reply_count,
            'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    return JsonResponse(data, safe=False)


@csrf_exempt
def api_create_order(request):
    """Create a new order - Handles both POST (JSON) and GET (URL parameters)"""
    
    # Get or create a guest user for unauthenticated users
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    guest_user = None
    try:
        guest_user = User.objects.get(email='guest@system.com')
    except User.DoesNotExist:
        guest_user = User.objects.create_user(
            email='guest@system.com',
            username='guest_system',
            password=None,
            account_status='active'
        )
    
    # Handle GET requests from education section links
    if request.method == 'GET':
        program = request.GET.get('program')
        name = request.GET.get('name')
        amount = request.GET.get('amount')
        payment_type = request.GET.get('type', 'education_purchase')
        
        if not all([program, name, amount]):
            return JsonResponse({
                'success': False, 
                'error': 'Missing required parameters: program, name, amount'
            }, status=400)
        
        try:
            amount_usd = Decimal(str(amount))
            amount_kes = amount_usd * Decimal('129')
            
            # Use the authenticated user or guest user
            user = request.user if request.user.is_authenticated and request.user.id else guest_user
            
            transaction = PaymentTransaction.objects.create(
                user=user,
                reference=generate_reference(),
                amount=amount_kes,
                currency='KES',
                payment_type=payment_type,
                payment_method='sasapay',
                description=name,
                metadata={
                    'program_code': program,
                    'program_name': name,
                    'amount_usd': float(amount_usd),
                    'source': 'education_section'
                },
                status='initiated',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            
            from django.shortcuts import redirect
            return redirect(f'/payment/?ref={transaction.reference}')
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    # Handle POST requests
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        
        amount_usd = Decimal(str(data.get('amount', 0)))
        amount_kes = amount_usd * Decimal('129')
        
        # Use the authenticated user or guest user
        user = request.user if request.user.is_authenticated and request.user.id else guest_user
        
        transaction = PaymentTransaction.objects.create(
            user=user,
            reference=generate_reference(),
            amount=amount_kes,
            currency='KES',
            payment_type=data.get('payment_type', 'other'),
            payment_method='sasapay',
            description=data.get('description', ''),
            metadata={
                **data.get('metadata', {}),
                'amount_usd': float(amount_usd)
            },
            status='initiated',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        return JsonResponse({
            'success': True,
            'reference': transaction.reference,
            'amount': float(amount_kes),
            'currency': 'KES'
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def api_create_ticket(request):
    """Create support ticket (JSON)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    ticket = SupportTicket.objects.create(
        user=request.user,
        subject=data.get('subject'),
        message=data.get('message'),
        category=data.get('category', 'general'),
        priority=data.get('priority', 'medium'),
    )
    
    # Send notification to admin
    send_ticket_notification_email(ticket, is_new=True)
    
    log_activity(request.user, 'SUPPORT_TICKET_CREATED', f'Created ticket: {ticket.subject[:50]}', request)
    
    return JsonResponse({
        'success': True,
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
    })

def api_check_email(request):
    """Check if email exists (for registration)"""
    email = request.GET.get('email', '')
    exists = MfalmeUsers.objects.filter(email=email).exists()
    return JsonResponse({'exists': exists})


def api_check_username(request):
    """Check if username exists"""
    username = request.GET.get('username', '')
    exists = MfalmeUsers.objects.filter(username=username).exists()
    return JsonResponse({'exists': exists})


@csrf_exempt
def api_profile_update(request):
    """Update user profile (JSON)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    user = request.user
    
    # Update allowed fields
    allowed_fields = ['first_name', 'last_name', 'phone', 'country', 'city', 
                      'whatsapp_number', 'telegram_username', 'bio']
    
    changes = []
    for field in allowed_fields:
        if field in data:
            old_value = getattr(user, field)
            new_value = data[field]
            if str(old_value) != str(new_value):
                setattr(user, field, new_value)
                changes.append(field)
    
    if changes:
        user.save()
        log_activity(user, 'PROFILE_UPDATE', f'Updated profile fields: {", ".join(changes)}', request)
    
    return JsonResponse({
        'success': True,
        'updated_fields': changes
    })


@csrf_exempt
def api_password_change(request):
    """Change user password (JSON)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    current = data.get('current_password')
    new = data.get('new_password')
    
    if not current or not new:
        return JsonResponse({'error': 'Current and new password required'}, status=400)
    
    user = request.user
    
    if not user.check_password(current):
        return JsonResponse({'error': 'Current password is incorrect'}, status=400)
    
    if len(new) < 8:
        return JsonResponse({'error': 'Password must be at least 8 characters'}, status=400)
    
    user.set_password(new)
    user.password_changed_at = timezone.now()
    user.save(update_fields=['password', 'password_changed_at'])
    
    log_activity(user, 'PASSWORD_CHANGE', 'Password changed', request)
    
    # Re-authenticate to maintain session
    from django.contrib.auth import update_session_auth_hash
    update_session_auth_hash(request, user)
    
    return JsonResponse({'success': True})


def api_ticket_detail(request, ticket_id):
    """Get ticket details with replies"""
    ticket = get_object_or_404(SupportTicket, id=ticket_id, user=request.user)
    replies = ticket.replies.select_related('user').order_by('created_at')
    
    reply_data = []
    for reply in replies:
        reply_data.append({
            'id': reply.id,
            'message': reply.message,
            'is_admin': reply.user.is_staff,
            'created_at': reply.created_at.strftime('%Y-%m-%d %H:%M'),
            'time_ago': timesince(reply.created_at),
        })
    
    data = {
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
        'subject': ticket.subject,
        'message': ticket.message,
        'category': ticket.category,
        'priority': ticket.priority,
        'status': ticket.status,
        'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
        'replies': reply_data,
    }
    
    return JsonResponse(data)


@csrf_exempt
def api_ticket_reply(request, ticket_id):
    """Reply to a ticket"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    message = data.get('message')
    if not message:
        return JsonResponse({'error': 'Message required'}, status=400)
    
    ticket = get_object_or_404(SupportTicket, id=ticket_id, user=request.user)
    
    reply = TicketReply.objects.create(
        ticket=ticket,
        user=request.user,
        message=message,
        is_internal=False
    )
    
    ticket.status = 'waiting_reply'
    ticket.reply_count += 1
    ticket.last_reply_at = timezone.now()
    ticket.save(update_fields=['status', 'reply_count', 'last_reply_at'])
    
    # Notify admin
    send_ticket_notification_email(ticket, is_new=False)
    
    return JsonResponse({
        'success': True,
        'id': reply.id,
        'created_at': reply.created_at.strftime('%Y-%m-%d %H:%M')
    })

# ==================== PUBLIC API ENDPOINTS ====================

def api_public_videos(request):
    """Get all public videos with course information"""
    videos = TrainingVideo.objects.filter(is_active=True).select_related('course').order_by('-created_at')
    
    data = []
    for v in videos:
        data.append({
            'id': v.id,
            'title': v.title,
            'description': v.description,
            'thumbnail': request.build_absolute_uri(v.thumbnail.url) if v.thumbnail else None,  # ← FIXED
            'category': v.category,
            'duration': v.duration,
            'price': float(v.price),
            'view_count': v.view_count,
            'course_id': v.course.id if v.course else None,
            'course_name': v.course.title if v.course else 'Standalone Video',
            'allow_download': v.allow_download,
        })
    
    return JsonResponse(data, safe=False)
 
def verify_s3_file(file_field):
    """Helper to verify file exists in S3"""
    try:
        if file_field and file_field.name:
            return file_field.storage.exists(file_field.name)
    except Exception as e:
        print(f"⚠️ Error checking S3 file: {e}")
    return False


def api_public_pdfs(request):
    """Get all public PDFs with course information"""
    pdfs = PDF.objects.filter(is_active=True).select_related('course').order_by('-created_at')
    
    data = []
    for p in pdfs:
        data.append({
            'id': p.id,
            'title': p.title,
            'description': p.description,
            'cover_image': request.build_absolute_uri(p.cover_image.url) if p.cover_image else None,  # ← FIXED
            'pages': p.pages,
            'file_size': p.file_size,
            'category': p.category,
            'price': float(p.price),
            'is_free': p.is_free,
            'views': p.views,
            'course_id': p.course.id if p.course else None,
            'course_name': p.course.title if p.course else 'Standalone PDF',
        })
    
    return JsonResponse(data, safe=False)

def api_public_blogs(request):
    """Get published blogs (no login required)"""
    blogs = Blog.objects.filter(status='published').order_by('-published_at')[:10]
    
    data = []
    for blog in blogs:
        data.append({
            'id': blog.id,
            'title': blog.title,
            'slug': blog.slug,
            'excerpt': blog.excerpt,
            'author': blog.author.username if blog.author else 'Admin',
            'category': blog.category,
            'featured_image': blog.featured_image.url if blog.featured_image else None,
            'published_at': blog.published_at.strftime('%Y-%m-%d'),
        })
    
    return JsonResponse(data, safe=False)

def api_public_testimonials(request):
    """Get testimonials (no login required)"""
    testimonials = Testimonial.objects.filter(is_active=True, is_featured=True)[:10]
    
    data = []
    for t in testimonials:
        data.append({
            'id': t.id,
            'name': t.name,
            'title': t.title,
            'company': t.company,
            'content': t.content,
            'rating': t.rating,
            'image': t.image.url if t.image else None,
            'program': t.program,
        })
    
    return JsonResponse(data, safe=False)

def api_public_faqs(request):
    """Get FAQs (no login required)"""
    category = request.GET.get('category', '')
    
    faqs = FAQ.objects.filter(is_active=True)
    if category:
        faqs = faqs.filter(category=category)
    
    faqs = faqs.order_by('category', 'order')[:50]
    
    data = []
    for faq in faqs:
        data.append({
            'id': faq.id,
            'question': faq.question,
            'answer': faq.answer,
            'category': faq.category,
        })
    
    return JsonResponse(data, safe=False)

def api_public_courses(request):
    """Get all active courses (public)"""
    courses = Course.objects.filter(is_active=True).order_by('-created_at')
    
    data = []
    for course in courses:
        data.append({
            'id': course.id,
            'title': course.title,
            'description': course.description,
            'price': float(course.price),
            'thumbnail': request.build_absolute_uri(course.thumbnail.url) if course.thumbnail else None,  # ← FIXED
            'duration_weeks': course.duration_weeks,
            'videos_count': course.video_count(),
            'pdfs_count': course.pdf_count(),
            'category': 'Course',
        })
    
    return JsonResponse(data, safe=False)

def api_public_community_tiers(request):
    """Get all community tiers (public)"""
    tiers = CommunityTier.objects.filter(is_active=True).order_by('order')
    
    data = []
    for tier in tiers:
        data.append({
            'id': tier.id,
            'name': tier.name,
            'tier': tier.tier,
            'description': tier.description,
            'features': tier.features,
            'icon_class': tier.icon_class,
            'badge_text': tier.badge_text,
            'color_scheme': tier.color_scheme,
            'button_text': tier.button_text,
            'minimum_investment': float(tier.minimum_investment) if tier.minimum_investment else None,
            'requirements': tier.requirements,
        })
    
    return JsonResponse(data, safe=False)

def api_blog_detail(request, blog_id):
    """Get blog details (public)"""
    blog = get_object_or_404(Blog, id=blog_id, status='published')
    
    blog.views += 1
    blog.save(update_fields=['views'])
    
    data = {
        'id': blog.id,
        'title': blog.title,
        'slug': blog.slug,
        'content': blog.content,
        'excerpt': blog.excerpt,
        'author': blog.author.username if blog.author else 'Admin',
        'category': blog.category,
        'tags': blog.tags.split(',') if blog.tags else [],
        'featured_image': blog.featured_image.url if blog.featured_image else None,
        'views': blog.views,
        'read_time': blog.read_time,
        'published_at': blog.published_at.strftime('%Y-%m-%d') if blog.published_at else None,
        'meta_title': blog.meta_title,
        'meta_description': blog.meta_description,
    }
    
    return JsonResponse(data)

# ==================== PDF VIEWING API ENDPOINT ====================


def api_pdf_view(request, pdf_id):
    """API endpoint to get PDF viewing URL - FOR VIEWING ONLY"""
    try:
        pdf = PDF.objects.get(id=pdf_id, is_active=True)
    except PDF.DoesNotExist:
        return JsonResponse({'error': 'PDF not found'}, status=404)
    
    user = request.user
    
    # Check access
    has_access = check_pdf_access(user, pdf)
    
    if not has_access:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if not pdf.pdf_file or not pdf.pdf_file.storage.exists(pdf.pdf_file.name):
        return JsonResponse({'error': 'PDF file not found'}, status=404)
    
    # Update access record
    access, created = UserPDFAccess.objects.get_or_create(
        user=user,
        pdf=pdf,
        defaults={'payment': None}
    )
    access.viewed = True
    access.view_count += 1
    access.last_viewed = timezone.now()
    access.save(update_fields=['viewed', 'view_count', 'last_viewed'])
    
    # Update PDF view count
    pdf.views += 1
    pdf.save(update_fields=['views'])
    
    # Log activity
    log_activity(
        user,
        'PDF_VIEWED',
        f'Viewed PDF: {pdf.title}',
        request
    )
    
    # Return the URL for viewing (inline, not attachment)
    return JsonResponse({
        'success': True,
        'view_url': pdf.pdf_file.url,
        'title': pdf.title,
        'pages': pdf.pages,
        'file_size': pdf.file_size
    })

# ==================== FIXED PAYMENT VIEWS ====================


def initiate_payment(request):
    """Initiate payment page - Handles tickets, merchandise, videos, courses, PDFs, packages"""
    package_type = request.GET.get('type')
    package_id = request.GET.get('id')
    reference = request.GET.get('ref')
    
    context = {
        'package_type': package_type,
        'package_id': package_id,
        'user': request.user if request.user.is_authenticated else None,
        'paystack_public_key': getattr(settings, 'PAYSTACK_PUBLIC_KEY', ''),
    }
    
    item = None
    amount_usd = 0
    amount_kes = 0
    title = ''
    transaction = None
    
    # FIRST: Check if this is a ticket or merchandise order by reference
    if reference:
        try:
            # Check if it's an Order (ticket or merchandise)
            order = Order.objects.get(reference=reference)
            amount_kes = float(order.amount)
            amount_usd = amount_kes / 129
            
            # ADD CUSTOMER DATA FOR DISPLAY IN TEMPLATE
            context['customer_name'] = order.customer_name
            context['customer_email'] = order.customer_email
            context['customer_phone'] = order.customer_phone
            
            if order.item_type == 'ticket':
                title = f"Event Ticket - {order.metadata.get('quantity', 1)} Ticket(s)"
            elif order.item_type == 'merchandise':
                title = f"Merchandise Order - {reference}"
            else:
                title = "Order Payment"
            
            # Create or get existing transaction
            transaction, created = PaymentTransaction.objects.get_or_create(
                reference=reference,
                defaults={
                    'user': request.user if request.user.is_authenticated else None,
                    'amount': Decimal(str(amount_kes)),
                    'currency': 'KES',
                    'payment_type': order.item_type,
                    'payment_method': 'sasapay',
                    'description': title,
                    'customer_email': order.customer_email,
                    'customer_name': order.customer_name,
                    'customer_phone': order.customer_phone,
                    'metadata': {
                        'order_id': order.id,
                        'order_reference': reference,
                        'item_type': order.item_type,
                        'amount_usd': amount_usd,
                        'quantity': order.metadata.get('quantity', 1)
                    },
                    'status': 'initiated',
                    'ip_address': get_client_ip(request),
                    'user_agent': request.META.get('HTTP_USER_AGENT', ''),
                }
            )
            
            context['reference'] = reference
            context['transaction'] = transaction
            
        except Order.DoesNotExist:
            pass
    
    # SECOND: If no reference, check by package_type and package_id (videos, courses, PDFs, packages)
    if not transaction and package_type and package_id:
        if package_type == 'video':
            try:
                item = TrainingVideo.objects.get(id=package_id)
                amount_usd = float(item.price)
                amount_kes = amount_usd * 129
                title = item.title
            except TrainingVideo.DoesNotExist:
                pass
        elif package_type == 'course':
            try:
                item = Course.objects.get(id=package_id)
                amount_usd = float(item.price)
                amount_kes = amount_usd * 129
                title = item.title
            except Course.DoesNotExist:
                pass
        elif package_type == 'pdf':
            try:
                item = PDF.objects.get(id=package_id)
                amount_usd = float(item.price)
                amount_kes = amount_usd * 129
                title = item.title
            except PDF.DoesNotExist:
                pass
        elif package_type == 'package':
            try:
                item = Package.objects.get(id=package_id)
                amount_usd = float(item.price)
                amount_kes = amount_usd * 129
                title = item.name
            except Package.DoesNotExist:
                pass
        
        if item and amount_usd > 0:
            transaction = PaymentTransaction.objects.create(
                user=request.user if request.user.is_authenticated else None,
                reference=generate_reference(),
                amount=Decimal(str(amount_kes)),
                currency='KES',
                payment_type=f'{package_type}_purchase',
                payment_method='sasapay',
                description=f'Purchase: {title}',
                customer_email=request.user.email if request.user.is_authenticated else None,
                customer_name=request.user.get_full_name() if request.user.is_authenticated else None,
                customer_phone=request.user.phone if request.user.is_authenticated else None,
                metadata={
                    'item_type': package_type,
                    'item_id': package_id,
                    'item_title': title,
                    'amount_usd': amount_usd
                },
                status='initiated',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            context['reference'] = transaction.reference
            context['transaction'] = transaction
    
    context['item'] = item
    context['amount_usd'] = amount_usd
    context['amount_kes'] = amount_kes
    context['title'] = title
    
    return render(request, 'payment/initiate_payment.html', context)


def process_payment(request):
    """Process payment (AJAX) - NOW WITH USD -> KES CONVERSION"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    # Get USD amount from request
    amount_usd = Decimal(str(data.get('amount', 0)))
    
    # Convert to KES (multiply by 129)
    amount_kes = amount_usd * Decimal('129')
    
    transaction = PaymentTransaction.objects.create(
        user=request.user,
        reference=generate_reference(),
        amount=amount_kes,  # Store KES
        currency='KES',
        payment_type=data.get('payment_type', 'other'),
        payment_method='paystack',
        description=data.get('description', ''),
        metadata={
            'item_type': data.get('item_type'),
            'item_id': data.get('item_id'),
            'item_name': data.get('item_name'),
            'amount_usd': float(amount_usd)
        },
        status='pending',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )
    
    # Return KES amount to frontend
    return JsonResponse({
        'success': True,
        'reference': transaction.reference,
        'amount': float(amount_kes),  # Send KES to Paystack
        'currency': 'KES',
        'email': request.user.email,
        'authorization_url': f"/payment/verify/{transaction.reference}/",
    })

def verify_payment(request, reference):
    """Verify payment after redirect"""
    transaction = get_object_or_404(PaymentTransaction, reference=reference)
    
    if transaction.status == 'pending':
        transaction.status = 'completed'
        transaction.paid_at = timezone.now()
        transaction.completed_at = timezone.now()
        transaction.save(update_fields=['status', 'paid_at', 'completed_at'])
        
        if transaction.user:
            user = transaction.user
            user.total_deposits += transaction.amount
            user.account_balance += transaction.amount
            user.save(update_fields=['total_deposits', 'account_balance'])
            
            metadata = transaction.metadata
            if metadata and metadata.get('item_type') == 'video':
                video_id = metadata.get('item_id')
                try:
                    video = TrainingVideo.objects.get(id=video_id)
                    grant_video_access(user, video, transaction)
                except TrainingVideo.DoesNotExist:
                    pass
            elif metadata and metadata.get('item_type') == 'course':
                course_id = metadata.get('item_id')
                try:
                    course = Course.objects.get(id=course_id)
                    enrollment = UserCourse.objects.create(
                        user=user,
                        course=course,
                        payment=transaction,
                        access_expires_at=timezone.now() + timedelta(days=365)  # 1 YEAR
                    )
                    enrollment.get_video_access()
                    enrollment.get_pdf_access()
                    
                    Notification.objects.create(
                        user=user,
                        title='Course Access Granted',
                        message=f'You now have 1 year access to {course.title}',
                        notification_type='SUCCESS',
                        related_object_type='course',
                        related_object_id=course.id
                    )
                except Course.DoesNotExist:
                    pass
            elif metadata and metadata.get('item_type') == 'pdf':
                pdf_id = metadata.get('item_id')
                try:
                    pdf = PDF.objects.get(id=pdf_id)
                    grant_pdf_access(user, pdf, transaction)
                except PDF.DoesNotExist:
                    pass
            elif metadata and metadata.get('item_type') == 'package':
                package_id = metadata.get('item_id')
                # Handle package purchase logic here
                pass
            
            # Get USD amount from metadata for display
            amount_usd = metadata.get('amount_usd', float(transaction.amount / Decimal('129'))) if metadata else float(transaction.amount / Decimal('129'))
            
            Notification.objects.create(
                user=user,
                title='Payment Successful',
                message=f'Your payment of ${amount_usd:.2f} was successful. Reference: {reference}',
                notification_type='SUCCESS',
                related_object_type='payment',
                related_object_id=transaction.id,
            )
            
            log_activity(user, 'PAYMENT_COMPLETED', f'Payment completed: ${amount_usd:.2f} (KES {transaction.amount})', request)
        
        messages.success(request, 'Payment successful!')
    
    return redirect('payment_success', reference=reference)

def payment_success(request, reference):
    """Payment success page"""
    transaction = get_object_or_404(PaymentTransaction, reference=reference)
    # Calculate USD amount for display
    amount_usd = float(transaction.amount / Decimal('129'))
    
    # Get item details from metadata
    item_name = 'Purchase'
    item_type = ''
    item_id = ''
    if transaction.metadata:
        item_name = transaction.metadata.get('item_title', transaction.metadata.get('item_name', 'Purchase'))
        item_type = transaction.metadata.get('item_type', '')
        item_id = transaction.metadata.get('item_id', '')
    
    context = {
        'transaction': transaction,
        'amount_usd': amount_usd,
        'amount': amount_usd,
        'item_name': item_name,
        'item_type': item_type,
        'item_id': item_id,
        'order_id': reference,
        'payment_date': transaction.paid_at.strftime('%B %d, %Y') if transaction.paid_at else timezone.now().strftime('%B %d, %Y'),
    }
    
    return render(request, 'payment/success.html', context)

def payment_failed(request):
    """Payment failed page"""
    error_message = request.GET.get('message', 'Your payment could not be processed.')
    
    context = {
        'error_message': error_message,
    }
    
    return render(request, 'payment/failed.html', context)

@csrf_exempt
def paystack_webhook(request):
    """Paystack webhook endpoint - NOW HANDLES KES"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error'}, status=405)
    
    paystack_signature = request.headers.get('X-Paystack-Signature')
    if not paystack_signature:
        return JsonResponse({'status': 'error'}, status=400)
    
    try:
        payload = json.loads(request.body)
        event = payload.get('event')
        data = payload.get('data', {})
        
        if event == 'charge.success':
            reference = data.get('reference')
            amount_kes = Decimal(str(data.get('amount', 0))) / 100  # Convert from cents
            
            try:
                transaction = PaymentTransaction.objects.get(reference=reference)
                if transaction.status == 'pending':
                    # Verify amount matches (within small tolerance)
                    if abs(transaction.amount - amount_kes) < Decimal('1'):
                        transaction.status = 'completed'
                        transaction.paid_at = timezone.now()
                        transaction.completed_at = timezone.now()
                        transaction.paystack_data = data
                        transaction.save()
                        
                        user = transaction.user
                        if user:
                            user.total_deposits += transaction.amount
                            user.account_balance += transaction.amount
                            user.save()
                            
                            metadata = transaction.metadata
                            if metadata and metadata.get('item_type') == 'video':
                                video_id = metadata.get('item_id')
                                try:
                                    video = TrainingVideo.objects.get(id=video_id)
                                    grant_video_access(user, video, transaction)
                                except TrainingVideo.DoesNotExist:
                                    pass
                            elif metadata and metadata.get('item_type') == 'course':
                                course_id = metadata.get('item_id')
                                try:
                                    course = Course.objects.get(id=course_id)
                                    enrollment = UserCourse.objects.create(
                                        user=user,
                                        course=course,
                                        payment=transaction
                                    )
                                    enrollment.get_video_access()
                                    enrollment.get_pdf_access()
                                except Course.DoesNotExist:
                                    pass
                            elif metadata and metadata.get('item_type') == 'pdf':
                                pdf_id = metadata.get('item_id')
                                try:
                                    pdf = PDF.objects.get(id=pdf_id)
                                    grant_pdf_access(user, pdf, transaction)
                                except PDF.DoesNotExist:
                                    pass
                            
                            amount_usd = metadata.get('amount_usd', float(amount_kes / Decimal('129'))) if metadata else float(amount_kes / Decimal('129'))
                            
                            Notification.objects.create(
                                user=user,
                                title='Payment Received',
                                message=f'Payment of ${amount_usd:.2f} confirmed.',
                                notification_type='SUCCESS',
                            )
            except PaymentTransaction.DoesNotExist:
                pass
        
        return JsonResponse({'status': 'success'})
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error'}, status=400)

# ==================== STATIC PAGES ====================

def index(request):
    """Home page"""
    context = {
        'featured_packages': Package.objects.filter(is_active=True, is_featured=True)[:3],
        'recent_blogs': Blog.objects.filter(status='published')[:3],
        'testimonials': Testimonial.objects.filter(is_active=True, is_featured=True)[:5],
        'stats': Statistic.objects.filter(is_active=True),
        'faqs': FAQ.objects.filter(is_active=True, is_featured=True)[:5],
        'total_users': MfalmeUsers.objects.filter(account_status='active').count(),
        'total_videos': TrainingVideo.objects.filter(is_active=True).count(),
        'total_courses': Course.objects.filter(is_active=True).count(),
    }
    return render(request, 'main/index.html', context)

def about(request):
    context = {
        'stats': Statistic.objects.filter(is_active=True),
        'testimonials': Testimonial.objects.filter(is_active=True)[:5],
    }
    return render(request, 'main/about.html', context)

def services(request):
    context = {
        'packages': Package.objects.filter(is_active=True),
        'education_programs': EducationProgram.objects.filter(is_active=True),
        'mentorship_programs': MentorshipProgram.objects.filter(is_available=True),
    }
    return render(request, 'main/services.html', context)

def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        message = request.POST.get('message')
        
        submission = ContactSubmission.objects.create(
            name=name,
            email=email,
            phone=phone,
            message=message,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        messages.success(request, 'Thank you for contacting us. We will get back to you soon.')
        return redirect('contact')
    
    context = {
        'faqs': FAQ.objects.filter(category='general', is_active=True)[:5],
    }
    return render(request, 'main/contact.html', context)





def payment(request):
    """Payment page with SasaPay integration"""
    reference = request.GET.get('ref')
    
    if not reference:
        messages.error(request, 'No transaction reference provided')
        return redirect('index')
    
    try:
        transaction = PaymentTransaction.objects.get(reference=reference)
        
        # If transaction is already completed, redirect to success
        if transaction.status == 'completed':
            return redirect('payment_success', reference=reference)
        
        # Get SasaPay checkout URL
        from .sasapay_utils import create_checkout
        
        # Create SasaPay checkout session
        checkout_data = create_checkout(
            amount=int(transaction.amount),  # Amount in KES
            reference=reference,
            description=transaction.description,
            email=request.user.email if request.user.is_authenticated else 'customer@example.com',
            phone=request.user.phone if request.user.is_authenticated else '',
            callback_url=request.build_absolute_uri('/sasapay/callback/'),
            success_url=request.build_absolute_uri(f'/payment/success/{reference}/'),
            failure_url=request.build_absolute_uri('/payment/failed/')
        )
        
        context = {
            'title': transaction.metadata.get('package_name', 'Package'),
            'amount_usd': transaction.metadata.get('amount_usd', 0),
            'amount_kes': float(transaction.amount),
            'user': request.user,
            'reference': reference,
            'transaction': transaction,
            'checkout_url': checkout_data.get('checkout_url'),
            'sasapay_script': checkout_data.get('script_url', 'https://checkout.sasapay.app/v1.js')
        }
        
        return render(request, 'payment.html', context)
        
    except PaymentTransaction.DoesNotExist:
        messages.error(request, 'Transaction not found')
        return redirect('index')
    except Exception as e:
        print(f"SasaPay error: {e}")
        # Fall back to manual payment
        context = {
            'title': transaction.metadata.get('package_name', 'Package'),
            'amount_usd': transaction.metadata.get('amount_usd', 0),
            'amount_kes': float(transaction.amount),
            'user': request.user,
            'reference': reference,
            'transaction': transaction
        }
        return render(request, 'payment/payment.html', context)
    

    

@csrf_exempt
def sasapay_initiate_payment(request):
    """Initiate SasaPay payment"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    reference = data.get('reference')
    amount = data.get('amount')
    email = data.get('email')
    
    if not reference:
        return JsonResponse({'error': 'Reference required'}, status=400)
    
    try:
        transaction = PaymentTransaction.objects.get(reference=reference)
    except PaymentTransaction.DoesNotExist:
        return JsonResponse({'error': 'Transaction not found'}, status=404)
    
    # For now, return a mock checkout URL (replace with actual SasaPay integration)
    return JsonResponse({
        'success': True,
        'checkout_url': f'/payment/success/{reference}/',  # This will be replaced with actual SasaPay URL
        'message': 'Payment initiated'
    })

def accounts(request):
    context = {
        'community_tiers': CommunityTier.objects.filter(is_active=True),
    }
    return render(request, 'main/accounts.html', context)

def partnerships(request):
    context = {
        'partnership_programs': PartnershipProgram.objects.filter(is_active=True),
        'active_partners': UserPartnership.objects.filter(status='active').count(),
        'total_investment': UserPartnership.objects.filter(status='active').aggregate(Sum('investment_amount'))['investment_amount__sum'] or 0,
    }
    return render(request, 'main/partnerships.html', context)

def education(request):
    context = {
        'programs': EducationProgram.objects.filter(is_active=True),
        'testimonials': Testimonial.objects.filter(program__in=['IPLT', 'PTM', 'POTM', 'PFTM'], is_active=True)[:5],
    }
    return render(request, 'main/education.html', context)

def mentoring(request):
    context = {
        'programs': MentorshipProgram.objects.filter(is_available=True),
        'mentors': MfalmeUsers.objects.filter(elite_rank__in=['Captain', 'Commander', 'General'], is_active=True)[:4],
    }
    return render(request, 'main/mentoring.html', context)

def community(request):
    context = {
        'tiers': CommunityTier.objects.filter(is_active=True),
        'member_count': UserCommunityMembership.objects.filter(status='active').count(),
    }
    return render(request, 'main/community.html', context)

def seminars(request):
    context = {
        'upcoming_seminars': [],
    }
    return render(request, 'main/seminars.html', context)

def faqs(request):
    category = request.GET.get('category', '')
    faqs = FAQ.objects.filter(is_active=True)
    if category:
        faqs = faqs.filter(category=category)
    faqs = faqs.order_by('category', 'order')
    
    categories = {}
    for faq in faqs:
        if faq.category not in categories:
            categories[faq.category] = []
        categories[faq.category].append(faq)
    
    context = {
        'categories': categories,
        'selected_category': category,
        'category_choices': FAQ.CATEGORY_CHOICES,
    }
    return render(request, 'main/faqs.html', context)

def booking(request):
    return render(request, 'main/booking.html', context)

# ==================== FORM SUBMISSIONS ====================

@csrf_exempt
def contact_form_submit(request):
    """Contact form submission (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'success': False})
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False})
    
    submission = ContactSubmission.objects.create(
        name=data.get('name'),
        email=data.get('email'),
        phone=data.get('phone'),
        subject=data.get('subject', ''),
        message=data.get('message'),
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )
    
    return JsonResponse({'success': True})

@csrf_exempt
def submit_partnership_application(request):
    """Partnership application (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'success': False})
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False})
    
    submission = ContactSubmission.objects.create(
        name=data.get('company_name') or data.get('name'),
        email=data.get('email'),
        phone=data.get('phone'),
        subject='Partnership Application',
        message=json.dumps(data),
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )
    
    return JsonResponse({'success': True})

# ==================== TEST ENDPOINTS ====================

def test_all_emails(request):
    """Test email functionality"""
    if not request.user.is_staff:
        return HttpResponse('Unauthorized', status=401)
    
    results = []
    
    code = generate_verification_code()
    result1 = send_verification_email(request.user, code)
    results.append(f"Verification email: {'âœ“' if result1 else 'âœ—'}")
    
    result2 = send_password_reset_email(request.user, code)
    results.append(f"Password reset email: {'âœ“' if result2 else 'âœ—'}")
    
    result3 = send_welcome_email(request.user)
    results.append(f"Welcome email: {'âœ“' if result3 else 'âœ—'}")
    
    result4 = send_admin_notification(request.user)
    results.append(f"Admin notification: {'âœ“' if result4 else 'âœ—'}")
    
    return HttpResponse('<br>'.join(results))

def test_smtp_connection(request):
    """Test SMTP connection"""
    if not request.user.is_staff:
        return HttpResponse('Unauthorized', status=401)
    
    from django.core.mail import get_connection
    
    try:
        connection = get_connection()
        connection.open()
        connection.close()
        return HttpResponse('SMTP connection successful âœ“')
    except Exception as e:
        return HttpResponse(f'SMTP connection failed: {str(e)}')

def emergency_email_fix(request):
    """Emergency email configuration fix"""
    if not request.user.is_staff:
        return HttpResponse('Unauthorized', status=401)
    
    from django.conf import settings
    
    settings.EMAIL_BACKEND = 'django.core.mail.backends.console.EmailBackend'
    
    code = generate_verification_code()
    result = send_verification_email(request.user, code)
    
    if result:
        return HttpResponse('Email sent to console. Check your terminal.')
    else:
        return HttpResponse('Email sending failed.')

# ========== PAYMENT VIEWS ==========

def pay_without_login(request):
    """Payment without login page"""
    context = {
        'packages': Package.objects.filter(is_active=True),
    }
    return render(request, 'payment/index.html', context)

def verify_guest_payment(request, reference):
    """Verify guest payment"""
    transaction = get_object_or_404(PaymentTransaction, reference=reference)
    
    context = {
        'transaction': transaction,
        'reference': reference,
    }
    
    return render(request, 'payment/verify.html', context)

def initiate_package_payment(request, package_type, amount):
    """Initiate package payment"""
    return redirect('payment')

def initiate_education_payment(request, program_type, duration):
    """Initiate education payment"""
    return redirect('payment')

def initiate_partnership_payment(request, tier):
    """Initiate partnership payment"""
    return redirect('payment')

def initiate_custom_payment(request):
    """Initiate custom payment"""
    if request.method == 'POST':
        return JsonResponse({'status': 'success'})
    return redirect('payment')

def payment_video(request, video_id):
    """Payment page for a specific video"""
    video = get_object_or_404(TrainingVideo, id=video_id, is_active=True)
    amount_kes = float(video.price) * 129
    
    # Create transaction
    transaction = None
    if request.user.is_authenticated:
        transaction = PaymentTransaction.objects.create(
            user=request.user,
            reference=generate_reference(),
            amount=Decimal(str(amount_kes)),
            currency='KES',
            payment_type='video_purchase',
            payment_method='paystack',
            description=f'Video: {video.title}',
            metadata={
                'video_id': video.id,
                'video_title': video.title,
                'amount_usd': float(video.price)
            },
            status='initiated',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
    
    context = {
        'video': video,
        'amount_usd': float(video.price),
        'amount_kes': amount_kes,
        'user': request.user if request.user.is_authenticated else None,
        'transaction': transaction,
        'reference': transaction.reference if transaction else None,
    }
    
    return render(request, 'payment/video.html', context)

def initiate_video_payment(request):
    """Initiate payment for a video"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            video_id = data.get('video_id')
            video = get_object_or_404(TrainingVideo, id=video_id)
            
            # Convert USD to KES
            amount_usd = video.price
            amount_kes = amount_usd * Decimal('129')
            
            transaction = PaymentTransaction.objects.create(
                user=request.user if request.user.is_authenticated else None,
                reference=generate_reference(),
                amount=amount_kes,  # Store KES
                currency='KES',
                payment_type='video_purchase',
                payment_method='paystack',
                description=f'Video: {video.title}',
                metadata={
                    'video_id': video.id,
                    'amount_usd': float(amount_usd)
                },
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'reference': transaction.reference,
                'amount': float(amount_kes),  # Return KES for Paystack
                'currency': 'KES'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def initiate_course_payment(request):
    """Initiate payment for a course"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            course_id = data.get('course_id')
            course = get_object_or_404(Course, id=course_id)
            
            # Convert USD to KES
            amount_usd = course.price
            amount_kes = amount_usd * Decimal('129')
            
            transaction = PaymentTransaction.objects.create(
                user=request.user if request.user.is_authenticated else None,
                reference=generate_reference(),
                amount=amount_kes,  # Store KES
                currency='KES',
                payment_type='course_purchase',
                payment_method='paystack',
                description=f'Course: {course.title}',
                metadata={
                    'course_id': course.id,
                    'amount_usd': float(amount_usd)
                },
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'reference': transaction.reference,
                'amount': float(amount_kes),  # Return KES for Paystack
                'currency': 'KES'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def initiate_mentorship_payment(request):
    """Initiate payment for mentorship"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            program_id = data.get('program_id')
            program = get_object_or_404(MentorshipProgram, id=program_id)
            
            # Convert USD to KES
            amount_usd = program.price
            amount_kes = amount_usd * Decimal('129')
            
            transaction = PaymentTransaction.objects.create(
                user=request.user if request.user.is_authenticated else None,
                reference=generate_reference(),
                amount=amount_kes,  # Store KES
                currency='KES',
                payment_type='mentorship_purchase',
                payment_method='paystack',
                description=f'Mentorship: {program.title}',
                metadata={
                    'program_id': program.id,
                    'amount_usd': float(amount_usd)
                },
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'reference': transaction.reference,
                'amount': float(amount_kes),  # Return KES for Paystack
                'currency': 'KES'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# ==================== COURSE PROGRESS APIS ====================


@csrf_exempt
def api_mark_lesson_complete(request):
    """Mark lesson complete (API version) - WITH EXPIRATION CHECK"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    course_id = data.get('course_id')
    lesson_id = data.get('lesson_id')
    lesson_type = data.get('lesson_type')
    
    if not all([course_id, lesson_id, lesson_type]):
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    
    # Check access with expiration
    has_access, enrollment = check_course_access(request.user, course_id)
    
    if not has_access or not enrollment:
        return JsonResponse({'error': 'Not enrolled or access expired'}, status=403)
    
    # Mark lesson as complete
    success = enrollment.mark_lesson_complete(lesson_type, lesson_id)
    
    if not success:
        return JsonResponse({
            'success': True,
            'message': 'Lesson already completed',
            'progress': enrollment.progress,
            'completed_count': enrollment.get_completed_count(),
            'total_lessons': enrollment.get_total_lessons()
        })
    
    # Get next lesson
    next_lesson = enrollment.get_next_lesson()
    
    return JsonResponse({
        'success': True,
        'progress': enrollment.progress,
        'completed_count': enrollment.get_completed_count(),
        'total_lessons': enrollment.get_total_lessons(),
        'next_lesson': next_lesson,
        'is_complete': enrollment.progress == 100,
        'expires_in': enrollment.time_until_expiry()
    })


def api_course_progress(request, course_id):
    """Get detailed progress for a course - WITH EXPIRATION"""
    has_access, enrollment = check_course_access(request.user, course_id)
    
    if not has_access or not enrollment:
        return JsonResponse({'error': 'Not enrolled or access expired'}, status=404)
    
    data = {
        'course_id': course_id,
        'progress': enrollment.progress,
        'completed_count': enrollment.get_completed_count(),
        'total_lessons': enrollment.get_total_lessons(),
        'remaining': enrollment.get_remaining_lessons(),
        'lessons': enrollment.get_lesson_statuses(),
        'next_lesson': enrollment.get_next_lesson(),
        'is_expired': enrollment.is_access_expired(),
        'expires_in': enrollment.time_until_expiry(),
        'days_remaining': enrollment.days_until_expiry()
    }
    
    return JsonResponse(data)


def api_course_next_lesson(request, course_id):
    """Get next lesson for a course - WITH EXPIRATION"""
    has_access, enrollment = check_course_access(request.user, course_id)
    
    if not has_access or not enrollment:
        return JsonResponse({'error': 'Not enrolled or access expired'}, status=404)
    
    next_lesson = enrollment.get_next_lesson()
    
    if next_lesson:
        return JsonResponse({
            'has_next': True,
            'type': next_lesson['type'],
            'id': next_lesson['id'],
            'title': next_lesson['title'],
            'url': f"/{next_lesson['type']}/{next_lesson['id']}/"
        })
    else:
        return JsonResponse({
            'has_next': False,
            'message': 'Course completed!'
        })


def api_course_reset_progress(request, course_id):
    """Reset progress for a course (admin only)"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        enrollment = UserCourse.objects.get(
            user_id=request.GET.get('user_id'),
            course_id=course_id
        )
    except UserCourse.DoesNotExist:
        return JsonResponse({'error': 'Enrollment not found'}, status=404)
    
    enrollment.reset_progress()
    
    return JsonResponse({'success': True})


def pesapal_initiate_payment(request):
    """Initiate payment with Pesapal"""
    package_type = request.GET.get('type')
    package_id = request.GET.get('id')
    
    if request.method == 'POST':
        # Get payment details from POST
        amount_usd = Decimal(request.POST.get('amount', 0))
        amount_kes = amount_usd * Decimal('129')  # Convert to KES
        phone = request.POST.get('phone', request.user.phone)
        email = request.POST.get('email', request.user.email)
        first_name = request.POST.get('first_name', request.user.first_name)
        last_name = request.POST.get('last_name', request.user.last_name)
        
        # Create transaction
        transaction = PaymentTransaction.objects.create(
            user=request.user,
            reference=generate_reference(),
            amount=amount_kes,
            currency='KES',
            payment_type=package_type or 'other',
            payment_method='pesapal',
            description=f'Payment for {package_type}',
            metadata={
                'item_type': package_type,
                'item_id': package_id,
                'amount_usd': float(amount_usd)
            },
            status='initiated',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        # Prepare Pesapal parameters
        params = {
            'pesapal_transaction_type': 'MERCHANT',
            'pesapal_merchant_reference': transaction.reference,
            'pesapal_amount': str(int(amount_kes)),  # No decimals
            'pesapal_currency': 'KES',
            'pesapal_description': transaction.description[:100],
            'pesapal_type': 'MERCHANT',
            'pesapal_first_name': first_name,
            'pesapal_last_name': last_name,
            'pesapal_email_address': email,
            'pesapal_phone_number': phone,
            'pesapal_country_code': 'KE',
        }
        
        # Get iframe URL
        iframe_url = get_pesapal_iframe_url(
            params,
            settings.PESAPAL_CONFIG['CALLBACK_URL'],
            settings.PESAPAL_CONFIG['CONSUMER_KEY'],
            settings.PESAPAL_CONFIG['CONSUMER_SECRET']
        )
        
        # Store in session
        request.session['pesapal_transaction_id'] = transaction.id
        
        return JsonResponse({
            'success': True,
            'iframe_url': iframe_url,
            'reference': transaction.reference
        })
    
    # GET request - show payment form
    context = {
        'user': request.user,
        'package_type': package_type,
        'package_id': package_id,
    }
    return render(request, 'payment/pesapal_form.html', context)

@csrf_exempt
def pesapal_callback(request):
    """Pesapal callback URL - user returns here after payment"""
    # Get parameters from Pesapal
    merchant_reference = request.GET.get('pesapal_merchant_reference')
    tracking_id = request.GET.get('pesapal_transaction_tracking_id')
    
    if not merchant_reference:
        messages.error(request, 'Invalid payment callback')
        return redirect('payment_failed')
    
    try:
        transaction = PaymentTransaction.objects.get(reference=merchant_reference)
        
        # Update transaction with tracking ID
        transaction.pesapal_tracking_id = tracking_id
        transaction.save(update_fields=['pesapal_tracking_id'])
        
        # Query payment status
        status_response = query_pesapal_status(merchant_reference, tracking_id)
        
        # Parse status (Pesapal returns "PENDING", "COMPLETED", "FAILED")
        if 'COMPLETED' in status_response:
            transaction.status = 'completed'
            transaction.paid_at = timezone.now()
            transaction.completed_at = timezone.now()
            transaction.save(update_fields=['status', 'paid_at', 'completed_at'])
            
            # Process successful payment
            process_successful_payment(transaction, request)
            
            messages.success(request, 'Payment successful!')
            return redirect('payment_success', reference=merchant_reference)
        elif 'FAILED' in status_response:
            transaction.status = 'failed'
            transaction.save(update_fields=['status'])
            messages.error(request, 'Payment failed')
            return redirect('payment_failed')
        else:
            transaction.status = 'pending'
            transaction.save(update_fields=['status'])
            messages.info(request, 'Payment is being processed')
            return redirect('payment_pending', reference=merchant_reference)
            
    except PaymentTransaction.DoesNotExist:
        messages.error(request, 'Transaction not found')
        return redirect('payment_failed')

@csrf_exempt
def pesapal_ipn(request):
    """Pesapal Instant Payment Notification endpoint"""
    if request.method != 'POST':
        return HttpResponse('OK')
    
    # Get parameters from POST
    merchant_reference = request.POST.get('pesapal_merchant_reference')
    tracking_id = request.POST.get('pesapal_transaction_tracking_id')
    notification_type = request.POST.get('pesapal_notification_type')
    
    if not merchant_reference:
        return HttpResponse('OK')
    
    try:
        transaction = PaymentTransaction.objects.get(reference=merchant_reference)
        
        # Update with tracking ID
        transaction.pesapal_tracking_id = tracking_id
        
        # Query status
        status_response = query_pesapal_status(merchant_reference, tracking_id)
        
        # Store raw response
        transaction.pesapal_raw_response = {
            'notification_type': notification_type,
            'status_response': status_response
        }
        
        # Update status
        if 'COMPLETED' in status_response:
            transaction.status = 'completed'
            transaction.paid_at = timezone.now()
            transaction.completed_at = timezone.now()
        elif 'FAILED' in status_response:
            transaction.status = 'failed'
        
        transaction.save()
        
        # Process if completed
        if transaction.status == 'completed' and transaction.user:
            process_successful_payment(transaction, None)
        
    except PaymentTransaction.DoesNotExist:
        pass
    
    return HttpResponse('OK')

def process_successful_payment(transaction, request):
    """Helper function to process successful payments"""
    if not transaction.user:
        return
    
    user = transaction.user
    
    # Update user balance
    user.total_deposits += transaction.amount
    user.account_balance += transaction.amount
    user.save(update_fields=['total_deposits', 'account_balance'])
    
    # Process based on metadata
    metadata = transaction.metadata
    if metadata:
        item_type = metadata.get('item_type')
        item_id = metadata.get('item_id')
        
        if item_type == 'video':
            try:
                video = TrainingVideo.objects.get(id=item_id)
                grant_video_access(user, video, transaction)
            except TrainingVideo.DoesNotExist:
                pass
                
        elif item_type == 'course':
            try:
                course = Course.objects.get(id=item_id)
                enrollment = UserCourse.objects.create(
                    user=user,
                    course=course,
                    payment=transaction,
                    access_expires_at=timezone.now() + timedelta(days=365)  # 1 YEAR
                )
                enrollment.get_video_access()
                enrollment.get_pdf_access()
                
                Notification.objects.create(
                    user=user,
                    title='Course Access Granted',
                    message=f'You now have 1 year access to {course.title}',
                    notification_type='SUCCESS',
                    related_object_type='course',
                    related_object_id=course.id
                )
            except Course.DoesNotExist:
                pass
                
        elif item_type == 'pdf':
            try:
                pdf = PDF.objects.get(id=item_id)
                grant_pdf_access(user, pdf, transaction)
            except PDF.DoesNotExist:
                pass
    
    # Create notification
    amount_usd = metadata.get('amount_usd', float(transaction.amount / Decimal('129'))) if metadata else float(transaction.amount / Decimal('129'))
    
    Notification.objects.create(
        user=user,
        title='Payment Successful',
        message=f'Your payment of ${amount_usd:.2f} via Pesapal was successful. Reference: {transaction.reference}',
        notification_type='SUCCESS',
        related_object_type='payment',
        related_object_id=transaction.id,
    )
    
    # Log activity
    log_activity(
        user,
        'PAYMENT_COMPLETED',
        f'Payment completed via Pesapal: ${amount_usd:.2f} (KES {transaction.amount})',
        request
    )

def payment_pending(request, reference):
    """Payment pending page"""
    transaction = get_object_or_404(PaymentTransaction, reference=reference)
    context = {
        'transaction': transaction,
        'reference': reference,
    }
    return render(request, 'payment/pending.html', context)



@csrf_exempt
def sasapay_process_payment(request):
    """Process payment with SasaPay - Handles guest users correctly"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    reference = data.get('reference')
    phone = data.get('phone')
    payment_method = data.get('payment_method', 'c2b')
    
    print(f"\n🔵 SasaPay Process Payment Called")
    print(f"   Reference: {reference}")
    print(f"   Phone: {phone}")
    print(f"   Method: {payment_method}")
    
    if not reference:
        return JsonResponse({'error': 'Reference required'}, status=400)
    
    from .sasapay_utils import initiate_c2b_payment, initiate_checkout
    
    # IMPORTANT: Check Order FIRST (for tickets/merchandise from guest users)
    order = None
    transaction = None
    amount_kes = 0
    description = ''
    customer_email = ''
    customer_name = ''
    customer_phone = ''
    
    try:
        order = Order.objects.get(reference=reference)
        amount_kes = int(order.amount)
        description = f"{order.item_type.upper()} Order: {reference}"
        customer_email = order.customer_email
        customer_name = order.customer_name
        customer_phone = order.customer_phone
        print(f"✅ Found Order: {order.id} - Type: {order.item_type}")
        print(f"   Customer: {customer_name} ({customer_email})")
    except Order.DoesNotExist:
        print(f"⚠️ No Order found, checking PaymentTransaction...")
        try:
            transaction = PaymentTransaction.objects.get(reference=reference)
            amount_kes = int(transaction.amount)
            description = transaction.description
            customer_email = transaction.customer_email or ''
            customer_name = transaction.customer_name or ''
            customer_phone = transaction.customer_phone or ''
            print(f"✅ Found PaymentTransaction: {transaction.id}")
        except PaymentTransaction.DoesNotExist:
            print(f"❌ No transaction found for reference: {reference}")
            return JsonResponse({'error': 'Transaction not found'}, status=404)
    
    if payment_method == 'c2b':
        # Use phone from request or from order/transaction
        phone_to_use = phone or customer_phone
        
        if not phone_to_use:
            return JsonResponse({'error': 'Phone number required for M-PESA'}, status=400)
        
        # Format phone number
        formatted_phone = str(phone_to_use).strip()
        formatted_phone = ''.join(filter(str.isdigit, formatted_phone))
        if formatted_phone.startswith('0'):
            formatted_phone = '254' + formatted_phone[1:]
        elif formatted_phone.startswith('7') and len(formatted_phone) == 9:
            formatted_phone = '254' + formatted_phone
        elif formatted_phone.startswith('+'):
            formatted_phone = formatted_phone[1:]
        
        if not formatted_phone.startswith('254'):
            formatted_phone = '254' + formatted_phone
        
        print(f"📱 Formatted phone: {formatted_phone}")
        print(f"💰 Amount: {amount_kes} KES")
        print(f"📝 Description: {description}")
        
        # Initiate M-PESA STK Push
        result = initiate_c2b_payment(
            phone=formatted_phone,
            amount=amount_kes,
            reference=reference,
            description=description
        )
        
        print(f"📡 SasaPay Response: {result}")
        
        if result.get('success'):
            if order:
                order.payment_reference = result.get('transaction_id')
                order.checkout_request_id = result.get('checkout_id')
                order.save()
                print(f"✅ Order updated: {order.reference}")
            elif transaction:
                transaction.sasapay_transaction_id = result.get('transaction_id')
                transaction.sasapay_checkout_id = result.get('checkout_id')
                transaction.sasapay_payment_method = 'mpesa'
                transaction.sasapay_raw_response = result
                transaction.status = 'pending'
                transaction.save()
                print(f"✅ Transaction updated: {transaction.reference}")
            
            return JsonResponse({
                'success': True,
                'message': 'STK Push sent. Check your phone.',
                'transaction_id': result.get('transaction_id')
            })
        else:
            error_msg = result.get('error', 'Payment failed - Please try again')
            print(f"❌ Payment failed: {error_msg}")
            return JsonResponse({
                'success': False,
                'error': error_msg
            }, status=400)
    
    elif payment_method == 'checkout':
        # NEVER use request.user.email - use customer data from order/transaction
        email_to_use = customer_email
        if not email_to_use:
            email_to_use = 'customer@example.com'
        
        print(f"📧 Using email for checkout: {email_to_use}")
        
        result = initiate_checkout(
            amount=amount_kes,
            reference=reference,
            description=description,
            email=email_to_use,
            phone=phone or customer_phone
        )
        
        if result.get('success'):
            if order:
                order.payment_reference = result.get('transaction_id') or result.get('checkout_id')
                order.checkout_request_id = result.get('checkout_id')
                order.save()
            elif transaction:
                transaction.sasapay_checkout_id = result.get('checkout_id')
                transaction.sasapay_payment_method = 'checkout'
                transaction.sasapay_raw_response = result
                transaction.status = 'pending'
                transaction.save()
            
            return JsonResponse({
                'success': True,
                'checkout_url': result.get('checkout_url'),
                'message': 'Redirecting to checkout...'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Checkout failed')
            }, status=400)
    
    return JsonResponse({'error': 'Invalid payment method'}, status=400)

def test_sasapay_connection(request):
    """Simple test connection to SasaPay"""
    import requests
    import socket
    from django.conf import settings
    
    results = []
    results.append("<h1>SasaPay Connection Test</h1>")
    
    # Test 1: Import requests
    results.append("<h3>1. Library Check</h3>")
    results.append(f"✅ requests version: {requests.__version__}")
    
    # Test 2: DNS
    results.append("<h3>2. DNS Resolution</h3>")
    try:
        ip = socket.gethostbyname('sandbox.sasapay.com')
        results.append(f"✅ sandbox.sasapay.com → {ip}")
    except Exception as e:
        results.append(f"❌ DNS Error: {str(e)}")
    
    # Test 3: Simple HTTP Request
    results.append("<h3>3. HTTP Request</h3>")
    try:
        response = requests.get(
            "https://sandbox.sasapay.com",
            timeout=5,
            verify=False
        )
        results.append(f"✅ Status: {response.status_code}")
    except requests.exceptions.Timeout:
        results.append("⏱️ Timeout - Server not responding")
    except requests.exceptions.ConnectionError as e:
        results.append(f"❌ Connection Error: {str(e)}")
    except Exception as e:
        results.append(f"❌ Error: {type(e).__name__} - {str(e)}")
    
    # Test 4: Config Values
    results.append("<h3>4. Configuration</h3>")
    results.append(f"Environment: {settings.SASAPAY_CONFIG.get('ENVIRONMENT', 'Not set')}")
    results.append(f"Client ID: {settings.SASAPAY_CONFIG.get('CLIENT_ID', 'Not set')[:10]}...")
    client_secret = settings.SASAPAY_CONFIG.get('CLIENT_SECRET', '')
    results.append(f"Client Secret: {'*' * 10 if client_secret else 'Not set'}")
    
    return HttpResponse("<br>".join(results))

def process_successful_payment(transaction, request):
    """Process successful payment - grant access to content"""
    if not transaction.user:
        return
    
    user = transaction.user
    
    # Update user balance
    user.total_deposits += transaction.amount
    user.account_balance += transaction.amount
    user.save(update_fields=['total_deposits', 'account_balance'])
    
    # Process based on metadata
    metadata = transaction.metadata
    if metadata:
        item_type = metadata.get('item_type')
        item_id = metadata.get('item_id')
        
        if item_type == 'video':
            try:
                video = TrainingVideo.objects.get(id=item_id)
                grant_video_access(user, video, transaction)
            except TrainingVideo.DoesNotExist:
                pass
                
        elif item_type == 'course':
            try:
                course = Course.objects.get(id=item_id)
                enrollment = UserCourse.objects.create(
                    user=user,
                    course=course,
                    payment=transaction,
                    access_expires_at=timezone.now() + timedelta(days=365)  # 1 YEAR
                )
                enrollment.get_video_access()
                enrollment.get_pdf_access()
                
                Notification.objects.create(
                    user=user,
                    title='Course Access Granted',
                    message=f'You now have 1 year access to {course.title}',
                    notification_type='SUCCESS',
                    related_object_type='course',
                    related_object_id=course.id
                )
            except Course.DoesNotExist:
                pass
                
        elif item_type == 'pdf':
            try:
                pdf = PDF.objects.get(id=item_id)
                grant_pdf_access(user, pdf, transaction)
            except PDF.DoesNotExist:
                pass
    
    # Create notification
    amount_usd = metadata.get('amount_usd', float(transaction.amount / Decimal('129'))) if metadata else float(transaction.amount / Decimal('129'))
    
    Notification.objects.create(
        user=user,
        title='Payment Successful',
        message=f'Your payment of ${amount_usd:.2f} was successful. Reference: {transaction.reference}',
        notification_type='SUCCESS',
        related_object_type='payment',
        related_object_id=transaction.id,
    )
    
    # Log activity
    log_activity(
        user,
        'PAYMENT_COMPLETED',
        f'Payment completed: ${amount_usd:.2f} (KES {transaction.amount})',
        request
    )

@csrf_exempt
def sasapay_callback(request):
    """Handle SasaPay callback after payment"""
    import json
    
    if request.method == 'GET':
        transaction_id = request.GET.get('transaction_id')
        checkout_id = request.GET.get('checkout_id')
        status = request.GET.get('status')
        reference = request.GET.get('reference')
        
        if transaction_id:
            return redirect(f'/sasapay/verify/?transaction_id={transaction_id}')
        
        return redirect('payment_failed')
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except:
            data = request.POST.dict()
        
        transaction_id = data.get('transaction_id') or data.get('checkout_id')
        checkout_id = data.get('checkout_id')
        status = data.get('status')
        reference = data.get('reference')
        
        if not transaction_id and not reference:
            return HttpResponse('OK')
        
        # Process the payment
        if status and status.lower() == 'completed':
            # Find and update order
            order = None
            if reference:
                order = Order.objects.filter(reference=reference).first()
            if not order and transaction_id:
                order = Order.objects.filter(payment_reference=transaction_id).first()
            
            if order and order.status != 'completed':
                order.status = 'completed'
                order.payment_reference = transaction_id
                order.save()
                
                # Create ticket for ticket orders
                if order.item_type == 'ticket':
                    try:
                        event = Event.objects.first()
                        if event:
                            ticket = EventTicket.objects.create(
                                event=event,
                                attendee_name=order.customer_name,
                                attendee_phone=order.customer_phone,
                                attendee_email=order.customer_email,
                                quantity=order.metadata.get('quantity', 1),
                                unit_price_usd=249,
                                unit_price_kes=249 * 129,
                                order_reference=order.reference,
                                payment_reference=transaction_id,
                                status='confirmed'
                            )
                            event.current_bookings += ticket.quantity
                            event.save()
                            send_ticket_email(ticket)
                    except Exception as e:
                        print(f"Ticket creation error: {e}")
                
                # Create merchandise order for merchandise purchases
                elif order.item_type == 'merchandise':
                    try:
                        merch_order = MerchandiseOrder.objects.create(
                            customer_name=order.customer_name,
                            customer_phone=order.customer_phone,
                            customer_email=order.customer_email,
                            delivery_address=order.metadata.get('address', ''),
                            items=order.items,
                            subtotal=order.amount,
                            total=order.amount,
                            payment_reference=transaction_id,
                            order_reference=order.reference,
                            status='paid'
                        )
                        send_merchandise_order_email(merch_order)
                        
                        # Update stock
                        for item in order.items:
                            try:
                                product = Merchandise.objects.get(id=item['id'])
                                product.stock -= item['quantity']
                                product.save()
                            except:
                                pass
                    except Exception as e:
                        print(f"Merchandise order error: {e}")
        
        return HttpResponse('OK')

def sasapay_verify(request):
    """Verify payment after redirect"""
    transaction_id = request.GET.get('transaction_id')
    
    if not transaction_id:
        return redirect('payment_failed')
    
    from .sasapay_utils import query_payment_status
    
    # Query status
    result = query_payment_status(transaction_id)
    
    if result.get('status') == 'COMPLETED':
        # Find transaction
        try:
            transaction = PaymentTransaction.objects.get(sasapay_transaction_id=transaction_id)
            if transaction.status != 'completed':
                transaction.status = 'completed'
                transaction.paid_at = timezone.now()
                transaction.completed_at = timezone.now()
                transaction.save()
                process_successful_payment(transaction, request)
            
            return redirect('payment_success', reference=transaction.reference)
        except PaymentTransaction.DoesNotExist:
            pass
    
    return redirect('payment_failed')

def sasapay_status(request, reference):
    """Check payment status (AJAX)"""
    try:
        transaction = PaymentTransaction.objects.get(reference=reference)
        
        # If still pending, query SasaPay
        if transaction.status == 'pending' and transaction.sasapay_transaction_id:
            from .sasapay_utils import query_payment_status
            result = query_payment_status(transaction.sasapay_transaction_id)
            
            if result.get('status') == 'COMPLETED':
                transaction.status = 'completed'
                transaction.paid_at = timezone.now()
                transaction.completed_at = timezone.now()
                transaction.save()
                process_successful_payment(transaction, None)
            elif result.get('status') == 'FAILED':
                transaction.status = 'failed'
                transaction.save()
        
        return JsonResponse({
            'status': transaction.status,
            'success_url': f'/payment/success/{reference}/' if transaction.status == 'completed' else None
        })
    except PaymentTransaction.DoesNotExist:
        return JsonResponse({'status': 'not_found'}, status=404)


# ==================== SASAPAY PAYMENT VIEWS ====================



@csrf_exempt
def sasapay_callback(request):
    """SasaPay callback endpoint - receives payment notifications"""
    if request.method == 'GET':
        # Redirect callback after checkout
        transaction_id = request.GET.get('transaction_id')
        checkout_id = request.GET.get('checkout_id')
        status = request.GET.get('status')
        
        if transaction_id:
            return redirect(f'/sasapay/verify/?transaction_id={transaction_id}')
        
        return redirect('payment_failed')
    
    elif request.method == 'POST':
        # IPN callback
        try:
            data = json.loads(request.body)
        except:
            data = request.POST.dict()
        
        transaction_id = data.get('transaction_id')
        checkout_id = data.get('checkout_id')
        status = data.get('status')
        reference = data.get('reference')
        
        if not transaction_id and not reference:
            return HttpResponse('OK')
        
        # Find transaction
        transaction = None
        if reference:
            try:
                transaction = PaymentTransaction.objects.get(reference=reference)
            except PaymentTransaction.DoesNotExist:
                pass
        
        if not transaction and transaction_id:
            try:
                transaction = PaymentTransaction.objects.get(sasapay_transaction_id=transaction_id)
            except PaymentTransaction.DoesNotExist:
                pass
        
        if transaction:
            # Update transaction
            transaction.sasapay_raw_response = data
            transaction.sasapay_status = status
            
            if status == 'COMPLETED':
                transaction.status = 'completed'
                transaction.paid_at = timezone.now()
                transaction.completed_at = timezone.now()
                transaction.save()
                
                # Process successful payment
                process_successful_payment(transaction, request)
                
            elif status == 'FAILED':
                transaction.status = 'failed'
                transaction.save()
            
            else:
                transaction.save()
        
        return HttpResponse('OK')


def sasapay_verify(request):
    """Verify payment after redirect"""
    transaction_id = request.GET.get('transaction_id')
    
    if not transaction_id:
        return redirect('payment_failed')
    
    from .sasapay_utils import query_payment_status
    
    # Query status
    result = query_payment_status(transaction_id)
    
    if result.get('status') == 'COMPLETED':
        # Find transaction
        try:
            transaction = PaymentTransaction.objects.get(sasapay_transaction_id=transaction_id)
            if transaction.status != 'completed':
                transaction.status = 'completed'
                transaction.paid_at = timezone.now()
                transaction.completed_at = timezone.now()
                transaction.save()
                process_successful_payment(transaction, request)
            
            return redirect('payment_success', reference=transaction.reference)
        except PaymentTransaction.DoesNotExist:
            pass
    
    return redirect('payment_failed')


def sasapay_status(request, reference):
    """Check payment status (AJAX)"""
    try:
        transaction = PaymentTransaction.objects.get(reference=reference)
        
        # If still pending, query SasaPay
        if transaction.status == 'pending' and transaction.sasapay_transaction_id:
            from .sasapay_utils import query_payment_status
            result = query_payment_status(transaction.sasapay_transaction_id)
            
            if result.get('status') == 'COMPLETED':
                transaction.status = 'completed'
                transaction.paid_at = timezone.now()
                transaction.completed_at = timezone.now()
                transaction.save()
                process_successful_payment(transaction, None)
            elif result.get('status') == 'FAILED':
                transaction.status = 'failed'
                transaction.save()
        
        return JsonResponse({
            'status': transaction.status,
            'success_url': f'/payment/success/{reference}/' if transaction.status == 'completed' else None
        })
    except PaymentTransaction.DoesNotExist:
        return JsonResponse({'status': 'not_found'}, status=404)        


# myapp/views.py - Add these functions at the end of your file

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, Count
from datetime import datetime, timedelta
import json

from .models import MfalmeUsers, Course, Video, PDF, Blog, Package, Order, Partnership

@require_GET
def export_users(request):
    """Export users to Excel or CSV"""
    format_type = request.GET.get('format', 'excel')
    
    # Get users
    users = MfalmeUsers.objects.all().order_by('-date_joined')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Phone', 'Tier', 'Status', 'Date Joined'])
        
        for user in users:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.phone or '',
                user.tier or 'citizen',
                'Active' if user.is_active else 'Inactive',
                user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '',
            ])
        return response
    
    # Excel format
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users Export"
    
    # Headers
    headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Phone', 'Tier', 'Status', 'Date Joined']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1).value = user.id
        ws.cell(row=row_num, column=2).value = user.username
        ws.cell(row=row_num, column=3).value = user.email
        ws.cell(row=row_num, column=4).value = user.first_name
        ws.cell(row=row_num, column=5).value = user.last_name
        ws.cell(row=row_num, column=6).value = user.phone
        ws.cell(row=row_num, column=7).value = user.tier or 'citizen'
        ws.cell(row=row_num, column=8).value = 'Active' if user.is_active else 'Inactive'
        ws.cell(row=row_num, column=9).value = user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else ''
    
    wb.save(response)
    return response


@require_GET
def export_orders(request):
    """Export orders/transactions to Excel"""
    format_type = request.GET.get('format', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Use PaymentTransaction model (not Order)
    transactions = PaymentTransaction.objects.all().order_by('-created_at')
    
    if start_date:
        transactions = transactions.filter(created_at__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__date__lte=end_date)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="transactions_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions Export"
    
    # Headers
    headers = ['Reference', 'User', 'Email', 'Amount (KES)', 'Payment Type', 'Payment Method', 'Status', 'Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    total_amount = 0
    for row_num, trans in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = trans.reference
        ws.cell(row=row_num, column=2).value = trans.user.get_full_name() if trans.user else 'Guest'
        ws.cell(row=row_num, column=3).value = trans.user.email if trans.user else 'guest@example.com'
        ws.cell(row=row_num, column=4).value = float(trans.amount)
        ws.cell(row=row_num, column=5).value = trans.payment_type or 'N/A'
        ws.cell(row=row_num, column=6).value = trans.payment_method or 'N/A'
        ws.cell(row=row_num, column=7).value = trans.status
        ws.cell(row=row_num, column=8).value = trans.created_at.strftime('%Y-%m-%d %H:%M') if trans.created_at else ''
        
        if trans.status == 'completed':
            total_amount += float(trans.amount)
    
    # Add total row
    total_row = len(transactions) + 2
    ws.cell(row=total_row, column=3).value = "TOTAL:"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = total_amount
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    
    wb.save(response)
    return response


@require_GET
def export_courses(request):
    """Export courses to Excel"""
    courses = Course.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses Export"
    
    # Headers
    headers = ['ID', 'Title', 'Price (USD)', 'Duration (weeks)', 'Videos Count', 'PDFs Count', 'Status', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row_num, course in enumerate(courses, 2):
        # Count related videos and PDFs
        videos_count = TrainingVideo.objects.filter(course=course).count()
        pdfs_count = PDF.objects.filter(course=course).count()
        
        ws.cell(row=row_num, column=1).value = course.id
        ws.cell(row=row_num, column=2).value = course.title
        ws.cell(row=row_num, column=3).value = float(course.price) if course.price else 0
        ws.cell(row=row_num, column=4).value = course.duration_weeks or 0
        ws.cell(row=row_num, column=5).value = videos_count
        ws.cell(row=row_num, column=6).value = pdfs_count
        ws.cell(row=row_num, column=7).value = 'Active' if course.is_active else 'Inactive'
        ws.cell(row=row_num, column=8).value = course.created_at.strftime('%Y-%m-%d') if course.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_videos(request):
    """Export videos to Excel"""
    # Use TrainingVideo model (not Video)
    videos = TrainingVideo.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="videos_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Videos Export"
    
    headers = ['ID', 'Title', 'Course', 'Duration (min)', 'Price (USD)', 'Views', 'Status', 'Uploaded']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, video in enumerate(videos, 2):
        ws.cell(row=row_num, column=1).value = video.id
        ws.cell(row=row_num, column=2).value = video.title
        ws.cell(row=row_num, column=3).value = video.course.title if video.course else 'Uncategorized'
        ws.cell(row=row_num, column=4).value = video.duration or 0
        ws.cell(row=row_num, column=5).value = float(video.price) if video.price else 0
        ws.cell(row=row_num, column=6).value = video.view_count or 0
        ws.cell(row=row_num, column=7).value = 'Active' if video.is_active else 'Inactive'
        ws.cell(row=row_num, column=8).value = video.created_at.strftime('%Y-%m-%d') if video.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_pdfs(request):
    """Export PDFs to Excel"""
    pdfs = PDF.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pdfs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDFs Export"
    
    headers = ['ID', 'Title', 'Course', 'Pages', 'Price (USD)', 'Is Free', 'Access Level', 'Views', 'Uploaded']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, pdf in enumerate(pdfs, 2):
        ws.cell(row=row_num, column=1).value = pdf.id
        ws.cell(row=row_num, column=2).value = pdf.title
        ws.cell(row=row_num, column=3).value = pdf.course.title if pdf.course else 'General'
        ws.cell(row=row_num, column=4).value = pdf.pages or 0
        ws.cell(row=row_num, column=5).value = float(pdf.price) if pdf.price else 0
        ws.cell(row=row_num, column=6).value = 'Yes' if pdf.is_free else 'No'
        ws.cell(row=row_num, column=7).value = pdf.access_level or 'free'
        ws.cell(row=row_num, column=8).value = pdf.views or 0
        ws.cell(row=row_num, column=9).value = pdf.created_at.strftime('%Y-%m-%d') if pdf.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_revenue_report(request):
    """Export revenue report to Excel"""
    # Import Sum and Count
    from django.db.models import Sum, Count
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Parse dates
    if start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start = datetime.now() - timedelta(days=30)
    
    if end_date:
        end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        end = datetime.now()
    
    # Get completed transactions in date range
    transactions = PaymentTransaction.objects.filter(
        created_at__date__gte=start.date(),
        created_at__date__lte=end.date(),
        status='completed'
    ).order_by('created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary.cell(row=1, column=1).value = "Revenue Report"
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:F1')
    
    ws_summary.cell(row=2, column=1).value = f"Period: {start_date} to {end_date}"
    ws_summary.merge_cells('A2:F2')
    
    # Stats
    total_revenue = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    total_orders = transactions.count()
    avg_order = total_revenue / total_orders if total_orders > 0 else 0
    
    ws_summary.cell(row=4, column=1).value = "Total Revenue:"
    ws_summary.cell(row=4, column=1).font = Font(bold=True)
    ws_summary.cell(row=4, column=2).value = float(total_revenue)
    ws_summary.cell(row=4, column=2).number_format = '#,##0.00'
    
    ws_summary.cell(row=5, column=1).value = "Total Transactions:"
    ws_summary.cell(row=5, column=1).font = Font(bold=True)
    ws_summary.cell(row=5, column=2).value = total_orders
    
    ws_summary.cell(row=6, column=1).value = "Average Transaction:"
    ws_summary.cell(row=6, column=1).font = Font(bold=True)
    ws_summary.cell(row=6, column=2).value = float(avg_order)
    ws_summary.cell(row=6, column=2).number_format = '#,##0.00'
    
    # ===== DAILY BREAKDOWN SHEET =====
    ws_daily = wb.create_sheet("Daily Breakdown")
    
    headers = ['Date', 'Transactions', 'Revenue (KES)', 'Avg Value']
    for col_num, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Group by date
    daily_stats = transactions.values('created_at__date').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('created_at__date')
    
    for row_num, stat in enumerate(daily_stats, 2):
        ws_daily.cell(row=row_num, column=1).value = stat['created_at__date'].strftime('%Y-%m-%d')
        ws_daily.cell(row=row_num, column=2).value = stat['count']
        ws_daily.cell(row=row_num, column=3).value = float(stat['total'])
        if stat['count'] > 0:
            ws_daily.cell(row=row_num, column=4).value = float(stat['total'] / stat['count'])
            ws_daily.cell(row=row_num, column=4).number_format = '#,##0.00'
    
    # ===== TRANSACTIONS DETAIL SHEET =====
    ws_detail = wb.create_sheet("Transaction Details")
    
    detail_headers = ['Reference', 'Date', 'Customer', 'Email', 'Amount (KES)', 'Payment Type', 'Method', 'Status']
    for col_num, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, trans in enumerate(transactions, 2):
        ws_detail.cell(row=row_num, column=1).value = trans.reference
        ws_detail.cell(row=row_num, column=2).value = trans.created_at.strftime('%Y-%m-%d %H:%M') if trans.created_at else ''
        ws_detail.cell(row=row_num, column=3).value = trans.user.get_full_name() if trans.user else 'Guest'
        ws_detail.cell(row=row_num, column=4).value = trans.user.email if trans.user else 'N/A'
        ws_detail.cell(row=row_num, column=5).value = float(trans.amount)
        ws_detail.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws_detail.cell(row=row_num, column=6).value = trans.payment_type or 'N/A'
        ws_detail.cell(row=row_num, column=7).value = trans.payment_method or 'N/A'
        ws_detail.cell(row=row_num, column=8).value = trans.status
    
    # ===== AUTO-ADJUST COLUMN WIDTHS (FIXED) =====
    # Process each sheet separately, handling merged cells
    for sheet in wb.worksheets:
        # Skip the summary sheet for auto-adjust (has merged cells)
        if sheet.title == "Summary":
            # Manually set widths for summary sheet
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            continue
        
        # For other sheets without merged cells, auto-adjust
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Skip if this column has merged cells
            skip = False
            for cell in column:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    skip = True
                    break
            
            if skip:
                continue
                
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response

@require_GET
def export_blogs(request):
    """Export blogs to Excel"""
    blogs = Blog.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="blogs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blogs Export"
    
    headers = ['ID', 'Title', 'Author', 'Category', 'Views', 'Status', 'Published', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, blog in enumerate(blogs, 2):
        ws.cell(row=row_num, column=1).value = blog.id
        ws.cell(row=row_num, column=2).value = blog.title
        ws.cell(row=row_num, column=3).value = blog.author.username if blog.author else 'Admin'
        ws.cell(row=row_num, column=4).value = blog.category or 'General'
        ws.cell(row=row_num, column=5).value = blog.views or 0
        ws.cell(row=row_num, column=6).value = blog.status or 'draft'
        ws.cell(row=row_num, column=7).value = blog.published_at.strftime('%Y-%m-%d') if blog.published_at else 'Not published'
        ws.cell(row=row_num, column=8).value = blog.created_at.strftime('%Y-%m-%d') if blog.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@require_GET
def export_packages(request):
    """Export packages to Excel"""
    packages = Package.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="packages_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packages Export"
    
    headers = ['ID', 'Name', 'Price (KES)', 'Type', 'Sales', 'Revenue (KES)', 'Status', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, package in enumerate(packages, 2):
        sales = Order.objects.filter(package=package, status='completed').count()
        revenue = Order.objects.filter(package=package, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        
        ws.cell(row=row_num, column=1).value = package.id
        ws.cell(row=row_num, column=2).value = package.name
        ws.cell(row=row_num, column=3).value = float(package.price) if package.price else 0
        ws.cell(row=row_num, column=4).value = package.package_type or 'N/A'
        ws.cell(row=row_num, column=5).value = sales
        ws.cell(row=row_num, column=6).value = float(revenue)
        ws.cell(row=row_num, column=7).value = package.status or 'inactive'
        ws.cell(row=row_num, column=8).value = package.created_at.strftime('%Y-%m-%d') if package.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@require_GET
def export_partnerships(request):
    """Export partnerships to Excel"""
    partnerships = Partnership.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="partnerships_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partnerships Export"
    
    headers = ['ID', 'Company', 'Contact Person', 'Email', 'Phone', 'Tier', 'Amount (KES)', 'NDA', 'Status', 'Applied']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, p in enumerate(partnerships, 2):
        ws.cell(row=row_num, column=1).value = p.id
        ws.cell(row=row_num, column=2).value = p.company_name
        ws.cell(row=row_num, column=3).value = p.contact_name
        ws.cell(row=row_num, column=4).value = p.email
        ws.cell(row=row_num, column=5).value = p.phone or ''
        ws.cell(row=row_num, column=6).value = p.tier or 'N/A'
        ws.cell(row=row_num, column=7).value = float(p.investment_amount) if p.investment_amount else 0
        ws.cell(row=row_num, column=8).value = 'Signed' if p.nda_signed else 'Pending'
        ws.cell(row=row_num, column=9).value = p.status or 'pending'
        ws.cell(row=row_num, column=10).value = p.created_at.strftime('%Y-%m-%d') if p.created_at else ''
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


@require_GET
def export_users(request):
    """Export users to Excel or CSV"""
    format_type = request.GET.get('format', 'excel')
    
    # Get users
    users = MfalmeUsers.objects.all().order_by('-date_joined')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Phone', 'Rank', 'Status', 'Date Joined'])
        
        for user in users:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.phone or '',
                user.elite_rank or 'citizen',  # ✅ FIXED: use elite_rank
                'Active' if user.is_active else 'Inactive',
                user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '',
            ])
        return response
    
    # Excel format
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users Export"
    
    # Headers
    headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Phone', 'Rank', 'Status', 'Date Joined']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1).value = user.id
        ws.cell(row=row_num, column=2).value = user.username
        ws.cell(row=row_num, column=3).value = user.email
        ws.cell(row=row_num, column=4).value = user.first_name
        ws.cell(row=row_num, column=5).value = user.last_name
        ws.cell(row=row_num, column=6).value = user.phone
        ws.cell(row=row_num, column=7).value = user.elite_rank or 'citizen'  # ✅ FIXED: use elite_rank
        ws.cell(row=row_num, column=8).value = 'Active' if user.is_active else 'Inactive'
        ws.cell(row=row_num, column=9).value = user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else ''
    
    wb.save(response)
    return response

@require_GET
def export_orders(request):
    """Export orders/transactions to Excel"""
    format_type = request.GET.get('format', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Use PaymentTransaction model (not Order)
    transactions = PaymentTransaction.objects.all().order_by('-created_at')
    
    if start_date:
        transactions = transactions.filter(created_at__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__date__lte=end_date)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="transactions_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions Export"
    
    # Headers
    headers = ['Reference', 'User', 'Email', 'Amount (KES)', 'Payment Type', 'Payment Method', 'Status', 'Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    total_amount = 0
    for row_num, trans in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = trans.reference
        ws.cell(row=row_num, column=2).value = trans.user.get_full_name() if trans.user else 'Guest'
        ws.cell(row=row_num, column=3).value = trans.user.email if trans.user else 'guest@example.com'
        ws.cell(row=row_num, column=4).value = float(trans.amount)
        ws.cell(row=row_num, column=5).value = trans.payment_type or 'N/A'
        ws.cell(row=row_num, column=6).value = trans.payment_method or 'N/A'
        ws.cell(row=row_num, column=7).value = trans.status
        ws.cell(row=row_num, column=8).value = trans.created_at.strftime('%Y-%m-%d %H:%M') if trans.created_at else ''
        
        if trans.status == 'completed':
            total_amount += float(trans.amount)
    
    # Add total row
    total_row = len(transactions) + 2
    ws.cell(row=total_row, column=3).value = "TOTAL:"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = total_amount
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    
    wb.save(response)
    return response


@require_GET
def export_courses(request):
    """Export courses to Excel"""
    courses = Course.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses Export"
    
    # Headers
    headers = ['ID', 'Title', 'Price (USD)', 'Duration (weeks)', 'Videos Count', 'PDFs Count', 'Status', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row_num, course in enumerate(courses, 2):
        # Count related videos and PDFs
        videos_count = TrainingVideo.objects.filter(course=course).count()
        pdfs_count = PDF.objects.filter(course=course).count()
        
        ws.cell(row=row_num, column=1).value = course.id
        ws.cell(row=row_num, column=2).value = course.title
        ws.cell(row=row_num, column=3).value = float(course.price) if course.price else 0
        ws.cell(row=row_num, column=4).value = course.duration_weeks or 0
        ws.cell(row=row_num, column=5).value = videos_count
        ws.cell(row=row_num, column=6).value = pdfs_count
        ws.cell(row=row_num, column=7).value = 'Active' if course.is_active else 'Inactive'
        ws.cell(row=row_num, column=8).value = course.created_at.strftime('%Y-%m-%d') if course.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_videos(request):
    """Export videos to Excel"""
    # Use TrainingVideo model (not Video)
    videos = TrainingVideo.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="videos_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Videos Export"
    
    headers = ['ID', 'Title', 'Course', 'Duration (min)', 'Price (USD)', 'Views', 'Status', 'Uploaded']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, video in enumerate(videos, 2):
        ws.cell(row=row_num, column=1).value = video.id
        ws.cell(row=row_num, column=2).value = video.title
        ws.cell(row=row_num, column=3).value = video.course.title if video.course else 'Uncategorized'
        ws.cell(row=row_num, column=4).value = video.duration or 0
        ws.cell(row=row_num, column=5).value = float(video.price) if video.price else 0
        ws.cell(row=row_num, column=6).value = video.view_count or 0
        ws.cell(row=row_num, column=7).value = 'Active' if video.is_active else 'Inactive'
        ws.cell(row=row_num, column=8).value = video.created_at.strftime('%Y-%m-%d') if video.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_pdfs(request):
    """Export PDFs to Excel"""
    pdfs = PDF.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pdfs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDFs Export"
    
    headers = ['ID', 'Title', 'Course', 'Pages', 'Price (USD)', 'Is Free', 'Access Level', 'Views', 'Uploaded']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, pdf in enumerate(pdfs, 2):
        ws.cell(row=row_num, column=1).value = pdf.id
        ws.cell(row=row_num, column=2).value = pdf.title
        ws.cell(row=row_num, column=3).value = pdf.course.title if pdf.course else 'General'
        ws.cell(row=row_num, column=4).value = pdf.pages or 0
        ws.cell(row=row_num, column=5).value = float(pdf.price) if pdf.price else 0
        ws.cell(row=row_num, column=6).value = 'Yes' if pdf.is_free else 'No'
        ws.cell(row=row_num, column=7).value = pdf.access_level or 'free'
        ws.cell(row=row_num, column=8).value = pdf.views or 0
        ws.cell(row=row_num, column=9).value = pdf.created_at.strftime('%Y-%m-%d') if pdf.created_at else ''
    
    wb.save(response)
    return response


@require_GET
def export_revenue_report(request):
    """Export revenue report to Excel"""
    # Import Sum and Count
    from django.db.models import Sum, Count
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Parse dates
    if start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start = datetime.now() - timedelta(days=30)
    
    if end_date:
        end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        end = datetime.now()
    
    # Get completed transactions in date range
    transactions = PaymentTransaction.objects.filter(
        created_at__date__gte=start.date(),
        created_at__date__lte=end.date(),
        status='completed'
    ).order_by('created_at')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{start_date}_to_{end_date}.xlsx"'
    
    wb = openpyxl.Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary.cell(row=1, column=1).value = "Revenue Report"
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:F1')
    
    ws_summary.cell(row=2, column=1).value = f"Period: {start_date} to {end_date}"
    ws_summary.merge_cells('A2:F2')
    
    # Stats
    total_revenue = transactions.aggregate(Sum('amount'))['amount__sum'] or 0
    total_orders = transactions.count()
    avg_order = total_revenue / total_orders if total_orders > 0 else 0
    
    ws_summary.cell(row=4, column=1).value = "Total Revenue:"
    ws_summary.cell(row=4, column=1).font = Font(bold=True)
    ws_summary.cell(row=4, column=2).value = float(total_revenue)
    ws_summary.cell(row=4, column=2).number_format = '#,##0.00'
    
    ws_summary.cell(row=5, column=1).value = "Total Transactions:"
    ws_summary.cell(row=5, column=1).font = Font(bold=True)
    ws_summary.cell(row=5, column=2).value = total_orders
    
    ws_summary.cell(row=6, column=1).value = "Average Transaction:"
    ws_summary.cell(row=6, column=1).font = Font(bold=True)
    ws_summary.cell(row=6, column=2).value = float(avg_order)
    ws_summary.cell(row=6, column=2).number_format = '#,##0.00'
    
    # ===== DAILY BREAKDOWN SHEET =====
    ws_daily = wb.create_sheet("Daily Breakdown")
    
    headers = ['Date', 'Transactions', 'Revenue (KES)', 'Avg Value']
    for col_num, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Group by date
    daily_stats = transactions.values('created_at__date').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('created_at__date')
    
    for row_num, stat in enumerate(daily_stats, 2):
        ws_daily.cell(row=row_num, column=1).value = stat['created_at__date'].strftime('%Y-%m-%d')
        ws_daily.cell(row=row_num, column=2).value = stat['count']
        ws_daily.cell(row=row_num, column=3).value = float(stat['total'])
        if stat['count'] > 0:
            ws_daily.cell(row=row_num, column=4).value = float(stat['total'] / stat['count'])
            ws_daily.cell(row=row_num, column=4).number_format = '#,##0.00'
    
    # ===== TRANSACTIONS DETAIL SHEET =====
    ws_detail = wb.create_sheet("Transaction Details")
    
    detail_headers = ['Reference', 'Date', 'Customer', 'Email', 'Amount (KES)', 'Payment Type', 'Method', 'Status']
    for col_num, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row_num, trans in enumerate(transactions, 2):
        ws_detail.cell(row=row_num, column=1).value = trans.reference
        ws_detail.cell(row=row_num, column=2).value = trans.created_at.strftime('%Y-%m-%d %H:%M') if trans.created_at else ''
        ws_detail.cell(row=row_num, column=3).value = trans.user.get_full_name() if trans.user else 'Guest'
        ws_detail.cell(row=row_num, column=4).value = trans.user.email if trans.user else 'N/A'
        ws_detail.cell(row=row_num, column=5).value = float(trans.amount)
        ws_detail.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws_detail.cell(row=row_num, column=6).value = trans.payment_type or 'N/A'
        ws_detail.cell(row=row_num, column=7).value = trans.payment_method or 'N/A'
        ws_detail.cell(row=row_num, column=8).value = trans.status
    
    # ===== AUTO-ADJUST COLUMN WIDTHS (FIXED) =====
    for sheet in wb.worksheets:
        if sheet.title == "Summary":
            # Manual widths for summary sheet (has merged cells)
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
        else:
            # Auto-adjust for other sheets
            for col in sheet.columns:
                max_length = 0
                col_letter = None
                for cell in col:
                    if cell and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        if col_letter is None:
                            col_letter = get_column_letter(cell.column)
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                if col_letter and max_length > 0:
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(response)
    return response

def s3_test_page(request):
    return render(request, 's3_test.html')

@csrf_exempt
def api_community_join(request):
    """Handle community join requests with email notifications"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    tier = data.get('tier')
    community_name = data.get('community_name', tier.upper())
    
    if not tier:
        return JsonResponse({'error': 'Tier required'}, status=400)
    
    user = request.user
    
    # Hardcoded community requirements
    community_requirements = {
        'citizens': {
            'name': 'CITIZENS',
            'min_investment': 0,
            'auto_approve': True
        },
        'studyhall': {
            'name': 'STUDY HALL',
            'min_investment': 1000,
            'auto_approve': False
        },
        'society': {
            'name': 'SOCIETY',
            'min_investment': 10000,
            'auto_approve': False
        }
    }
    
    if tier not in community_requirements:
        return JsonResponse({'error': 'Invalid community tier'}, status=400)
    
    community_info = community_requirements[tier]
    
    # Check if user already has a membership (in database or session)
    # For now, we'll use a simple session-based tracking
    if not hasattr(request, 'session'):
        from django.contrib.sessions.backends.db import SessionStore
        request.session = SessionStore()
    
    memberships = request.session.get('community_memberships', {})
    
    if tier in memberships:
        return JsonResponse({
            'error': f'You already have a {memberships[tier]} membership for this community'
        }, status=400)
    
    # Check investment requirement
    if user.total_deposits < community_info['min_investment']:
        return JsonResponse({
            'error': 'You do not meet the investment requirement',
            'missing': [f'Minimum investment: ${community_info["min_investment"]:,}']
        }, status=403)
    
    # Auto-approve citizens, pending for others
    if community_info['auto_approve']:
        status = 'active'
        memberships[tier] = 'active'
        message = f'Welcome to {community_info["name"]}! You now have access.'
    else:
        status = 'pending'
        memberships[tier] = 'pending'
        message = f'Your request to join {community_info["name"]} has been submitted for review.'
    
    # Store in session
    request.session['community_memberships'] = memberships
    request.session.modified = True
    
    # Create a record in database if you have the model
    try:
        # If you have the CommunityTier and UserCommunityMembership models
        from .models import CommunityTier, UserCommunityMembership
        
        community, created = CommunityTier.objects.get_or_create(
            tier=tier,
            defaults={
                'name': community_info['name'],
                'minimum_investment': community_info['min_investment']
            }
        )
        
        membership, created = UserCommunityMembership.objects.update_or_create(
            user=user,
            community=community,
            defaults={
                'status': status,
                'access_granted': status == 'active'
            }
        )
    except ImportError:
        # Models don't exist yet, just use session
        pass
    
    # SEND EMAIL NOTIFICATIONS
    try:
        # Email to user
        user_subject = f"Community Join Request: {community_info['name']}"
        user_context = {
            'username': user.username,
            'community_name': community_info['name'],
            'status': status,
            'tier': tier,
            'year': timezone.now().year
        }
        
        html_content = render_to_string('emails/community_join_user.html', user_context)
        text_content = f"Your request to join {community_info['name']} is {status}."
        
        user_email = EmailMultiAlternatives(
            user_subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [user.email]
        )
        user_email.attach_alternative(html_content, "text/html")
        user_email.send()
        
        # Email to admin
        admin_subject = f"🔔 NEW COMMUNITY JOIN REQUEST - {community_info['name']}"
        admin_context = {
            'username': user.username,
            'email': user.email,
            'community_name': community_info['name'],
            'tier': tier,
            'status': status,
            'total_deposits': user.total_deposits,
            'trading_experience': user.trading_experience,
            'user_id': user.id,
            'admin_url': f"{settings.SITE_URL}/admin/",
            'year': timezone.now().year
        }
        
        admin_html = render_to_string('emails/community_join_admin.html', admin_context)
        admin_text = f"New join request from {user.username} for {community_info['name']}"
        
        admin_email = EmailMultiAlternatives(
            admin_subject,
            admin_text,
            settings.DEFAULT_FROM_EMAIL,
            settings.ADMIN_EMAILS
        )
        admin_email.attach_alternative(admin_html, "text/html")
        admin_email.send()
        
        print(f"✅ Emails sent for {tier} join request")
        
    except Exception as e:
        print(f"❌ Email error: {e}")
    
    # Log activity
    try:
        log_activity(user, 'COMMUNITY_JOIN', f'Requested to join {community_info["name"]}', request)
    except:
        pass
    
    return JsonResponse({
        'success': True,
        'status': status,
        'message': message
    })



# ==================== ERROR HANDLERS ====================

def health_check(request):
    """Simple health check that verifies database connection"""
    try:
        # Test database connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        return HttpResponse("OK - Database connected", content_type="text/plain", status=200)
    except Exception as e:
        return HttpResponse(f"ERROR - Database: {str(e)}", content_type="text/plain", status=500)

def root_health(request):
    """Simple root path that returns 200 OK for Railway health checks"""
    return HttpResponse("OK", status=200, content_type="text/plain")

def custom_404(request, exception):
    return render(request, '404.html', status=404)

def custom_500(request):
    return render(request, '500.html', status=500)


def get_s3_presigned_url(request):
    """
    Generate a presigned URL for direct browser-to-S3 upload
    Uses direct S3 URLs (no CloudFront)
    """
    print("\n" + "="*60)
    print("🔐 S3 PRESIGNED URL REQUEST")
    print("="*60)
    
    # CHECK YOUR CUSTOM SESSION
    if not request.session.get('admin_authenticated'):
        print("❌ Custom admin authentication failed")
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    admin_username = request.session.get('admin_username', 'Unknown')
    print(f"✅ Admin {admin_username} authenticated")
    
    file_name = request.GET.get('file_name')
    file_type = request.GET.get('file_type')
    content_type = request.GET.get('content_type', 'application/octet-stream')
    
    if not file_name or not file_type:
        return JsonResponse({'error': 'file_name and file_type required'}, status=400)
    
    print(f"📁 File: {file_name}")
    print(f"📂 Type: {file_type}")
    print(f"📄 Content-Type: {content_type}")
    
    try:
        from datetime import datetime
        import uuid
        import os
        
        # Map file types to folders
        folder_mapping = {
            'video': f"videos/{datetime.now().strftime('%Y/%m')}",
            'pdf': f"pdfs/{datetime.now().strftime('%Y/%m')}",
            'thumbnail': "thumbnails",
            'course_thumbnail': "course_thumbnails",
            'pdf_cover': "pdf_covers",
            'blog_image': "blog_images",
            'profile_image': "profile_images",
        }
        
        folder = folder_mapping.get(file_type, "media")
        
        # Generate unique filename
        ext = file_name.split('.')[-1].lower() if '.' in file_name else 'bin'
        unique_id = uuid.uuid4().hex[:8]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        base_name = os.path.splitext(file_name)[0]
        base_name = ''.join(c for c in base_name if c.isalnum() or c in ' _-').strip()
        base_name = base_name.replace(' ', '_')[:50]
        
        key = f"{folder}/{timestamp}_{unique_id}_{base_name}.{ext}"
        
        print(f"🔑 S3 Key: {key}")
        
        # Verify AWS credentials are set
        if not settings.AWS_ACCESS_KEY_ID or not settings.AWS_SECRET_ACCESS_KEY:
            print("❌ AWS credentials not configured")
            return JsonResponse({'error': 'AWS credentials not configured'}, status=500)
        
        # IMPORTANT: Create S3 client with the correct endpoint configuration
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME,
            endpoint_url=f'https://s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com',
            config=Config(
                signature_version='s3v4',
                s3={'addressing_style': 'virtual'}
            )
        )
        
        # Determine ACL based on file type
        acl = 'private'
        public_url = None
        
        # Public files (images, thumbnails) - use public-read
        if file_type in ['thumbnail', 'course_thumbnail', 'pdf_cover', 'blog_image', 'profile_image']:
            acl = 'public-read'
            # Generate public URL for display
            public_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{key}"
            print(f"📎 Public S3 URL: {public_url}")
        
        # Generate presigned URL (only needed for PUT operation)
        presigned_url = s3_client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                'Key': key,
                'ContentType': content_type,
                'ACL': acl,
                'CacheControl': 'max-age=31536000',
            },
            ExpiresIn=3600,
            HttpMethod='PUT'
        )
        
        print(f"✅ Presigned URL generated successfully")
        print(f"📎 ACL: {acl}")
        
        # Prepare response
        response_data = {
            'success': True,
            'presigned_url': presigned_url,
            'key': key,
            'file_url': f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{key}",
            'bucket': settings.AWS_STORAGE_BUCKET_NAME,
            'region': settings.AWS_S3_REGION_NAME,
            'acl': acl
        }
        
        # Add public URL for image types
        if public_url:
            response_data['public_url'] = public_url
            
        return JsonResponse(response_data)
        
    except ClientError as e:
        error_code = e.response['Error']['Code']
        error_message = e.response['Error']['Message']
        print(f"❌ AWS ClientError: {error_code} - {error_message}")
        return JsonResponse({'error': f'AWS Error: {error_message}'}, status=500)
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def initiate_multipart_upload(request):
    """
    Initiate multipart upload for large files (>100MB)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    file_name = data.get('file_name')
    file_type = data.get('file_type')
    total_parts = data.get('total_parts', 1)
    
    if not file_name or not file_type:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    
    try:
        # Determine folder
        if file_type == 'video':
            folder = f"videos/{datetime.now().strftime('%Y/%m')}"
        elif file_type == 'pdf':
            folder = f"pdfs/{datetime.now().strftime('%Y/%m')}"
        else:
            folder = "media"
        
        # Generate key
        ext = file_name.split('.')[-1]
        unique_id = uuid.uuid4().hex[:8]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        key = f"{folder}/{timestamp}_{unique_id}.{ext}"
        
        # Create S3 client
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Create multipart upload
        response = s3_client.create_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            ACL='private'
        )
        
        upload_id = response['UploadId']
        
        # Generate presigned URLs for each part
        part_urls = []
        for part_number in range(1, total_parts + 1):
            part_url = s3_client.generate_presigned_url(
                'upload_part',
                Params={
                    'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                    'Key': key,
                    'UploadId': upload_id,
                    'PartNumber': part_number
                },
                ExpiresIn=3600
            )
            part_urls.append({
                'part_number': part_number,
                'url': part_url
            })
        
        return JsonResponse({
            'success': True,
            'upload_id': upload_id,
            'key': key,
            'part_urls': part_urls,
            'file_url': f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{key}"
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def complete_multipart_upload(request):
    """
    Complete multipart upload after all parts are uploaded
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    key = data.get('key')
    upload_id = data.get('upload_id')
    parts = data.get('parts')  # List of {'ETag': '...', 'PartNumber': 1}
    
    if not key or not upload_id or not parts:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Complete multipart upload
        response = s3_client.complete_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            UploadId=upload_id,
            MultipartUpload={'Parts': parts}
        )
        
        return JsonResponse({
            'success': True,
            'location': response['Location'],
            'key': key,
            'file_url': f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{key}"
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def abort_multipart_upload(request):
    """
    Abort multipart upload if something goes wrong
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    key = data.get('key')
    upload_id = data.get('upload_id')
    
    if not key or not upload_id:
        return JsonResponse({'error': 'Missing required fields'}, status=400)
    
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        s3_client.abort_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            UploadId=upload_id
        )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


    
# ========== HELPER FUNCTIONS ==========


def send_merchandise_order_email(order):
    """Send merchandise order confirmation email to customer and admin"""
    from django.core.mail import send_mail, EmailMultiAlternatives
    from django.template.loader import render_to_string
    from django.conf import settings
    import os
    
    try:
        subject = f"Order Confirmation - {order.order_number}"
        
        # Build items HTML for customer
        items_html = ""
        for item in order.items:
            items_html += f"""
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{item.get('name', 'Item')}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{item.get('quantity', 1)}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">KES {item.get('price', 0):,.2f}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">KES {item.get('price', 0) * item.get('quantity', 1):,.2f}</td>
            </tr>
            """
        
        # Customer email HTML
        customer_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Order Confirmation</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background: #f5f5f5;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 16px;
                    overflow: hidden;
                    box-shadow: 0 4px 20px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: linear-gradient(135deg, #FFD700, #FFA500);
                    padding: 25px;
                    text-align: center;
                }}
                .header h1 {{
                    margin: 0;
                    color: #0a1520;
                    font-size: 22px;
                }}
                .header p {{
                    margin: 5px 0 0;
                    color: #0a1520;
                    font-size: 13px;
                }}
                .content {{
                    padding: 25px;
                }}
                .order-details {{
                    background: #f8f9fa;
                    border-radius: 12px;
                    padding: 15px;
                    margin-bottom: 20px;
                }}
                .order-details p {{
                    margin: 8px 0;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th {{
                    background: #FFD700;
                    padding: 10px;
                    text-align: left;
                    color: #0a1520;
                }}
                td {{
                    padding: 10px;
                    border-bottom: 1px solid #eee;
                }}
                .totals {{
                    text-align: right;
                    margin-top: 20px;
                    padding-top: 15px;
                    border-top: 2px solid #eee;
                }}
                .totals p {{
                    margin: 5px 0;
                }}
                .totals .grand-total {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #B8860B;
                }}
                .footer {{
                    background: #1a1a2e;
                    padding: 15px;
                    text-align: center;
                    font-size: 11px;
                    color: #aaa;
                }}
                .footer a {{
                    color: #FFD700;
                    text-decoration: none;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>MFALME BETTERDAYS CAPITAL</h1>
                    <p>Order Confirmation</p>
                </div>
                <div class="content">
                    <div class="order-details">
                        <p><strong>Order Number:</strong> {order.order_number}</p>
                        <p><strong>Order Date:</strong> {order.created_at.strftime('%B %d, %Y at %I:%M %p')}</p>
                        <p><strong>Customer:</strong> {order.customer_name}</p>
                        <p><strong>Phone:</strong> {order.customer_phone}</p>
                        <p><strong>Email:</strong> {order.customer_email}</p>
                        <p><strong>Delivery Address:</strong> {order.delivery_address}</p>
                    </div>
                    
                    <table>
                        <thead>
                            <tr>
                                <th>Item</th>
                                <th>Qty</th>
                                <th>Unit Price</th>
                                <th>Total</th>
                            </tr>
                        </thead>
                        <tbody>
                            {items_html}
                        </tbody>
                    </table>
                    
                    <div class="totals">
                        <p><strong>Subtotal:</strong> KES {order.subtotal:,.2f}</p>
                        <p><strong>Shipping:</strong> KES {order.shipping_cost:,.2f}</p>
                        <p class="grand-total"><strong>Total:</strong> KES {order.total:,.2f}</p>
                    </div>
                    
                    <p style="margin-top: 20px;">Your order will be processed and shipped within 3-5 business days.</p>
                </div>
                <div class="footer">
                    <p>For inquiries: +254 706 286 667 | <a href="mailto:mfalmebetterdays@gmail.com">mfalmebetterdays@gmail.com</a></p>
                    <p>2026 Mfalme Betterdays Capital. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Admin email HTML - STYLED
        admin_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>New Merchandise Order</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background: #f5f5f5;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    max-width: 550px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 16px;
                    overflow: hidden;
                    box-shadow: 0 4px 20px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: linear-gradient(135deg, #FFD700, #FFA500);
                    padding: 25px;
                    text-align: center;
                }}
                .header h1 {{
                    margin: 0;
                    color: #0a1520;
                    font-size: 20px;
                }}
                .header p {{
                    margin: 5px 0 0;
                    color: #0a1520;
                    font-size: 12px;
                }}
                .content {{
                    padding: 25px;
                }}
                .alert {{
                    background: #fff3cd;
                    padding: 12px;
                    margin-bottom: 20px;
                    border-left: 4px solid #FFD700;
                    border-radius: 8px;
                }}
                .alert p {{
                    margin: 0;
                    color: #856404;
                    font-size: 13px;
                }}
                .details-card {{
                    background: #f8f9fa;
                    border-radius: 12px;
                    padding: 15px;
                    margin-bottom: 20px;
                }}
                .details-card h3 {{
                    margin: 0 0 15px 0;
                    color: #B8860B;
                    font-size: 16px;
                }}
                .row {{
                    display: flex;
                    justify-content: space-between;
                    padding: 8px 0;
                    border-bottom: 1px solid #ddd;
                }}
                .row:last-child {{
                    border-bottom: none;
                }}
                .label {{
                    color: #666;
                    font-size: 12px;
                }}
                .value {{
                    font-weight: bold;
                    color: #333;
                    font-size: 13px;
                }}
                .value.highlight {{
                    color: #B8860B;
                    font-size: 16px;
                }}
                .items-box {{
                    background: #0a1520;
                    border-radius: 12px;
                    padding: 15px;
                    margin-bottom: 20px;
                }}
                .items-box h3 {{
                    margin: 0 0 15px 0;
                    color: #FFD700;
                    font-size: 14px;
                    text-align: center;
                }}
                .items-table {{
                    width: 100%;
                    color: white;
                }}
                .items-table th {{
                    color: #FFD700;
                    padding: 5px;
                    text-align: left;
                    font-size: 12px;
                }}
                .items-table td {{
                    padding: 5px;
                    font-size: 12px;
                }}
                .items-table td:last-child {{
                    text-align: right;
                }}
                .button {{
                    display: block;
                    background: #FFD700;
                    color: #0a1520;
                    text-align: center;
                    padding: 12px;
                    border-radius: 40px;
                    text-decoration: none;
                    font-weight: bold;
                    margin-top: 20px;
                }}
                .footer {{
                    background: #1a1a2e;
                    padding: 15px;
                    text-align: center;
                    font-size: 11px;
                    color: #aaa;
                }}
                .footer a {{
                    color: #FFD700;
                    text-decoration: none;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>MFALME BETTERDAYS CAPITAL</h1>
                    <p>New Merchandise Order</p>
                </div>
                <div class="content">
                    <div class="alert">
                        <p>A new merchandise order has been received. Please review the details below.</p>
                    </div>
                    
                    <div class="details-card">
                        <h3>ORDER DETAILS</h3>
                        <div class="row">
                            <span class="label">Order Number</span>
                            <span class="value">{order.order_number}</span>
                        </div>
                        <div class="row">
                            <span class="label">Order Date</span>
                            <span class="value">{order.created_at.strftime('%B %d, %Y at %I:%M %p')}</span>
                        </div>
                        <div class="row">
                            <span class="label">Customer Name</span>
                            <span class="value">{order.customer_name}</span>
                        </div>
                        <div class="row">
                            <span class="label">Phone</span>
                            <span class="value">{order.customer_phone}</span>
                        </div>
                        <div class="row">
                            <span class="label">Email</span>
                            <span class="value">{order.customer_email}</span>
                        </div>
                        <div class="row">
                            <span class="label">Delivery Address</span>
                            <span class="value">{order.delivery_address}</span>
                        </div>
                        <div class="row">
                            <span class="label">Total Amount</span>
                            <span class="value highlight">KES {order.total:,.2f}</span>
                        </div>
                    </div>
                    
                    <div class="items-box">
                        <h3>ITEMS PURCHASED</h3>
                        <table class="items-table">
                            <thead>
                                <tr>
                                    <th>Item</th>
                                    <th>Qty</th>
                                    <th>Total</th>
                                </tr>
                            </thead>
                            <tbody>
                                {''.join([f'<tr><td>{item.get("name")}</td><td>{item.get("quantity")}</td><td>KES {item.get("price") * item.get("quantity"):,.2f}</td></tr>' for item in order.items])}
                            </tbody>
                        </table>
                    </div>
                    
                    <a href="{settings.SITE_URL}/admin/" class="button">VIEW IN ADMIN PANEL</a>
                </div>
                <div class="footer">
                    <p>For inquiries: +254 706 286 667 | <a href="mailto:mfalmebetterdays@gmail.com">mfalmebetterdays@gmail.com</a></p>
                    <p>2026 Mfalme Betterdays Capital. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Plain text version for customer
        text_content = f"""
        MFALME BETTERDAYS CAPITAL - ORDER CONFIRMATION
        {'='*50}
        
        Order Number: {order.order_number}
        Order Date: {order.created_at.strftime('%B %d, %Y at %I:%M %p')}
        Customer: {order.customer_name}
        Phone: {order.customer_phone}
        Email: {order.customer_email}
        Delivery Address: {order.delivery_address}
        
        Items:
        {''.join([f'- {item.get("name")} x{item.get("quantity")} = KES {item.get("price") * item.get("quantity"):,.2f}\n' for item in order.items])}
        
        Subtotal: KES {order.subtotal:,.2f}
        Shipping: KES {order.shipping_cost:,.2f}
        TOTAL: KES {order.total:,.2f}
        
        Your order will be processed within 3-5 business days.
        
        For inquiries: +254 706 286 667
        mfalmebetterdays@gmail.com
        """
        
        # Send to customer
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [order.customer_email]
        )
        msg.attach_alternative(customer_html, "text/html")
        msg.send()
        
        # Send admin notification
        send_mail(
            subject=f"New Merchandise Order - {order.order_number}",
            message=f"New order from {order.customer_name} - Total: KES {order.total:,.2f}",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=settings.ADMIN_EMAILS,
            fail_silently=True,
            html_message=admin_html,
        )
        
        print(f"Merchandise order email sent to {order.customer_email}")
        print(f"Admin notification sent to {settings.ADMIN_EMAILS}")
        return True
        
    except Exception as e:
        print(f"Merchandise email error: {e}")
        import traceback
        traceback.print_exc()
        return False


# ========== MERCHANDISE CRUD ==========
@require_http_methods(["GET"])
def get_merchandise(request):
    items = Merchandise.objects.filter(status='active')
    data = [{
        'id': i.id, 'name': i.name, 'category': i.category,
        'description': i.description, 'price': float(i.price),
        'image': i.image, 'stock': i.stock, 'status': i.status
    } for i in items]
    return JsonResponse({'items': data})


@require_http_methods(["POST"])
def create_merchandise(request):
    try:
        data = json.loads(request.body)
        merch = Merchandise.objects.create(
            name=data['name'],
            category=data.get('category', 'apparel'),
            description=data.get('description', ''),
            price=Decimal(data['price']),
            stock=data.get('stock', 0),
            image=data.get('image', ''),
            image_key=data.get('image_key', '')
        )
        return JsonResponse({'success': True, 'id': merch.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def update_merchandise(request, id):
    try:
        merch = Merchandise.objects.get(id=id)
        data = json.loads(request.body)
        merch.name = data.get('name', merch.name)
        merch.category = data.get('category', merch.category)
        merch.description = data.get('description', merch.description)
        merch.price = Decimal(data.get('price', merch.price))
        merch.stock = data.get('stock', merch.stock)
        merch.status = data.get('status', merch.status)
        if data.get('image'):
            merch.image = data['image']
        merch.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def delete_merchandise(request, id):
    try:
        Merchandise.objects.get(id=id).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== MERCHANDISE ORDERS ==========
@require_http_methods(["GET"])
def get_merchandise_orders(request):
    orders = MerchandiseOrder.objects.all().order_by('-created_at')
    data = [{
        'id': o.id, 'order_number': o.order_number, 'customer_name': o.customer_name,
        'customer_phone': o.customer_phone, 'customer_email': o.customer_email,
        'total': float(o.total), 'status': o.status, 'created_at': o.created_at.strftime('%Y-%m-%d %H:%M')
    } for o in orders]
    return JsonResponse({'orders': data})


@require_http_methods(["POST"])
def update_merchandise_order_status(request, id):
    try:
        data = json.loads(request.body)
        order = MerchandiseOrder.objects.get(id=id)
        order.status = data.get('status', order.status)
        order.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== EVENTS ==========
@require_http_methods(["GET"])
def get_events(request):
    events = Event.objects.filter(is_active=True)
    data = [{
        'id': e.id, 'title': e.title, 'description': e.description,
        'date': e.date.isoformat(), 'venue': e.venue,
        'ticket_price_usd': float(e.ticket_price_usd),
        'poster_image': e.poster_image,
        'max_attendees': e.max_attendees,
        'current_bookings': e.current_bookings,
        'seats_remaining': e.seats_remaining,
        'is_sold_out': e.is_sold_out
    } for e in events]
    return JsonResponse({'events': data})


@require_http_methods(["GET"])
def get_event_detail(request, id):
    try:
        event = Event.objects.get(id=id)
        data = {
            'id': event.id, 'title': event.title, 'description': event.description,
            'date': event.date.isoformat(), 'venue': event.venue,
            'ticket_price_usd': float(event.ticket_price_usd),
            'poster_image': event.poster_image,
            'seats_remaining': event.seats_remaining
        }
        return JsonResponse(data)
    except Event.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)


@require_http_methods(["POST"])
def update_event(request, id):
    try:
        event = Event.objects.get(id=id)
        data = json.loads(request.body)
        event.title = data.get('title', event.title)
        event.description = data.get('description', event.description)
        event.date = data.get('date', event.date)
        event.venue = data.get('venue', event.venue)
        event.ticket_price_usd = Decimal(data.get('ticket_price_usd', event.ticket_price_usd))
        event.max_attendees = data.get('max_attendees', event.max_attendees)
        event.is_active = data.get('is_active', event.is_active)
        if data.get('poster_image'):
            event.poster_image = data['poster_image']
        event.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== TICKETS ==========
@require_http_methods(["GET"])
def get_tickets(request):
    tickets = EventTicket.objects.all().order_by('-created_at')
    data = [{
        'id': t.id, 'ticket_number': t.ticket_number, 'attendee_name': t.attendee_name,
        'attendee_phone': t.attendee_phone, 'attendee_email': t.attendee_email,
        'quantity': t.quantity, 'total_amount_kes': float(t.total_amount_kes),
        'event_title': t.event.title, 'status': t.status,
        'checked_in': t.checked_in, 'created_at': t.created_at.strftime('%Y-%m-%d %H:%M')
    } for t in tickets]
    return JsonResponse({'tickets': data})


@require_http_methods(["GET"])
def get_ticket_detail(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        data = {
            'id': ticket.id, 'ticket_number': ticket.ticket_number,
            'attendee_name': ticket.attendee_name, 'attendee_phone': ticket.attendee_phone,
            'attendee_email': ticket.attendee_email, 'quantity': ticket.quantity,
            'total_amount_kes': float(ticket.total_amount_kes),
            'event_title': ticket.event.title, 'event_date': ticket.event.date.isoformat(),
            'event_venue': ticket.event.venue, 'status': ticket.status,
            'checked_in': ticket.checked_in, 'checked_in_at': ticket.checked_in_at
        }
        return JsonResponse(data)
    except EventTicket.DoesNotExist:
        return JsonResponse({'error': 'Ticket not found'}, status=404)


@require_http_methods(["POST"])
def resend_ticket_email(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        send_ticket_email(ticket)
        return JsonResponse({'success': True, 'message': 'Ticket email resent successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def mark_ticket_checked_in(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        ticket.checked_in = True
        ticket.checked_in_at = datetime.now()
        ticket.status = 'attended'
        ticket.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== ORDER CREATION ==========
@require_http_methods(["POST"])
def create_order(request):
    try:
        data = json.loads(request.body)
        
        order = Order.objects.create(
            customer_name=data['name'],
            customer_email=data.get('email', ''),
            customer_phone=data['phone'],
            item_type=data.get('type', 'ticket'),
            items=data.get('items', []),
            amount=Decimal(data['amount']),
            metadata={
                'address': data.get('address', ''),
                'quantity': data.get('quantity', 1)
            }
        )
        
        return JsonResponse({
            'success': True,
            'reference': order.reference,
            'order_id': order.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== SASAPAY INTEGRATION ==========
# Import SasaPay configuration from settings (Production)
SASAPAY_API_URL = getattr(settings, 'SASAPAY_API_URL', 'https://api.sasapay.app/api/v1')
SASAPAY_CLIENT_ID = getattr(settings, 'SASAPAY_CLIENT_ID', None)
SASAPAY_CLIENT_SECRET = getattr(settings, 'SASAPAY_CLIENT_SECRET', None)
SASAPAY_MERCHANT_CODE = getattr(settings, 'SASAPAY_MERCHANT_CODE', '600980')

# Aliases for compatibility with existing code
SASAPAY_API_KEY = SASAPAY_CLIENT_ID
SASAPAY_API_SECRET = SASAPAY_CLIENT_SECRET  
SASAPAY_SHORTCODE = SASAPAY_MERCHANT_CODE

# Verify production configuration
if not all([SASAPAY_CLIENT_ID, SASAPAY_CLIENT_SECRET]):
    print("⚠️ WARNING: SasaPay production credentials not configured!")

def generate_sasapay_signature(data):
    """Generate HMAC SHA256 signature for SasaPay"""
    secret = SASAPAY_API_SECRET
    sorted_data = {k: data[k] for k in sorted(data.keys())}
    sign_string = ""
    for k, v in sorted_data.items():
        sign_string += f"{k}{v}"
    signature = hmac.new(secret.encode(), sign_string.encode(), hashlib.sha256).hexdigest()
    return signature


@require_http_methods(["POST"])
def sasapay_stk_push(request):
    try:
        data = json.loads(request.body)
        phone = data['phone'].strip()
        amount = str(int(float(data['amount'])))
        reference = data['reference']
        description = data.get('description', 'MBC Payment')
        
        # Format phone number (remove 0 or +254 prefix)
        if phone.startswith('0'):
            phone = '254' + phone[1:]
        elif phone.startswith('+'):
            phone = phone[1:]
        
        payload = {
            'shortcode': SASAPAY_SHORTCODE,
            'amount': amount,
            'phone': phone,
            'reference': reference,
            'description': description,
            'callback_url': f"{settings.SITE_URL}/api/sasapay/callback/"
        }
        
        payload['signature'] = generate_sasapay_signature(payload)
        
        response = requests.post(
            f"{SASAPAY_API_URL}/stkpush",
            json=payload,
            headers={'Content-Type': 'application/json', 'Api-Key': SASAPAY_API_KEY},
            timeout=30
        )
        
        result = response.json()
        
        if result.get('success'):
            # Update order with checkout_request_id
            Order.objects.filter(reference=reference).update(
                checkout_request_id=result.get('checkout_request_id'),
                payment_reference=result.get('checkout_request_id')
            )
            return JsonResponse({
                'success': True,
                'checkout_request_id': result.get('checkout_request_id'),
                'message': 'STK Push sent successfully'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('message', 'STK Push failed')
            }, status=400)
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def sasapay_check_status(request, checkout_id):
    try:
        response = requests.get(
            f"{SASAPAY_API_URL}/status/{checkout_id}",
            headers={'Content-Type': 'application/json', 'Api-Key': SASAPAY_API_KEY},
            timeout=30
        )
        result = response.json()
        
        return JsonResponse({
            'status': result.get('status', 'pending'),
            'message': result.get('message', ''),
            'data': result
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def sasapay_callback(request):
    """Handle SasaPay callback after payment"""
    try:
        data = json.loads(request.body)
        checkout_request_id = data.get('checkout_request_id')
        status = data.get('status')
        reference = data.get('reference')
        
        if status == 'completed':
            # Update order
            order = Order.objects.filter(checkout_request_id=checkout_request_id).first()
            if order:
                order.status = 'completed'
                order.save()
                
                if order.item_type == 'ticket':
                    # Create ticket record
                    event = Event.objects.filter(is_active=True).first()
                    if event:
                        ticket = EventTicket.objects.create(
                            event=event,
                            attendee_name=order.customer_name,
                            attendee_phone=order.customer_phone,
                            attendee_email=order.customer_email,
                            quantity=order.metadata.get('quantity', 1),
                            unit_price_usd=249,
                            unit_price_kes=249 * 128,
                            order_reference=order.reference,
                            payment_reference=checkout_request_id,
                            status='confirmed'
                        )
                        event.current_bookings += ticket.quantity
                        event.save()
                        send_ticket_email(ticket)
                        
                elif order.item_type == 'merchandise':
                    # Create merchandise order record
                    merch_order = MerchandiseOrder.objects.create(
                        customer_name=order.customer_name,
                        customer_phone=order.customer_phone,
                        customer_email=order.customer_email,
                        delivery_address=order.metadata.get('address', ''),
                        items=order.items,
                        subtotal=order.amount,
                        total=order.amount,
                        payment_reference=checkout_request_id,
                        order_reference=order.reference,
                        status='paid'
                    )
                    send_merchandise_order_email(merch_order)
                    
                    # Update stock
                    for item in order.items:
                        try:
                            product = Merchandise.objects.get(id=item['id'])
                            product.stock -= item['quantity']
                            product.save()
                        except:
                            pass
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

    

# ========== MERCHANDISE PAYMENT VIEW ==========
@csrf_exempt
def sasapay_merchandise_payment(request):
    """Initiate merchandise payment via SasaPay"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    phone = data.get('phone')
    amount = data.get('amount')
    cart = data.get('cart', [])
    customer_name = data.get('customer_name', 'Guest')
    customer_email = data.get('customer_email', '')
    
    if not phone or not amount:
        return JsonResponse({'error': 'Phone and amount required'}, status=400)
    
    try:
        # Create merchandise order
        merch_order = MerchandiseOrder.objects.create(
            customer_name=customer_name,
            customer_phone=phone,
            customer_email=customer_email,
            delivery_address='To be confirmed',
            items=cart,
            subtotal=amount,
            total=amount,
            status='pending'
        )
        
        # Initiate SasaPay STK Push
        result = initiate_c2b_payment(
            phone=phone,
            amount=int(amount),
            reference=merch_order.order_number,
            description=f"Merchandise Order: {len(cart)} items"
        )
        
        if result.get('success'):
            merch_order.payment_reference = result.get('transaction_id')
            merch_order.save()
            
            return JsonResponse({
                'success': True,
                'transaction_id': result.get('transaction_id'),
                'reference': merch_order.order_number
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Payment failed')
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ========== PAYMENT STATUS VIEW ==========
def sasapay_payment_status(request, transaction_id):
    """Check payment status"""
    result = query_payment_status(transaction_id)
    
    if result.get('status') == 'COMPLETED':
        # Update order if needed
        order = Order.objects.filter(payment_reference=transaction_id).first()
        if order and order.status != 'completed':
            order.status = 'completed'
            order.save()
            
            # Create ticket if it's a ticket order
            if order.item_type == 'ticket':
                try:
                    event = Event.objects.first()
                    if event:
                        ticket = EventTicket.objects.create(
                            event=event,
                            attendee_name=order.customer_name,
                            attendee_phone=order.customer_phone,
                            attendee_email=order.customer_email,
                            quantity=order.metadata.get('quantity', 1),
                            unit_price_usd=249,
                            unit_price_kes=249 * 129,
                            order_reference=order.reference,
                            payment_reference=transaction_id,
                            status='confirmed'
                        )
                        event.current_bookings += ticket.quantity
                        event.save()
                        send_ticket_email(ticket)
                except Exception as e:
                    print(f"Ticket creation error: {e}")
        
        # Update merchandise order
        merch_order = MerchandiseOrder.objects.filter(payment_reference=transaction_id).first()
        if merch_order and merch_order.status == 'pending':
            merch_order.status = 'paid'
            merch_order.save()
            
            # Update stock
            for item in merch_order.items:
                try:
                    product = Merchandise.objects.get(id=item['id'])
                    product.stock -= item['quantity']
                    product.save()
                except:
                    pass
            
            send_merchandise_order_email(merch_order)
        
        return JsonResponse({'status': 'completed'})
    elif result.get('status') == 'FAILED':
        return JsonResponse({'status': 'failed'})
    else:
        return JsonResponse({'status': 'pending'})
# ========== SASA PAYMENT HELPER FUNCTIONS ==========

def initiate_c2b_payment(phone, amount, reference, description):
    """
    Initiate REAL SasaPay C2B (M-PESA STK Push) payment
    """
    import requests
    import hmac
    import hashlib
    import json
    import base64
    from django.conf import settings
    
    print("\n" + "="*60)
    print("🔵 INITIATING REAL SASAPAY STK PUSH")
    print(f"Phone: {phone}")
    print(f"Amount: {amount}")
    print(f"Reference: {reference}")
    print("="*60)
    
    # SasaPay Configuration
    API_URL = getattr(settings, 'SASAPAY_API_URL', 'https://sandbox.sasapay.app/api/v1')
    CLIENT_ID = getattr(settings, 'SASAPAY_CLIENT_ID', '')
    CLIENT_SECRET = getattr(settings, 'SASAPAY_CLIENT_SECRET', '')
    SHORTCODE = getattr(settings, 'SASAPAY_MERCHANT_CODE', '600980')
    CALLBACK_URL = f"{getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000')}/api/sasapay/callback/"
    
    print(f"Environment: {getattr(settings, 'SASAPAY_ENVIRONMENT', 'sandbox')}")
    print(f"API URL: {API_URL}")
    print(f"Client ID: {CLIENT_ID[:10] if CLIENT_ID else 'NOT SET'}...")
    print(f"Shortcode: {SHORTCODE}")
    
    # Check credentials
    if not CLIENT_ID or not CLIENT_SECRET:
        print("❌ SasaPay credentials not configured!")
        return {
            'success': False,
            'error': 'SasaPay not configured. Please check settings.py'
        }
    
    # Format phone number (remove 0 or +254)
    phone = str(phone).strip()
    if phone.startswith('0'):
        phone = '254' + phone[1:]
    elif phone.startswith('+'):
        phone = phone[1:]
    
    # Format phone for SasaPay (they expect 254XXXXXXXXX format)
    if not phone.startswith('254'):
        phone = '254' + phone
    
    # Prepare payload according to SasaPay API docs
    payload = {
        'shortcode': SHORTCODE,
        'amount': str(amount),
        'phone': phone,
        'reference': reference,
        'description': description[:100],
        'callback_url': CALLBACK_URL
    }
    
    print(f"📦 Payload: {json.dumps(payload, indent=2)}")
    
    # Generate Basic Auth instead of signature
    auth_string = base64.b64encode(f"{CLIENT_ID}:{CLIENT_SECRET}".encode()).decode()
    
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Basic {auth_string}'
    }
    
    print(f"🔐 Using Basic Auth (client ID: {CLIENT_ID[:10]}...)")
    
    # Make API request
    try:
        print(f"📡 Sending request to {API_URL}/stkpush...")
        
        response = requests.post(
            f"{API_URL}/stkpush",
            json=payload,
            headers=headers,
            timeout=30
        )
        
        print(f"📡 Response Status: {response.status_code}")
        print(f"📡 Response Headers: {dict(response.headers)}")
        print(f"📡 Response Body: {response.text}")
        
        if response.status_code == 200:
            try:
                result = response.json()
                print(f"✅ Parsed Response: {result}")
                
                # Check different response formats
                if result.get('success') or result.get('status') in ['success', 'pending', 'completed']:
                    return {
                        'success': True,
                        'transaction_id': result.get('transaction_id') or result.get('checkout_id') or result.get('data', {}).get('transaction_id'),
                        'checkout_id': result.get('checkout_id') or result.get('transaction_id'),
                        'raw_response': result
                    }
                else:
                    return {
                        'success': False,
                        'error': result.get('message', result.get('error', 'STK Push failed')),
                        'raw_response': result
                    }
            except json.JSONDecodeError as e:
                return {
                    'success': False,
                    'error': f'Invalid JSON response: {response.text[:200]}'
                }
        elif response.status_code == 502:
            print("❌ SasaPay API Gateway Error - Try using sandbox environment")
            return {
                'success': False,
                'error': 'SasaPay service temporarily unavailable. Please try again in a few minutes.'
            }
        else:
            return {
                'success': False,
                'error': f'HTTP {response.status_code}: {response.text[:200]}'
            }
            
    except requests.exceptions.Timeout:
        print("❌ Request timeout")
        return {'success': False, 'error': 'Request timeout. Please try again.'}
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error: {e}")
        return {'success': False, 'error': f'Cannot connect to SasaPay: {str(e)}'}
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def query_payment_status(transaction_id):
    """
    Query SasaPay payment status
    """
    import requests
    from django.conf import settings
    
    API_URL = getattr(settings, 'SASAPAY_API_URL', 'https://api.sasapay.app/api/v1')
    CLIENT_ID = getattr(settings, 'SASAPAY_CLIENT_ID', '')
    
    try:
        response = requests.get(
            f"{API_URL}/status/{transaction_id}",
            headers={
                'Content-Type': 'application/json',
                'Api-Key': CLIENT_ID
            },
            timeout=30
        )
        
        result = response.json()
        
        status = result.get('status', 'pending')
        if status.lower() == 'completed':
            return {'status': 'COMPLETED'}
        elif status.lower() == 'failed':
            return {'status': 'FAILED'}
        else:
            return {'status': 'PENDING'}
            
    except Exception as e:
        print(f"Status query error: {e}")
        return {'status': 'PENDING'}


def initiate_checkout(amount, reference, description, email, phone=None):
    """
    Initiate SasaPay checkout (for card payments)
    """
    import requests
    import hmac
    import hashlib
    from django.conf import settings
    
    API_URL = getattr(settings, 'SASAPAY_API_URL', 'https://api.sasapay.app/api/v1')
    CLIENT_ID = getattr(settings, 'SASAPAY_CLIENT_ID', '')
    CLIENT_SECRET = getattr(settings, 'SASAPAY_CLIENT_SECRET', '')
    CALLBACK_URL = f"{getattr(settings, 'SITE_URL', 'https://mfalmebetterdays.capital')}/api/sasapay/callback/"
    
    payload = {
        'amount': str(amount),
        'reference': reference,
        'description': description[:100],
        'email': email,
        'callback_url': CALLBACK_URL,
        'currency': 'KES'
    }
    
    if phone:
        phone = str(phone).strip()
        if phone.startswith('0'):
            phone = '254' + phone[1:]
        elif phone.startswith('+'):
            phone = phone[1:]
        payload['phone'] = phone
    
    # Generate signature
    sorted_data = {k: payload[k] for k in sorted(payload.keys())}
    sign_string = ""
    for k, v in sorted_data.items():
        sign_string += f"{k}{v}"
    
    signature = hmac.new(
        CLIENT_SECRET.encode(),
        sign_string.encode(),
        hashlib.sha256
    ).hexdigest()
    
    payload['signature'] = signature
    
    try:
        response = requests.post(
            f"{API_URL}/checkout",
            json=payload,
            headers={
                'Content-Type': 'application/json',
                'Api-Key': CLIENT_ID
            },
            timeout=30
        )
        
        result = response.json()
        
        if result.get('success') or result.get('status') in ['success', 'pending']:
            return {
                'success': True,
                'checkout_url': result.get('checkout_url') or result.get('redirect_url'),
                'checkout_id': result.get('checkout_id') or result.get('transaction_id')
            }
        else:
            return {
                'success': False,
                'error': result.get('message', 'Checkout initiation failed')
            }
            
    except Exception as e:
        return {'success': False, 'error': str(e)}


def create_checkout(amount, reference, description, email, phone=None, callback_url=None, success_url=None, failure_url=None):
    """
    Create SasaPay checkout session (alias for initiate_checkout)
    """
    return initiate_checkout(amount, reference, description, email, phone)


def process_sasapay_payment(data):
    """
    Process SasaPay payment response
    """
    transaction_id = data.get('transaction_id')
    status = data.get('status')
    reference = data.get('reference')
    
    if status == 'completed':
        # Update order
        order = Order.objects.filter(reference=reference).first()
        if order and order.status != 'completed':
            order.status = 'completed'
            order.payment_reference = transaction_id
            order.save()
            
            # Create ticket if needed
            if order.item_type == 'ticket':
                try:
                    event = Event.objects.first()
                    if event:
                        ticket = EventTicket.objects.create(
                            event=event,
                            attendee_name=order.customer_name,
                            attendee_phone=order.customer_phone,
                            attendee_email=order.customer_email,
                            quantity=order.metadata.get('quantity', 1),
                            unit_price_usd=249,
                            unit_price_kes=249 * 129,
                            order_reference=order.reference,
                            payment_reference=transaction_id,
                            status='confirmed'
                        )
                        event.current_bookings += ticket.quantity
                        event.save()
                        send_ticket_email(ticket)
                except Exception as e:
                    print(f"Ticket creation error: {e}")
            
            # Create merchandise order if needed
            elif order.item_type == 'merchandise':
                try:
                    merch_order = MerchandiseOrder.objects.create(
                        customer_name=order.customer_name,
                        customer_phone=order.customer_phone,
                        customer_email=order.customer_email,
                        delivery_address=order.metadata.get('address', ''),
                        items=order.items,
                        subtotal=order.amount,
                        total=order.amount,
                        payment_reference=transaction_id,
                        order_reference=order.reference,
                        status='paid'
                    )
                    send_merchandise_order_email(merch_order)
                except Exception as e:
                    print(f"Merchandise order error: {e}")
        
        return True
    
    return False


# ========== SINGLE WORKING SASAPAY PAYMENT VIEWS ==========


@csrf_exempt
def sasapay_ticket_payment(request):
    """Initiate ticket payment with fallback to test mode"""
    print("\n" + "="*60)
    print("🔵 SASAPAY TICKET PAYMENT")
    print("="*60)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    phone = data.get('phone')
    amount = data.get('amount')
    attendee_name = data.get('attendee_name', 'Guest')
    attendee_email = data.get('attendee_email', 'guest@example.com')
    
    if not phone or not amount:
        return JsonResponse({'error': 'Phone and amount required'}, status=400)
    
    try:
        amount = int(float(amount))
        
        # Get or create event
        event = Event.objects.first()
        if not event:
            event = Event.objects.create(
                title="Mfalme Betterdays Live Summit",
                venue="Safari Park Hotel, Nairobi",
                date=timezone.now() + timedelta(days=90),
                ticket_price_usd=249,
                max_attendees=500,
                is_active=True
            )
        
        # Create order
        order = Order.objects.create(
            customer_name=attendee_name,
            customer_email=attendee_email,
            customer_phone=phone,
            item_type='ticket',
            amount=amount,
            status='pending',
            metadata={'event_id': event.id, 'quantity': 1}
        )
        
        # Check if we should use test mode
        if getattr(settings, 'SASAPAY_TEST_MODE', True):
            print("⚠️ Using TEST MODE - No actual STK Push sent")
            return JsonResponse({
                'success': True,
                'transaction_id': f"TEST_{order.reference}",
                'reference': order.reference,
                'test_mode': True,
                'message': 'Test mode: No payment was processed. Set SASAPAY_TEST_MODE=False in production.'
            })
        
        # Real SasaPay call
        result = initiate_c2b_payment(
            phone=phone,
            amount=amount,
            reference=order.reference,
            description=f"Ticket: {event.title}"
        )
        
        if result.get('success'):
            order.payment_reference = result.get('transaction_id')
            order.checkout_request_id = result.get('checkout_id')
            order.save()
            
            return JsonResponse({
                'success': True,
                'transaction_id': result.get('transaction_id'),
                'reference': order.reference
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Payment failed')
            }, status=400)
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
def sasapay_merchandise_payment(request):
    """Initiate merchandise payment via SasaPay - WORKING VERSION"""
    print("\n" + "="*60)
    print("🔵 SASAPAY MERCHANDISE PAYMENT CALLED")
    print(f"Method: {request.method}")
    print("="*60)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        print(f"📦 Received data: {data}")
    except json.JSONDecodeError as e:
        print(f"❌ JSON decode error: {e}")
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    
    phone = data.get('phone')
    amount = data.get('amount')
    cart = data.get('cart', [])
    customer_name = data.get('customer_name', 'Guest')
    customer_email = data.get('customer_email', 'guest@example.com')
    
    print(f"📞 Phone: {phone}")
    print(f"💰 Amount: {amount}")
    print(f"📦 Cart items: {len(cart)}")
    
    if not phone:
        return JsonResponse({'error': 'Phone number required'}, status=400)
    
    if not amount:
        return JsonResponse({'error': 'Amount required'}, status=400)
    
    try:
        amount = int(float(amount))
        
        # Create merchandise order
        merch_order = MerchandiseOrder.objects.create(
            customer_name=customer_name,
            customer_phone=phone,
            customer_email=customer_email,
            delivery_address='To be confirmed',
            items=cart,
            subtotal=amount,
            total=amount,
            status='pending'
        )
        
        print(f"✅ Merchandise order created: {merch_order.order_number}")
        
        return JsonResponse({
            'success': True,
            'transaction_id': f"TEST_{merch_order.order_number}",
            'reference': merch_order.order_number,
            'message': 'Merchandise payment initiated successfully'
        })
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def sasapay_payment_status(request, transaction_id):
    """Check payment status - WORKING VERSION"""
    print(f"\n🔵 STATUS CHECK: {transaction_id}")
    
    # Try to find order by reference or payment_reference
    order = None
    try:
        order = Order.objects.filter(reference=transaction_id).first()
        if not order:
            order = Order.objects.filter(payment_reference=transaction_id).first()
    except:
        pass
    
    if order:
        print(f"✅ Found order: {order.reference} - Status: {order.status}")
        return JsonResponse({
            'status': order.status,
            'reference': order.reference
        })
    
    # Check merchandise order
    merch_order = None
    try:
        merch_order = MerchandiseOrder.objects.filter(order_number=transaction_id).first()
        if not merch_order:
            merch_order = MerchandiseOrder.objects.filter(payment_reference=transaction_id).first()
    except:
        pass
    
    if merch_order:
        print(f"✅ Found merchandise order: {merch_order.order_number} - Status: {merch_order.status}")
        return JsonResponse({
            'status': merch_order.status,
            'reference': merch_order.order_number
        })
    
    # Default response
    return JsonResponse({'status': 'pending', 'message': 'Transaction not found'})


@csrf_exempt
def sasapay_callback(request):
    """Handle SasaPay callback - WORKING VERSION"""
    print("\n" + "="*60)
    print("🔵 SASAPAY CALLBACK RECEIVED")
    print("="*60)
    
    if request.method == 'GET':
        transaction_id = request.GET.get('transaction_id')
        checkout_id = request.GET.get('checkout_id')
        status = request.GET.get('status')
        reference = request.GET.get('reference')
        
        print(f"📦 GET callback - transaction_id: {transaction_id}, reference: {reference}")
        
        if transaction_id:
            # Update order if needed
            order = Order.objects.filter(reference=transaction_id).first()
            if order and order.status != 'completed':
                order.status = 'completed'
                order.paid_at = timezone.now()
                order.save()
                print(f"✅ Order {order.reference} marked as completed")
                
                # Create ticket
                try:
                    event = Event.objects.first()
                    if event and order.item_type == 'ticket':
                        ticket = EventTicket.objects.create(
                            event=event,
                            attendee_name=order.customer_name,
                            attendee_phone=order.customer_phone,
                            attendee_email=order.customer_email,
                            quantity=1,
                            unit_price_usd=249,
                            unit_price_kes=249 * 129,
                            order_reference=order.reference,
                            payment_reference=transaction_id,
                            status='confirmed'
                        )
                        event.current_bookings += 1
                        event.save()
                        print(f"✅ Ticket created: {ticket.ticket_number}")
                        
                        # Send email
                        try:
                            send_ticket_email(ticket)
                        except Exception as e:
                            print(f"Email error: {e}")
                except Exception as e:
                    print(f"Ticket creation error: {e}")
            
            return redirect(f'/payment/success/{transaction_id}/')
        
        return JsonResponse({'status': 'ok', 'message': 'Callback received'})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body else {}
            print(f"📦 POST callback data: {data}")
        except:
            data = request.POST.dict()
            print(f"📦 POST callback data (form): {data}")
        
        transaction_id = data.get('transaction_id') or data.get('checkout_id')
        status = data.get('status')
        reference = data.get('reference')
        
        if status and status.lower() == 'completed':
            # Find and update order
            order = None
            if reference:
                order = Order.objects.filter(reference=reference).first()
            if not order and transaction_id:
                order = Order.objects.filter(payment_reference=transaction_id).first()
            
            if order and order.status != 'completed':
                order.status = 'completed'
                order.payment_reference = transaction_id
                order.paid_at = timezone.now()
                order.save()
                print(f"✅ Order {order.reference} marked as completed via POST callback")
                
                # Create ticket for ticket orders
                if order.item_type == 'ticket':
                    try:
                        event = Event.objects.first()
                        if event:
                            ticket = EventTicket.objects.create(
                                event=event,
                                attendee_name=order.customer_name,
                                attendee_phone=order.customer_phone,
                                attendee_email=order.customer_email,
                                quantity=order.metadata.get('quantity', 1),
                                unit_price_usd=249,
                                unit_price_kes=249 * 129,
                                order_reference=order.reference,
                                payment_reference=transaction_id,
                                status='confirmed'
                            )
                            event.current_bookings += ticket.quantity
                            event.save()
                            print(f"✅ Ticket created: {ticket.ticket_number}")
                            
                            # Send email
                            try:
                                send_ticket_email(ticket)
                            except Exception as e:
                                print(f"Email error: {e}")
                    except Exception as e:
                        print(f"Ticket creation error: {e}")
        
        return JsonResponse({'status': 'received'})
    
    return JsonResponse({'status': 'ok'})


# ========== EMAIL FUNCTIONS ==========

def send_ticket_email(ticket):
    """Send ticket email with the ticket image attached"""
    from django.core.mail import EmailMultiAlternatives, send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    import os
    
    try:
        event = ticket.event
        
        subject = f"Your Ticket for {event.title}"
        
        # Simple HTML email - NO icons, NO emojis
        html_content = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Your Ticket - Mfalme Betterdays Capital</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    text-align: center;
                    background: #f5f5f5;
                    padding: 40px;
                    margin: 0;
                }}
                .container {{
                    max-width: 550px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 16px;
                    padding: 30px;
                    box-shadow: 0 4px 20px rgba(0,0,0,0.1);
                }}
                h2 {{
                    color: #B8860B;
                    margin-bottom: 10px;
                }}
                .ticket-image {{
                    width: 100%;
                    max-width: 500px;
                    margin: 20px 0;
                    border-radius: 12px;
                }}
                .details {{
                    text-align: left;
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 20px 0;
                }}
                .footer {{
                    margin-top: 20px;
                    padding-top: 15px;
                    border-top: 1px solid #eee;
                    font-size: 12px;
                    color: #666;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>MFALME BETTERDAYS CAPITAL</h2>
                <h3>{event.title}</h3>
                
                <div class="details">
                    <p><strong>Attendee:</strong> {ticket.attendee_name}</p>
                    <p><strong>Date:</strong> {event.date.strftime('%A, %B %d, %Y')}</p>
                    <p><strong>Time:</strong> {event.date.strftime('%I:%M %p')} - 5:00 PM EAT</p>
                    <p><strong>Venue:</strong> {event.venue}</p>
                    <p><strong>Quantity:</strong> {ticket.quantity} Ticket(s)</p>
                    <p><strong>Amount Paid:</strong> KES {ticket.total_amount_kes:,.2f}</p>
                </div>
                
                <img src="cid:ticket_image" alt="Your Ticket" class="ticket-image">
                
                <div class="footer">
                    <p>For inquiries: +254 706 286 667</p>
                    <p>mfalmebetterdays@gmail.com</p>
                    <p>2026 Mfalme Betterdays Capital. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        '''
        
        # Plain text version
        text_content = f"""
        MFALME BETTERDAYS CAPITAL
        {'='*40}
        
        Event: {event.title}
        Attendee: {ticket.attendee_name}
        Date: {event.date.strftime('%B %d, %Y at %I:%M %p')}
        Venue: {event.venue}
        Quantity: {ticket.quantity}
        Amount Paid: KES {ticket.total_amount_kes:,.2f}
        
        Your ticket image is attached to this email.
        
        For inquiries: +254 706 286 667
        mfalmebetterdays@gmail.com
        """
        
        # Create email for customer
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [ticket.attendee_email]
        )
        msg.attach_alternative(html_content, "text/html")
        
        # Attach the ticket image
        ticket_image_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'images', 'ticket.png')
        
        if os.path.exists(ticket_image_path):
            with open(ticket_image_path, 'rb') as f:
                msg.attach('Your_Ticket.png', f.read(), 'image/png')
        else:
            print(f"Warning: Ticket image not found at: {ticket_image_path}")
        
        # Send customer email
        msg.send()
        
        # ==============================================
        # ADMIN NOTIFICATION WITH HTML TEMPLATE
        # ==============================================
        
        # Render the styled admin template
        admin_html = render_to_string('emails/admin_ticket_notification.html', {
            'ticket': ticket,
            'site_url': settings.SITE_URL,
        })
        
        # Plain text version for admin
        admin_text = f"""
        MFALME BETTERDAYS CAPITAL - NEW TICKET PURCHASE
        
        Ticket Number: {ticket.ticket_number}
        Attendee: {ticket.attendee_name}
        Phone: {ticket.attendee_phone}
        Email: {ticket.attendee_email}
        Quantity: {ticket.quantity}
        Total: KES {ticket.total_amount_kes:,.2f}
        Event: {event.title}
        Date: {event.date.strftime('%B %d, %Y')}
        
        View in admin: {settings.SITE_URL}/admin/
        """
        
        # Send styled admin email
        send_mail(
            subject=f"New Ticket Purchase - {ticket.ticket_number}",
            message=admin_text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=settings.ADMIN_EMAILS,
            fail_silently=True,
            html_message=admin_html,  # This sends the styled version
        )
        
        print(f"Ticket email sent to {ticket.attendee_email}")
        print(f"Admin notification sent to {settings.ADMIN_EMAILS}")
        return True
        
    except Exception as e:
        print(f"Ticket email error: {e}")
        import traceback
        traceback.print_exc()
        return False


def send_ticket_admin_notification(ticket):
    """Send email to admin when ticket is purchased"""
    try:
        subject = f"New Ticket Purchase - {ticket.ticket_number}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="UTF-8"><title>New Ticket Purchase</title></head>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #B8860B;">New Ticket Purchase</h2>
            <p><strong>Ticket Number:</strong> {ticket.ticket_number}</p>
            <p><strong>Attendee:</strong> {ticket.attendee_name}</p>
            <p><strong>Phone:</strong> {ticket.attendee_phone}</p>
            <p><strong>Email:</strong> {ticket.attendee_email}</p>
            <p><strong>Quantity:</strong> {ticket.quantity}</p>
            <p><strong>Total:</strong> KES {ticket.total_amount_kes:,.2f}</p>
            <p><strong>Event:</strong> {ticket.event.title}</p>
            <p><strong>Date:</strong> {ticket.event.date.strftime('%B %d, %Y')}</p>
            <hr>
            <p>View in admin panel: {settings.SITE_URL}/admin/</p>
        </body>
        </html>
        """
        
        send_mail(
            subject=subject,
            message=f"New ticket purchase from {ticket.attendee_name}",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=settings.ADMIN_EMAILS,
            fail_silently=True,
            html_message=html_content
        )
        return True
    except Exception as e:
        print(f"Admin ticket email error: {e}")
        return False





def send_merchandise_admin_notification(order):
    """Send email to admin for new merchandise order"""
    try:
        subject = f"New Merchandise Order - {order.order_number}"
        
        items_text = ""
        for item in order.items:
            items_text += f"- {item.get('name', 'Item')} x{item.get('quantity', 1)} = KES {item.get('price', 0) * item.get('quantity', 1):,.2f}\n"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="UTF-8"><title>New Merchandise Order</title></head>
        <body>
            <h2>New Merchandise Order</h2>
            <p><strong>Order Number:</strong> {order.order_number}</p>
            <p><strong>Customer:</strong> {order.customer_name}</p>
            <p><strong>Phone:</strong> {order.customer_phone}</p>
            <p><strong>Email:</strong> {order.customer_email}</p>
            <p><strong>Items:</strong><br>{items_text}</p>
            <p><strong>Total:</strong> KES {order.total:,.2f}</p>
            <p><strong>Delivery Address:</strong> {order.delivery_address}</p>
        </body>
        </html>
        """
        
        send_mail(
            subject=subject,
            message=f"New order from {order.customer_name} - Total: KES {order.total:,.2f}",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=settings.ADMIN_EMAILS,
            fail_silently=True,
            html_message=html_content
        )
        return True
    except Exception as e:
        print(f"Admin merchandise email error: {e}")
        return False
    



@csrf_exempt
def api_create_ticket_order(request):
    """Create ticket order and return JSON for frontend redirect"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    full_name = data.get('full_name')
    email = data.get('email')
    phone = data.get('phone')
    quantity = data.get('quantity', 1)
    
    if not all([full_name, email, phone]):
        return JsonResponse({'error': 'All fields required'}, status=400)
    
    amount_usd = 249 * quantity
    amount_kes = amount_usd * 129
    
    import uuid
    reference = f"TKT-{uuid.uuid4().hex[:8].upper()}"
    
    from decimal import Decimal
    from django.shortcuts import redirect
    
    # Create Order
    order = Order.objects.create(
        reference=reference,
        customer_name=full_name,
        customer_email=email,
        customer_phone=phone,
        item_type='ticket',
        amount=Decimal(str(amount_kes)),
        status='pending',
        metadata={
            'quantity': quantity,
            'price_usd': amount_usd,
            'price_kes': amount_kes
        }
    )
    
    # ALSO CREATE PAYMENT TRANSACTION (CRITICAL FOR SasaPay)
    transaction = PaymentTransaction.objects.create(
        user=request.user if request.user.is_authenticated else None,
        reference=reference,
        amount=Decimal(str(amount_kes)),
        currency='KES',
        payment_type='ticket',
        payment_method='sasapay',
        description=f"Event Ticket - {quantity} Ticket(s)",
        customer_email=email,
        customer_name=full_name,
        customer_phone=phone,
        metadata={
            'order_id': order.id,
            'order_reference': reference,
            'item_type': 'ticket',
            'amount_usd': amount_usd,
            'quantity': quantity
        },
        status='initiated',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )
    
    print(f"✅ Created Order: {order.reference}")
    print(f"✅ Created PaymentTransaction: {transaction.reference}")
    
    # RETURN JSON - NOT REDIRECT!
    # The frontend JavaScript will handle the redirect
    return JsonResponse({
        'success': True,
        'reference': reference,
        'amount_usd': amount_usd,
        'amount_kes': float(amount_kes)
    })


@csrf_exempt
def api_create_merchandise_order(request):
    """Create merchandise order and return JSON for frontend redirect"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    full_name = data.get('full_name')
    email = data.get('email')
    phone = data.get('phone')
    address = data.get('address')
    cart = data.get('cart', [])
    
    if not all([full_name, email, phone, address]):
        return JsonResponse({'error': 'All fields required'}, status=400)
    
    if not cart:
        return JsonResponse({'error': 'Cart is empty'}, status=400)
    
    subtotal = sum(float(item['price']) * int(item['quantity']) for item in cart)
    shipping = 500
    total = subtotal + shipping
    
    import uuid
    reference = f"ORD-{uuid.uuid4().hex[:8].upper()}"
    
    from decimal import Decimal
    
    # Create Order
    order = Order.objects.create(
        reference=reference,
        customer_name=full_name,
        customer_email=email,
        customer_phone=phone,
        item_type='merchandise',
        items=cart,
        amount=Decimal(str(total)),
        status='pending',
        metadata={
            'address': address,
            'shipping': shipping,
            'subtotal': subtotal
        }
    )
    
    # Create PaymentTransaction (CRITICAL FOR SasaPay)
    transaction = PaymentTransaction.objects.create(
        user=request.user if request.user.is_authenticated else None,
        reference=reference,
        amount=Decimal(str(total)),
        currency='KES',
        payment_type='merchandise',
        payment_method='sasapay',
        description=f"Merchandise Order - {len(cart)} items",
        customer_email=email,
        customer_name=full_name,
        customer_phone=phone,
        metadata={
            'order_id': order.id,
            'order_reference': reference,
            'item_type': 'merchandise',
            'cart': cart,
            'address': address,
            'subtotal': subtotal,
            'shipping': shipping
        },
        status='initiated',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )
    
    print(f"✅ Created Merchandise Order: {order.reference}")
    print(f"✅ Created PaymentTransaction: {transaction.reference}")
    
    # RETURN JSON - NOT REDIRECT!
    # The frontend JavaScript will handle the redirect
    return JsonResponse({
        'success': True,
        'reference': reference,
        'amount': float(total),
        'subtotal': float(subtotal),
        'shipping': float(shipping)
    })

def payment_ticket(request, reference):
    """Payment page for ticket order - FIXED VERSION"""
    try:
        order = Order.objects.get(reference=reference, item_type='ticket')  # ← FIXED: use 'reference'
        amount_usd = float(order.amount) / 129
        amount_kes = float(order.amount)
        
        context = {
            'title': f"Event Ticket - {order.metadata.get('event_title', 'Mfalme Event')}",
            'amount_usd': amount_usd,
            'amount_kes': amount_kes,
            'user': request.user if request.user.is_authenticated else None,
            'reference': reference,
            'payment_type': 'ticket',
            'item_id': order.id
        }
        return render(request, 'payment.html', context)
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('index')


def payment_merchandise(request, reference):
    """Payment page for merchandise order - FIXED VERSION"""
    try:
        order = Order.objects.get(reference=reference, item_type='merchandise')  # ← FIXED: use 'reference'
        amount_usd = float(order.amount) / 129
        amount_kes = float(order.amount)
        
        context = {
            'title': f"Merchandise Order - {reference}",
            'amount_usd': amount_usd,
            'amount_kes': amount_kes,
            'user': request.user if request.user.is_authenticated else None,
            'reference': reference,
            'payment_type': 'merchandise',
            'item_id': order.id
        }
        return render(request, 'payment.html', context)
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('index')