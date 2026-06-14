from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.files.storage import default_storage
from django.contrib import messages
from django.conf import settings
from django.views.decorators.http import require_http_methods
import json
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
import os
import re

from .models import (
    MfalmeUsers, PaymentTransaction, Package, TrainingVideo, 
    UserVideoAccess, Course, UserCourse, PDF, UserPDFAccess,
    SupportTicket, TicketReply, ActivityLog, Notification,
    PartnershipProgram, UserPartnership, ContactSubmission,
    Blog, FAQ, Testimonial, Statistic, EducationProgram,
    MentorshipProgram, CommunityTier, UserCommunityMembership,
    VerificationCode, Event, EventTicket, Merchandise,
    MerchandiseOrder, Order, Book, BookOrder, Watchlist,
    InstituteApplication, CommunityJoinRequest, SiteContent,
    HeroSlider, AboutSection, Brokerage, Subscription,
    UserEducationEnrollment, UserSession, SystemSettings,
    PaymentMethod, ContactInfo, Logo, UserActivity,
    SiteContentVersion,
)
# ==================== HELPER FUNCTIONS ====================

def sanitize_filename(filename):
    """Remove special characters and spaces from filename"""
    # Remove path
    filename = os.path.basename(filename)
    # Split name and extension
    name, ext = os.path.splitext(filename)
    # Replace spaces and special characters with underscore
    name = re.sub(r'[^a-zA-Z0-9]', '_', name)
    # Add timestamp to ensure uniqueness
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{name}_{timestamp}{ext}"

def get_admin_user(request):
    """
    Get a proper MfalmeUsers instance for the currently logged-in admin.
    This solves the issue where request.user is AnonymousUser but we need
    a real user for ActivityLog and other operations.
    """
    admin_username = request.session.get('admin_username', 'System')
    
    # Try to find existing admin user by username or email pattern
    admin_user = MfalmeUsers.objects.filter(
        Q(username__iexact=admin_username) | 
        Q(email__iexact=f'admin@{admin_username.lower()}.com') |
        Q(email__iexact=f'{admin_username.lower()}@mfalmebetterdays.com')
    ).first()
    
    # If not found, try to find any superuser
    if not admin_user:
        admin_user = MfalmeUsers.objects.filter(is_superuser=True).first()
    
    # If still not found, try to find any staff user
    if not admin_user:
        admin_user = MfalmeUsers.objects.filter(is_staff=True).first()
    
    # If absolutely no admin user exists, create one
    if not admin_user:
        # Check if we already have a superuser in the database
        existing_superuser = MfalmeUsers.objects.filter(is_superuser=True).first()
        if existing_superuser:
            admin_user = existing_superuser
        else:
            # Create a new admin user
            admin_user = MfalmeUsers.objects.create_user(
                email=f'admin@{admin_username.lower()}.com',
                password='TemporaryPassword123!',
                username=f'Admin_{admin_username}',
                first_name='System',
                last_name='Administrator',
                phone='0000000000',
                account_status='active',
                is_staff=True,
                is_superuser=True,
                email_verified=True,
                elite_rank='General'
            )
            print(f"✅ Created new admin user: {admin_user.email}")
    
    return admin_user

# ==================== CUSTOM ADMIN AUTHENTICATION ====================
# Hardcoded admin credentials - ONLY MESH AND MFALME 
VALID_ADMINS = {
    'Mesh': '1234',
    'Mfalme': 'Mfalme@2026!'
}

def is_admin_authenticated(request):
    """Check if admin is authenticated via custom session"""
    return (request.session.get('admin_authenticated') and 
            request.session.get('admin_username') in VALID_ADMINS)

# ==================== ADMIN ACCESS DECORATOR ====================

def admin_required(view_func):
    """Decorator to check if user is authenticated via custom session"""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if is_admin_authenticated(request):
            return view_func(request, *args, **kwargs)
        # For API endpoints that expect JSON
        if request.path.startswith('/admin/api/'):
            return JsonResponse({'error': 'Admin access required'}, status=403)
        # For HTML views, redirect to admin login
        return redirect('admin_login')
    return wrapper

# ==================== ADMIN AUTHENTICATION ====================

def admin_login_view(request):
    """Admin login page - ONLY MESH AND MFALME - NO DATABASE REQUIRED"""
    
    # If already logged in via custom session, go to dashboard
    if is_admin_authenticated(request):
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        # Check if it's an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Get form data
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                username = data.get('username', '').strip()
                password = data.get('password', '')
            except json.JSONDecodeError:
                if is_ajax:
                    return JsonResponse({'success': False, 'message': 'Invalid JSON'})
                messages.error(request, 'Invalid request')
                return redirect('admin_login')
        else:
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
        
        if not username or not password:
            if is_ajax:
                return JsonResponse({'success': False, 'message': 'Please enter both fields'})
            messages.error(request, 'Please enter both fields')
            return redirect('admin_login')
        
        # Check credentials against hardcoded list (ONLY MESH AND MFALME)
        if username in VALID_ADMINS and password == VALID_ADMINS[username]:
            # MANUALLY set session variables - NO DATABASE INVOLVED
            request.session.flush()  # Clear any existing session
            request.session['admin_authenticated'] = True
            request.session['admin_username'] = username
            request.session['admin_login_time'] = str(timezone.now())
            request.session.save()
            
            print(f"✅ Admin {username} logged in successfully via hardcoded credentials")
            print(f"✅ Session ID: {request.session.session_key}")
            print(f"✅ Session data: {dict(request.session.items())}")
            
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': 'Login successful',
                    'redirect': '/admin/'
                })
            
            messages.success(request, f'Welcome back, {username}!')
            return redirect('admin_dashboard')
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'message': 'Invalid admin credentials'})
            
            messages.error(request, 'Invalid admin credentials')
            return redirect('admin_login')
    
    # GET request - show login page
    return render(request, 'admin_login.html')

def admin_logout_view(request):
    """Admin logout - clear custom session"""
    if is_admin_authenticated(request):
        username = request.session.get('admin_username')
        print(f"✅ Admin {username} logged out")
        request.session.flush()  # Clear the session completely
        messages.success(request, 'Logged out successfully')
    return redirect('admin_login')

# ==================== MAIN ADMIN DASHBOARD ====================

def admin_dashboard_view(request):
    """Main admin dashboard - checks custom session authentication"""
    
    import sys
    print("\n" + "="*50, file=sys.stderr)
    print("ADMIN DASHBOARD ACCESS", file=sys.stderr)
    print("="*50, file=sys.stderr)
    print(f"Session data: {dict(request.session.items())}", file=sys.stderr)
    
    # Check custom authentication
    if not is_admin_authenticated(request):
        print("❌ Not authenticated via custom method", file=sys.stderr)
        return redirect('admin_login')
    
    admin_username = request.session.get('admin_username')
    print(f"✅ Admin {admin_username} authenticated via custom session", file=sys.stderr)
    print("="*50 + "\n", file=sys.stderr)
    
    # Get statistics
    context = {
        # User stats
        'total_users': MfalmeUsers.objects.count(),
        'active_users': MfalmeUsers.objects.filter(account_status='active').count(),
        'pending_users': MfalmeUsers.objects.filter(account_status='pending').count(),
        'verified_users': MfalmeUsers.objects.filter(email_verified=True).count(),
        
        # Revenue stats
        'total_revenue': float(PaymentTransaction.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0),
        'total_transactions': PaymentTransaction.objects.count(),
        'completed_transactions': PaymentTransaction.objects.filter(status='completed').count(),
        
        # Content stats
        'total_videos': TrainingVideo.objects.count(),
        'total_pdfs': PDF.objects.count(),
        'total_courses': Course.objects.count(),
        'total_packages': Package.objects.count(),
        
        # Support stats
        'open_tickets': SupportTicket.objects.filter(status='open').count(),
        'resolved_tickets': SupportTicket.objects.filter(status='resolved').count(),
        
        # Recent data
        'recent_users': MfalmeUsers.objects.order_by('-date_joined')[:5],
        'recent_transactions': PaymentTransaction.objects.select_related('user').order_by('-created_at')[:5],
        'recent_activities': ActivityLog.objects.select_related('user').order_by('-created_at')[:10],
        
        # Today's stats
        'today_users': MfalmeUsers.objects.filter(date_joined__date=timezone.now().date()).count(),
        'today_revenue': float(PaymentTransaction.objects.filter(
            status='completed',
            paid_at__date=timezone.now().date()
        ).aggregate(Sum('amount'))['amount__sum'] or 0),
        
        # Admin info
        'admin_username': admin_username,
        
        # ADD MEDIA_URL for images
        'MEDIA_URL': settings.MEDIA_URL,
    }
    
    return render(request, 'admin.html', context)

# ==================== SESSION DEBUG VIEWS ====================

def debug_session(request):
    """Debug session to check authentication status"""
    from django.http import HttpResponse
    
    html = f"""
    <html>
    <head><title>Session Debug</title></head>
    <body>
        <h1>Session Debug Information</h1>
        <p><strong>Session Key:</strong> {request.session.session_key}</p>
        <p><strong>Session Data:</strong> {dict(request.session.items())}</p>
        <p><strong>Admin Authenticated:</strong> {is_admin_authenticated(request)}</p>
    """
    
    if is_admin_authenticated(request):
        html += f"""
        <p><strong>Admin Username:</strong> {request.session.get('admin_username')}</p>
        <p><a href="/admin/">Go to Admin Dashboard</a></p>
        """
    else:
        html += f"""
        <p><strong>Not logged in</strong></p>
        <p><a href="/admin/login/">Go to Login</a></p>
        """
    
    html += """
        <p><a href="/admin/logout/">Logout</a></p>
        <p><a href="/admin/clear-session/">Clear Session</a></p>
    </body>
    </html>
    """
    
    return HttpResponse(html)

@csrf_exempt
def clear_session(request):
    """Clear session for debugging"""
    if request.method == 'POST':
        request.session.flush()
        return HttpResponse("Session cleared")
    return HttpResponse("""
    <html>
    <body>
        <h1>Clear Session</h1>
        <form method="post">
            <button type="submit">Clear Session</button>
        </form>
        <p><a href="/admin/debug-session/">Back to Debug</a></p>
    </body>
    </html>
    """)

# ==================== TEST VIEW ====================

def test_admin_access(request):
    """Simple test to check if admin is authenticated"""
    from django.http import HttpResponse
    
    if not is_admin_authenticated(request):
        return HttpResponse("""
        <html>
        <head><title>Not Authenticated</title></head>
        <body>
            <h1>Not Authenticated</h1>
            <p>You are not logged in as admin.</p>
            <p><a href="/admin/login/">Go to Login</a></p>
            <p><a href="/admin/debug-session/">Check Session</a></p>
        </body>
        </html>
        """)
    
    return HttpResponse(f"""
    <html>
    <head><title>Admin Test</title></head>
    <body>
        <h1>Admin Access Test</h1>
        <p><strong>Admin Username:</strong> {request.session.get('admin_username')}</p>
        <p><strong>Session ID:</strong> {request.session.session_key}</p>
        <p><strong>Session Data:</strong> {dict(request.session.items())}</p>
        <p><a href="/admin/">Go to Admin Dashboard</a></p>
        <p><a href="/admin/debug-session/">Check Session</a></p>
        <p><a href="/admin/logout/">Logout</a></p>
    </body>
    </html>
    """)

# ==================== ADMIN API - DASHBOARD STATS ====================

@admin_required
def admin_api_dashboard_stats(request):
    """Get real dashboard statistics"""
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    
    # User stats
    total_users = MfalmeUsers.objects.count()
    active_users = MfalmeUsers.objects.filter(account_status='active').count()
    pending_users = MfalmeUsers.objects.filter(account_status='pending').count()
    new_users_30d = MfalmeUsers.objects.filter(date_joined__gte=thirty_days_ago).count()
    
    # Revenue stats
    total_revenue = PaymentTransaction.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    revenue_30d = PaymentTransaction.objects.filter(
        status='completed', 
        paid_at__gte=thirty_days_ago
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Content stats
    videos = TrainingVideo.objects.filter(is_active=True).count()
    pdfs = PDF.objects.filter(is_active=True).count()
    courses = Course.objects.filter(is_active=True).count()
    blogs = Blog.objects.filter(status='published').count()
    
    # Support stats
    open_tickets = SupportTicket.objects.filter(status='open').count()
    in_progress_tickets = SupportTicket.objects.filter(status='in_progress').count()
    
    return JsonResponse({
        'users': {
            'total': total_users,
            'active': active_users,
            'pending': pending_users,
            'new_30d': new_users_30d,
        },
        'revenue': {
            'total': float(total_revenue),
            'total_formatted': f"${float(total_revenue):,.2f}",
            'monthly': float(revenue_30d),
            'monthly_formatted': f"${float(revenue_30d):,.2f}",
        },
        'content': {
            'videos': videos,
            'pdfs': pdfs,
            'courses': courses,
            'blogs': blogs,
        },
        'support': {
            'open': open_tickets,
            'in_progress': in_progress_tickets,
        }
    })

@admin_required
def admin_api_activities(request):
    """Get recent activities"""
    limit = int(request.GET.get('limit', 10))
    
    activities = ActivityLog.objects.select_related('user').order_by('-created_at')[:limit]
    
    data = []
    for act in activities:
        icon_map = {
            'LOGIN': 'sign-in-alt',
            'LOGOUT': 'sign-out-alt',
            'REGISTER': 'user-plus',
            'ADMIN_LOGIN': 'user-tie',
            'PAYMENT_COMPLETED': 'dollar-sign',
            'VIDEO_WATCH': 'play',
            'PDF_DOWNLOAD': 'file-pdf',
            'SUPPORT_TICKET_CREATED': 'ticket-alt',
            'COURSE_ENROLLED': 'graduation-cap',
            'ADMIN_ACTION': 'cog',
        }
        
        type_map = {
            'LOGIN': 'user',
            'LOGOUT': 'user',
            'REGISTER': 'user',
            'ADMIN_LOGIN': 'admin',
            'PAYMENT_COMPLETED': 'payment',
            'VIDEO_WATCH': 'course',
            'PDF_DOWNLOAD': 'pdf',
            'SUPPORT_TICKET_CREATED': 'support',
            'COURSE_ENROLLED': 'course',
            'ADMIN_ACTION': 'admin',
        }
        
        data.append({
            'id': act.id,
            'type': type_map.get(act.action, 'user'),
            'icon': icon_map.get(act.action, 'bell'),
            'title': act.description,
            'time': timesince(act.created_at),
            'user': act.user.username if act.user else 'System',
            'created_at': act.created_at.isoformat(),
        })
    
    return JsonResponse({'activities': data})

# ==================== ADMIN API - USER MANAGEMENT ====================

@admin_required
def admin_api_users(request):
    """Get all users with filters"""
    # Get query parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    rank = request.GET.get('rank', '')
    limit = int(request.GET.get('limit', 50))
    offset = int(request.GET.get('offset', 0))
    
    # Build query
    users = MfalmeUsers.objects.all()
    
    if search:
        users = users.filter(
            Q(email__icontains=search) |
            Q(username__icontains=search) |
            Q(phone__icontains=search) |
            Q(soldier_id__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    if status:
        users = users.filter(account_status=status)
    
    if rank:
        users = users.filter(elite_rank=rank)
    
    total = users.count()
    users = users.order_by('-date_joined')[offset:offset+limit]
    
    data = []
    for user in users:
        profile_image_url = None
        if user.profile_image:
            try:
                profile_image_url = user.profile_image.url
            except:
                profile_image_url = None
        
        data.append({
            'id': user.id,
            'soldier_id': user.soldier_id,
            'name': user.get_full_name() or user.username,
            'username': user.username,
            'email': user.email,
            'phone': user.phone,
            'rank': user.elite_rank,
            'status': user.account_status,
            'tier': user.elite_rank.lower(),
            'joined': user.date_joined.strftime('%Y-%m-%d'),
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else None,
            'email_verified': user.email_verified,
            'country': user.country,
            'referral_count': user.referral_count,
            'total_deposits': float(user.total_deposits),
            'profile_image': profile_image_url,
        })
    
    return JsonResponse({
        'total': total,
        'users': data
    })

@admin_required
def admin_api_user_detail(request, user_id):
    """Get single user details"""
    user = get_object_or_404(MfalmeUsers, id=user_id)
    
    # Get user stats
    transactions = PaymentTransaction.objects.filter(user=user)
    videos_watched = UserVideoAccess.objects.filter(user=user).count()
    pdfs_downloaded = UserPDFAccess.objects.filter(user=user, downloaded=True).count()
    courses_enrolled = UserCourse.objects.filter(user=user).count()
    
    profile_image_url = None
    if user.profile_image:
        try:
            profile_image_url = user.profile_image.url
        except:
            profile_image_url = None
    
    data = {
        'id': user.id,
        'soldier_id': user.soldier_id,
        'username': user.username,
        'email': user.email,
        'phone': user.phone,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'full_name': user.get_full_name(),
        'elite_rank': user.elite_rank,
        'account_status': user.account_status,
        'email_verified': user.email_verified,
        'country': user.country,
        'city': user.city,
        'address': user.address,
        'whatsapp_number': user.whatsapp_number,
        'telegram_username': user.telegram_username,
        'trading_experience': user.trading_experience,
        'account_balance': float(user.account_balance),
        'referral_code': user.referral_code,
        'referral_count': user.referral_count,
        'referral_earnings': float(user.referral_earnings),
        'total_deposits': float(user.total_deposits),
        'total_withdrawals': float(user.total_withdrawals),
        'total_profit': float(user.total_profit),
        'success_rate': user.success_rate,
        'bio': user.bio,
        'admin_notes': user.admin_notes,
        'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
        'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else None,
        'verified_at': user.verified_at.strftime('%Y-%m-%d') if user.verified_at else None,
        'registration_ip': user.registration_ip,
        'profile_image': profile_image_url,
        
        # Stats
        'transactions_count': transactions.count(),
        'transactions_total': float(transactions.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0),
        'videos_watched': videos_watched,
        'pdfs_downloaded': pdfs_downloaded,
        'courses_enrolled': courses_enrolled,
        'support_tickets': SupportTicket.objects.filter(user=user).count(),
    }
    
    return JsonResponse(data)

@admin_required
@csrf_exempt
def admin_api_user_create(request):
    """Create a new user"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    # Check if email exists
    if MfalmeUsers.objects.filter(email=data.get('email')).exists():
        return JsonResponse({'error': 'Email already exists'}, status=400)
    
    # Check if username exists
    if MfalmeUsers.objects.filter(username=data.get('username')).exists():
        return JsonResponse({'error': 'Username already exists'}, status=400)
    
    # Create user
    user = MfalmeUsers.objects.create_user(
        email=data.get('email'),
        password=data.get('password'),
        username=data.get('username'),
        phone=data.get('phone', ''),
        first_name=data.get('first_name', ''),
        last_name=data.get('last_name', ''),
        country=data.get('country', ''),
        city=data.get('city', ''),
        account_status=data.get('account_status', 'active'),
        elite_rank=data.get('elite_rank', 'Recruit'),
        email_verified=data.get('email_verified', True),
    )
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Created user: {user.email}',
        metadata={
            'admin_username': request.session.get('admin_username'),
            'user_id': user.id
        }
    )
    
    return JsonResponse({
        'success': True,
        'id': user.id,
        'message': 'User created successfully'
    })

@admin_required
@csrf_exempt
def admin_api_user_update(request, user_id):
    """Update user"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    user = get_object_or_404(MfalmeUsers, id=user_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    # Update fields
    user.first_name = data.get('first_name', user.first_name)
    user.last_name = data.get('last_name', user.last_name)
    user.phone = data.get('phone', user.phone)
    user.country = data.get('country', user.country)
    user.city = data.get('city', user.city)
    user.address = data.get('address', user.address)
    user.whatsapp_number = data.get('whatsapp_number', user.whatsapp_number)
    user.telegram_username = data.get('telegram_username', user.telegram_username)
    user.elite_rank = data.get('elite_rank', user.elite_rank)
    user.account_status = data.get('account_status', user.account_status)
    user.email_verified = data.get('email_verified', user.email_verified)
    user.bio = data.get('bio', user.bio)
    user.admin_notes = data.get('admin_notes', user.admin_notes)
    
    # Update password if provided
    if data.get('password'):
        user.set_password(data['password'])
    
    user.save()
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Updated user: {user.email}',
        metadata={
            'admin_username': request.session.get('admin_username'),
            'user_id': user.id
        }
    )
    
    return JsonResponse({'success': True, 'message': 'User updated successfully'})

@admin_required
@require_POST
def admin_api_user_delete(request, user_id):
    """Delete user"""
    user = get_object_or_404(MfalmeUsers, id=user_id)
    
    # Get admin user before potentially deleting
    admin_user = get_admin_user(request)
    
    # Don't allow deleting yourself (check against admin user)
    if admin_user and user.id == admin_user.id:
        return JsonResponse({'error': 'Cannot delete yourself'}, status=400)
    
    email = user.email
    user.delete()
    
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted user: {email}',
        metadata={
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({'success': True})

@admin_required
@require_POST
def admin_api_user_activate(request, user_id):
    """Activate/deactivate user"""
    user = get_object_or_404(MfalmeUsers, id=user_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = {}
    
    activate = data.get('activate', True)
    
    user.account_status = 'active' if activate else 'inactive'
    user.save(update_fields=['account_status'])
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f"{'Activated' if activate else 'Deactivated'} user: {user.email}",
        metadata={
            'admin_username': request.session.get('admin_username'),
            'user_id': user.id
        }
    )
    
    return JsonResponse({'success': True})

# ==================== ADMIN API - COURSE MANAGEMENT ====================

@admin_required
def admin_api_courses(request):
    """Get all courses with proper thumbnail URLs"""
    try:
        courses = Course.objects.all().order_by('-created_at')
        data = []
        
        for course in courses:
            try:
                # Get thumbnail URL - use the property that returns full S3 URL
                thumbnail_url = course.thumbnail_url if hasattr(course, 'thumbnail_url') else None
                
                # If thumbnail_url is None but thumbnail exists, construct S3 URL manually
                if not thumbnail_url and course.thumbnail and course.thumbnail.name:
                    from django.conf import settings
                    if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') and settings.AWS_S3_CUSTOM_DOMAIN:
                        thumbnail_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{course.thumbnail.name}"
                    else:
                        thumbnail_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{course.thumbnail.name}"
                
                data.append({
                    'id': course.id,
                    'name': str(course.title),
                    'title': str(course.title),
                    'code': f"CRS-{str(course.id).zfill(4)}",
                    'price': float(course.price),
                    'price_formatted': f"${float(course.price):,.2f}",
                    'videos': course.video_count(),
                    'pdfs': course.pdf_count(),
                    'status': 'active' if course.is_active else 'inactive',
                    'description': course.description or "",
                    'created_at': course.created_at.strftime('%Y-%m-%d') if course.created_at else "",
                    'duration_weeks': course.duration_weeks or 4,
                    'thumbnail': thumbnail_url,  # ← This should now be full URL
                })
            except Exception as e:
                print(f"⚠️ Error processing course {course.id}: {e}")
                continue
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        print(f"❌ Error: {e}")
        return JsonResponse([], safe=False)


@admin_required
def admin_api_course_detail(request, course_id):
    """Get single course details with all related content"""
    course = get_object_or_404(Course, id=course_id)
    
    # Get videos in this course
    videos = course.videos.all().values(
        'id', 'title', 'description', 'duration', 'order', 
        'is_active', 'view_count', 'created_at'
    ).order_by('order')
    
    # Get PDFs in this course
    pdfs = course.pdf_resources.all().values(
        'id', 'title', 'description', 'pages', 'file_size', 
        'order', 'downloads', 'is_active', 'created_at'
    ).order_by('order')
    
    # Get enrollment stats
    total_enrollments = UserCourse.objects.filter(course=course).count()
    active_enrollments = UserCourse.objects.filter(
        course=course, 
        access_expires_at__gt=timezone.now()
    ).count()
    
    # Calculate total revenue from this course
    total_revenue = PaymentTransaction.objects.filter(
        course=course,
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Get thumbnail URL
    thumbnail_url = None
    if hasattr(course, 'thumbnail') and course.thumbnail:
        try:
            thumbnail_url = course.thumbnail.url
        except:
            thumbnail_url = None
    
    data = {
        'id': course.id,
        'title': course.title,
        'description': course.description,
        # Price fields
        'price': float(course.price),
        'price_formatted': f"${float(course.price):,.2f}",
        # For backward compatibility
        'price_1_month': float(course.price),
        'price_12_months': float(course.price) * 10,  # Just for compatibility
        # Course details
        'materials': course.materials,
        'duration_weeks': course.duration_weeks,
        'is_active': course.is_active,
        'status': 'active' if course.is_active else 'inactive',
        'created_at': course.created_at.strftime('%Y-%m-%d'),
        'created_at_full': course.created_at.strftime('%Y-%m-%d %H:%M'),
        # Thumbnail
        'thumbnail': thumbnail_url,
        # Related content
        'videos': list(videos),
        'pdfs': list(pdfs),
        'videos_count': len(videos),
        'pdfs_count': len(pdfs),
        # Stats
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
        'total_revenue': float(total_revenue),
        'total_revenue_formatted': f"${float(total_revenue):,.2f}",
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_course_create(request):
    """Create a new course - OPTIMIZED FOR S3 DIRECT UPLOADS"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    print("\n=== COURSE CREATE DEBUG ===")
    print(f"Content type: {request.content_type}")
    
    try:
        # ========== HANDLE JSON DATA WITH S3 KEYS (NEW FLOW - PREFERRED) ==========
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
            print(f"📦 JSON data received: {data}")
            
            # Get data from JSON
            title = data.get('title') or data.get('name', '')
            description = data.get('description', '')
            price = data.get('price', '0')
            duration_weeks = data.get('duration_weeks', 4)
            thumbnail_key = data.get('thumbnail_key')  # S3 key from direct upload
            
            print(f"Title: {title}")
            print(f"Price: {price}")
            print(f"Thumbnail S3 key: {thumbnail_key}")
            
            # Validate required fields
            if not title:
                return JsonResponse({'error': 'Course title is required'}, status=400)
            
            if not thumbnail_key:
                return JsonResponse({'error': 'Thumbnail S3 key is required'}, status=400)
            
            # Create course instance
            course = Course(
                title=title,
                description=description,
                price=Decimal(str(price)) if price else Decimal('0'),
                duration_weeks=int(duration_weeks) if duration_weeks else 4,
                is_active=True,
            )
            
            # CRITICAL: Store the S3 key in both fields for compatibility
            # 1. Set thumbnail_s3_key (if your model has this field - RECOMMENDED)
            if hasattr(course, 'thumbnail_s3_key'):
                course.thumbnail_s3_key = thumbnail_key
                print(f"✅ Set thumbnail_s3_key: {thumbnail_key}")
            
            # 2. Set thumbnail.name for Django's FileField (backward compatibility)
            course.thumbnail.name = thumbnail_key
            print(f"✅ Set thumbnail.name: {course.thumbnail.name}")
            
            # Save the course with bypass_validation
            # This prevents Django from trying to validate the file existence
            course.save(bypass_validation=True)
            print(f"✅ Course saved with ID: {course.id}")
            
            # Log activity
            admin_user = get_admin_user(request)
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created course: {course.title} (S3 upload)',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'course_id': course.id,
                    'thumbnail_key': thumbnail_key
                }
            )
            
            # Generate thumbnail URL for response
            from django.conf import settings
            if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') and settings.AWS_S3_CUSTOM_DOMAIN:
                thumbnail_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{thumbnail_key}"
            else:
                thumbnail_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{thumbnail_key}"
            
            return JsonResponse({
                'success': True, 
                'id': course.id,
                'title': course.title,
                'message': 'Course created successfully with S3 thumbnail',
                'thumbnail_url': thumbnail_url,
                'thumbnail_key': thumbnail_key
            })
        
        # ========== HANDLE MULTIPART/FORM-DATA (LEGACY FLOW) ==========
        elif request.content_type and 'multipart/form-data' in request.content_type:
            print(f"📦 Form data received: {request.POST}")
            print(f"FILES: {request.FILES}")
            
            # Get data from FormData
            title = request.POST.get('name', request.POST.get('title', ''))
            description = request.POST.get('description', '')
            price = request.POST.get('price', '0')
            duration_weeks = request.POST.get('duration_weeks', 4)
            
            print(f"Title: {title}")
            print(f"Price: {price}")
            
            # Validate required fields
            if not title:
                return JsonResponse({'error': 'Course title is required'}, status=400)
            
            # Create course
            course = Course(
                title=title,
                description=description,
                price=Decimal(str(price)) if price else Decimal('0'),
                duration_weeks=int(duration_weeks),
                is_active=True,
            )
            
            # Handle thumbnail upload if present
            if 'thumbnail' in request.FILES:
                thumbnail = request.FILES['thumbnail']
                print(f"Thumbnail received: {thumbnail.name}")
                print(f"Thumbnail size: {thumbnail.size}")
                print(f"Thumbnail content type: {thumbnail.content_type}")
                
                # Sanitize filename
                thumbnail.name = sanitize_filename(thumbnail.name)
                
                # Save the thumbnail (this will use Django's storage)
                course.thumbnail = thumbnail
                print(f"Thumbnail saved with name: {thumbnail.name}")
            
            course.save()
            print(f"✅ Course saved with ID: {course.id}")
            
            # Log activity
            admin_user = get_admin_user(request)
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created course: {course.title} (form upload)',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'course_id': course.id
                }
            )
            
            # Return thumbnail URL
            thumbnail_url = None
            if course.thumbnail:
                try:
                    thumbnail_url = course.thumbnail.url
                except:
                    thumbnail_url = None
            
            return JsonResponse({
                'success': True, 
                'id': course.id,
                'title': course.title,
                'message': 'Course created successfully',
                'thumbnail_url': thumbnail_url
            })
        
        # ========== UNSUPPORTED CONTENT TYPE ==========
        else:
            return JsonResponse({
                'error': f'Unsupported content type: {request.content_type}. Expected application/json or multipart/form-data'
            }, status=400)
            
    except json.JSONDecodeError as e:
        print(f"❌ JSON decode error: {e}")
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)
    
@admin_required
def debug_courses_api(request):
    """Debug endpoint to see raw course data"""
    from django.http import HttpResponse
    import traceback
    from .models import Course
    
    output = "<h1>Courses API Debug</h1>"
    
    try:
        courses = Course.objects.all()
        output += f"<p>Total courses in DB: {courses.count()}</p>"
        
        for course in courses:
            output += f"<h3>Course ID: {course.id}</h3>"
            output += f"<p>Title: {course.title}</p>"
            output += f"<p>Price: {course.price}</p>"
            
            # Test thumbnail
            if hasattr(course, 'thumbnail') and course.thumbnail:
                try:
                    thumbnail_url = course.thumbnail.url
                    output += f"<p>thumbnail: {thumbnail_url} ✅</p>"
                except Exception as e:
                    output += f"<p>thumbnail URL error: {str(e)} ❌</p>"
            elif hasattr(course, 'thumbnail_s3_key') and course.thumbnail_s3_key:
                output += f"<p>thumbnail_s3_key: {course.thumbnail_s3_key} ✅</p>"
            else:
                output += f"<p>thumbnail: None</p>"
            
            # Test thumbnail_url property
            if hasattr(course, 'thumbnail_url'):
                try:
                    thumb_url = course.thumbnail_url
                    output += f"<p>thumbnail_url property: {thumb_url}</p>"
                except Exception as e:
                    output += f"<p>thumbnail_url property error: {str(e)} ❌</p>"
            
            # Test video_count
            try:
                v_count = course.video_count()
                output += f"<p>video_count(): {v_count} ✅</p>"
            except Exception as e:
                output += f"<p>video_count() ERROR: {str(e)} ❌</p>"
            
            # Test pdf_count
            try:
                p_count = course.pdf_count()
                output += f"<p>pdf_count(): {p_count} ✅</p>"
            except Exception as e:
                output += f"<p>pdf_count() ERROR: {str(e)} ❌</p>"
            
            output += "<hr>"
        
    except Exception as e:
        output += f"<p>FATAL ERROR: {str(e)}</p>"
        output += f"<pre>{traceback.format_exc()}</pre>"
    
    return HttpResponse(output)    


@admin_required
@csrf_exempt
def admin_api_course_update(request, course_id):
    """Update course - handles both JSON and multipart/form-data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    course = get_object_or_404(Course, id=course_id)
    
    # Track changes for logging
    changes = []
    
    try:
        # Handle multipart/form-data (for file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            print(f"🔵 Handling multipart form data update for course {course_id}")
            
            # Update fields from POST data
            if 'title' in request.POST:
                new_title = request.POST.get('title')
                if new_title and new_title != course.title:
                    changes.append(f'title: {course.title} → {new_title}')
                    course.title = new_title
            elif 'name' in request.POST:
                new_title = request.POST.get('name')
                if new_title and new_title != course.title:
                    changes.append(f'title: {course.title} → {new_title}')
                    course.title = new_title
            
            if 'description' in request.POST:
                course.description = request.POST.get('description', '')
            
            if 'price' in request.POST:
                old_price = float(course.price)
                new_price = float(request.POST.get('price', 0))
                if old_price != new_price:
                    changes.append(f'price: ${old_price} → ${new_price}')
                course.price = Decimal(str(request.POST.get('price', 0)))
            
            if 'duration_weeks' in request.POST:
                course.duration_weeks = int(request.POST.get('duration_weeks', 4))
            
            if 'is_active' in request.POST:
                is_active = request.POST.get('is_active') == 'true'
                old_status = 'active' if course.is_active else 'inactive'
                new_status = 'active' if is_active else 'inactive'
                if old_status != new_status:
                    changes.append(f'status: {old_status} → {new_status}')
                course.is_active = is_active
            
            # Handle thumbnail upload if present
            if 'thumbnail' in request.FILES:
                thumbnail = request.FILES['thumbnail']
                # Sanitize filename
                thumbnail.name = sanitize_filename(thumbnail.name)
                course.thumbnail = thumbnail
                changes.append('thumbnail: updated')
                print(f"✅ Thumbnail uploaded: {thumbnail.name}")
            
            course.save()
            
        else:
            # Handle JSON data
            try:
                data = json.loads(request.body)
                print(f"🔵 Handling JSON update for course {course_id}: {data}")
                
                # Update fields if provided
                if 'title' in data and data['title'] != course.title:
                    changes.append(f'title: {course.title} → {data["title"]}')
                    course.title = data['title']
                
                if 'description' in data:
                    course.description = data['description']
                
                if 'price' in data:
                    old_price = float(course.price)
                    new_price = float(data['price'])
                    if old_price != new_price:
                        changes.append(f'price: ${old_price} → ${new_price}')
                    course.price = Decimal(str(data['price']))
                
                if 'duration_weeks' in data:
                    course.duration_weeks = int(data['duration_weeks'])
                
                if 'is_active' in data:
                    old_status = 'active' if course.is_active else 'inactive'
                    new_status = 'active' if data['is_active'] else 'inactive'
                    if old_status != new_status:
                        changes.append(f'status: {old_status} → {new_status}')
                    course.is_active = data['is_active']
                
                if 'materials' in data:
                    course.materials = data['materials']
                
                course.save()
                
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
        
        # Log activity
        admin_user = get_admin_user(request)
        if changes:
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Updated course: {course.title}',
                metadata={
                    'course_id': course.id,
                    'changes': changes,
                    'admin_username': request.session.get('admin_username')
                }
            )
            print(f"✅ Course updated with changes: {changes}")
        else:
            print("ℹ️ No changes detected")
        
        # Get thumbnail URL
        thumbnail_url = None
        if course.thumbnail:
            try:
                thumbnail_url = course.thumbnail.url
            except:
                thumbnail_url = None
        
        # Return the updated course data
        return JsonResponse({
            'success': True, 
            'message': 'Course updated successfully',
            'changes': changes,
            'course': {
                'id': course.id,
                'title': course.title,
                'price': float(course.price),
                'thumbnail': thumbnail_url
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating course: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@require_POST
def admin_api_course_delete(request, course_id):
    """Delete course and handle related content"""
    course = get_object_or_404(Course, id=course_id)
    title = course.title
    
    # Check if course has enrollments
    enrollments_count = UserCourse.objects.filter(course=course).count()
    if enrollments_count > 0:
        return JsonResponse({
            'error': f'Cannot delete course with {enrollments_count} active enrollments'
        }, status=400)
    
    # Get counts for logging
    videos_count = course.videos.count()
    pdfs_count = course.pdf_resources.count()
    
    # Delete the course
    course.delete()
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted course: {title}',
        metadata={
            'videos_affected': videos_count,
            'pdfs_affected': pdfs_count,
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({
        'success': True,
        'message': f'Course "{title}" deleted successfully'
    })


@admin_required
def admin_api_course_stats(request, course_id):
    """Get detailed statistics for a specific course"""
    course = get_object_or_404(Course, id=course_id)
    
    # Enrollment stats
    total_enrollments = UserCourse.objects.filter(course=course)
    active_enrollments = total_enrollments.filter(access_expires_at__gt=timezone.now())
    expired_enrollments = total_enrollments.filter(access_expires_at__lte=timezone.now())
    
    # Revenue stats
    total_revenue = PaymentTransaction.objects.filter(
        course=course,
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Monthly revenue for this course (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        month = timezone.now() - timedelta(days=30*i)
        month_start = month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i > 0:
            month_end = (month + timedelta(days=30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
        else:
            month_end = timezone.now()
        
        revenue = PaymentTransaction.objects.filter(
            course=course,
            status='completed',
            paid_at__range=[month_start, month_end]
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        monthly_revenue.append({
            'month': month.strftime('%B %Y'),
            'revenue': float(revenue)
        })
    
    # Content engagement
    video_views = sum([v.view_count for v in course.videos.all()])
    pdf_downloads = sum([p.downloads for p in course.pdf_resources.all()])
    
    data = {
        'course_id': course.id,
        'course_title': course.title,
        'enrollments': {
            'total': total_enrollments.count(),
            'active': active_enrollments.count(),
            'expired': expired_enrollments.count(),
        },
        'revenue': {
            'total': float(total_revenue),
            'total_formatted': f"${float(total_revenue):,.2f}",
            'monthly': monthly_revenue,
            'average_per_enrollment': float(total_revenue / total_enrollments.count()) if total_enrollments.count() > 0 else 0,
        },
        'content': {
            'videos': course.video_count(),
            'video_views': video_views,
            'pdfs': course.pdf_count(),
            'pdf_downloads': pdf_downloads,
        }
    }
    
    return JsonResponse(data)


@admin_required
def admin_api_course_enrollments(request, course_id):
    """Get all enrollments for a specific course"""
    course = get_object_or_404(Course, id=course_id)
    
    enrollments = UserCourse.objects.filter(course=course).select_related('user', 'payment').order_by('-enrolled_at')
    
    data = []
    for enrollment in enrollments:
        data.append({
            'id': enrollment.id,
            'user_id': enrollment.user.id,
            'user_name': enrollment.user.get_full_name() or enrollment.user.username,
            'user_email': enrollment.user.email,
            'purchase_type': enrollment.purchase_type,
            'enrolled_at': enrollment.enrolled_at.strftime('%Y-%m-%d %H:%M'),
            'expires_at': enrollment.access_expires_at.strftime('%Y-%m-%d') if enrollment.access_expires_at else 'Never',
            'is_active': enrollment.access_expires_at > timezone.now() if enrollment.access_expires_at else True,
            'progress': enrollment.progress,
            'payment_id': enrollment.payment.id if enrollment.payment else None,
            'payment_amount': float(enrollment.payment.amount) if enrollment.payment else 0,
        })
    
    return JsonResponse({
        'course_id': course.id,
        'course_title': course.title,
        'total_enrollments': len(data),
        'enrollments': data
    })

# ==================== ADMIN API - VIDEO MANAGEMENT (FIXED) ====================

@admin_required
def admin_api_videos(request):
    """Get all videos with thumbnails"""
    videos = TrainingVideo.objects.select_related('course').order_by('-created_at')
    
    data = []
    for video in videos:
        data.append({
            'id': video.id,
            'title': video.title,
            'description': video.description[:100] + '...' if len(video.description) > 100 else video.description,
            'course_id': video.course.id if video.course else None,
            'course_name': video.course.title if video.course else 'Uncategorized',
            'category': video.category,
            'duration': f"{video.duration} min",
            'price': float(video.price),
            'view_count': video.view_count,
            'thumbnail': video.thumbnail_url,  # Use the property
            'is_active': video.is_active,
            'order': video.order,
            'created_at': video.created_at.strftime('%Y-%m-%d'),
            'uploaded': video.created_at.strftime('%Y-%m-%d'),
            'status': 'published' if video.is_active else 'draft',
        })
    
    return JsonResponse(data, safe=False)

@admin_required
def admin_api_video_detail(request, video_id):
    """Get single video details"""
    video = get_object_or_404(TrainingVideo, id=video_id)
    
    thumbnail_url = None
    if video.thumbnail:
        try:
            thumbnail_url = video.thumbnail.url
        except:
            thumbnail_url = None
    
    data = {
        'id': video.id,
        'title': video.title,
        'description': video.description,
        'course_id': video.course.id if video.course else None,
        'course_name': video.course.title if video.course else None,
        'category': video.category,
        'duration': f"{video.duration} min",
        'price': float(video.price),
        'video_url': video.video_url,
        'thumbnail': thumbnail_url,
        'allow_download': video.allow_download,
        'disable_screenshots': video.disable_screenshots,
        'order': video.order,
        'is_active': video.is_active,
        'view_count': video.view_count,
        'created_at': video.created_at.strftime('%Y-%m-%d %H:%M'),
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_video_create(request):
    """Create a new video - OPTIMIZED FOR S3 DIRECT UPLOADS"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # ==================== HANDLE JSON DATA (S3 DIRECT UPLOADS) ====================
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
            print(f"📹 Creating video from JSON (S3 direct upload): {data}")
            
            # ===== REQUIRED FIELDS =====
            title = data.get('title')
            s3_key = data.get('s3_key') or data.get('video_key')  # Consistent naming
            
            if not title:
                return JsonResponse({'error': 'Title is required'}, status=400)
            
            if not s3_key:
                return JsonResponse({'error': 'Video S3 key is required (s3_key)'}, status=400)
            
            # ===== OPTIONAL FIELDS =====
            course_id = data.get('course_id')
            duration = data.get('duration', 30)
            price = Decimal(str(data.get('price', 0)))
            category = data.get('category', 'PTM')
            description = data.get('description', '')
            module = data.get('module', '')
            order = data.get('order', 0)
            is_active = data.get('is_active', True)
            
            # ===== THUMBNAIL HANDLING =====
            thumbnail_key = data.get('thumbnail_key') or data.get('thumbnail_s3_key')
            
            # ===== CREATE VIDEO INSTANCE =====
            video = TrainingVideo(
                title=title,
                description=description,
                category=category,
                duration=int(duration),
                price=price,
                order=int(order),
                is_active=is_active,
            )
            
            # CRITICAL: Store S3 keys - ONLY set s3_key, NOT video_url (it's a property)
            video.s3_key = s3_key
            # ❌ REMOVED: video.video_url = s3_key  ← This line caused the AttributeError
            
            # Handle thumbnail S3 key
            if thumbnail_key:
                video.thumbnail_s3_key = thumbnail_key
                # Also set the thumbnail field for Django compatibility
                video.thumbnail.name = thumbnail_key
            
            # Set module if provided (assuming your model has this field)
            if hasattr(video, 'module') and module:
                video.module = module
            
            # Set course if provided
            if course_id and course_id != 'null' and course_id != '':
                try:
                    video.course = Course.objects.get(id=int(course_id))
                    print(f"📚 Linked to course: {video.course.title}")
                except (Course.DoesNotExist, ValueError) as e:
                    print(f"⚠️ Course not found: {e}")
            
            # Save with bypass_validation=True for S3 uploads
            # This prevents Django from trying to validate the file existence
            video.save(bypass_validation=True)
            
            print(f"✅ Video created successfully with ID: {video.id}")
            print(f"   S3 Key: {s3_key}")
            print(f"   Thumbnail Key: {thumbnail_key}")
            
            # ===== LOG ACTIVITY =====
            admin_user = get_admin_user(request)
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created video: {video.title} (S3 direct upload)',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'video_id': video.id,
                    's3_key': s3_key,
                    'thumbnail_key': thumbnail_key
                }
            )
            
            # ===== GENERATE VIDEO URL FOR RESPONSE =====
            from django.conf import settings
            if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') and settings.AWS_S3_CUSTOM_DOMAIN:
                video_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{s3_key}"
                thumbnail_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{thumbnail_key}" if thumbnail_key else None
            else:
                video_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}"
                thumbnail_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{thumbnail_key}" if thumbnail_key else None
            
            return JsonResponse({
                'success': True,
                'id': video.id,
                'title': video.title,
                'message': 'Video created successfully from S3 upload',
                'video_url': video_url,
                'thumbnail_url': thumbnail_url,
                's3_key': s3_key,
                'thumbnail_key': thumbnail_key
            })
        
        # ==================== HANDLE MULTIPART/FORM-DATA (LEGACY DIRECT UPLOADS) ====================
        elif request.content_type and 'multipart/form-data' in request.content_type:
            print(f"📹 Creating video from form data (legacy upload): {request.POST}")
            
            # ===== REQUIRED FIELDS =====
            title = request.POST.get('title')
            
            if not title:
                return JsonResponse({'error': 'Title is required'}, status=400)
            
            # ===== OPTIONAL FIELDS =====
            description = request.POST.get('description', '')
            category = request.POST.get('category', 'PTM')
            duration = int(request.POST.get('duration', 30))
            price = Decimal(str(request.POST.get('price', 0)))
            allow_download = request.POST.get('allow_download') == 'true'
            disable_screenshots = request.POST.get('disable_screenshots') == 'true'
            order = int(request.POST.get('order', 0))
            is_active = request.POST.get('is_active', 'true') == 'true'
            course_id = request.POST.get('course_id')
            module = request.POST.get('module', '')
            
            # Get uploaded files
            video_file = request.FILES.get('video_file')
            thumbnail_file = request.FILES.get('thumbnail_file')
            
            # Validate files
            if not video_file and not request.POST.get('video_url'):
                return JsonResponse({'error': 'Either video file or video URL is required'}, status=400)
            
            # ===== CREATE VIDEO INSTANCE =====
            video = TrainingVideo(
                title=title,
                description=description,
                category=category,
                duration=duration,
                price=price,
                allow_download=allow_download,
                disable_screenshots=disable_screenshots,
                order=order,
                is_active=is_active,
            )
            
            # Set module if field exists
            if hasattr(video, 'module') and module:
                video.module = module
            
            # ===== HANDLE VIDEO FILE/UPLOAD =====
            if video_file:
                # For direct uploads, Django will handle the file
                video.video_file = video_file
                print(f"📁 Video file attached: {video_file.name}")
            elif request.POST.get('video_url'):
                # For external URLs (YouTube, etc.)
                video.video_url = request.POST.get('video_url')
            
            # ===== HANDLE THUMBNAIL =====
            if thumbnail_file:
                video.thumbnail = thumbnail_file
                print(f"🖼️ Thumbnail file attached: {thumbnail_file.name}")
            
            # ===== SET COURSE =====
            if course_id and course_id != 'null' and course_id != '':
                try:
                    video.course = Course.objects.get(id=int(course_id))
                    print(f"📚 Linked to course: {video.course.title}")
                except (Course.DoesNotExist, ValueError) as e:
                    print(f"⚠️ Course not found: {e}")
            
            # ===== SAVE VIDEO =====
            video.save()  # Normal save with validation
            print(f"✅ Video created successfully with ID: {video.id}")
            
            # ===== LOG ACTIVITY =====
            admin_user = get_admin_user(request)
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created video: {video.title} (form upload)',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'video_id': video.id,
                    'has_video_file': bool(video_file),
                    'has_thumbnail': bool(thumbnail_file)
                }
            )
            
            # ===== GENERATE RESPONSE URLS =====
            response_data = {
                'success': True,
                'id': video.id,
                'title': video.title,
                'message': 'Video created successfully',
            }
            
            # Add video URL if available
            if video.video_url:
                response_data['video_url'] = video.video_url
            
            # Add thumbnail URL if available
            if video.thumbnail and hasattr(video.thumbnail, 'url'):
                try:
                    response_data['thumbnail_url'] = video.thumbnail.url
                except:
                    pass
            
            return JsonResponse(response_data)
        
        # ==================== UNSUPPORTED CONTENT TYPE ====================
        else:
            return JsonResponse({
                'error': f'Unsupported content type: {request.content_type}. Expected application/json or multipart/form-data'
            }, status=400)
    
    except json.JSONDecodeError as e:
        print(f"❌ JSON decode error: {e}")
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    
    except Exception as e:
        print(f"❌ Error creating video: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@admin_required
@csrf_exempt
def admin_api_video_update(request, video_id):
    """Update video - handles both JSON and multipart/form-data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    video = get_object_or_404(TrainingVideo, id=video_id)
    
    # Track changes for logging
    changes = []
    
    try:
        # Handle multipart/form-data (for file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            print(f"🔵 Handling multipart form data update for video {video_id}")
            
            # Update fields from POST data
            if 'title' in request.POST:
                new_title = request.POST.get('title')
                if new_title != video.title:
                    changes.append(f'title: {video.title} → {new_title}')
                    video.title = new_title
            
            if 'description' in request.POST:
                video.description = request.POST.get('description', video.description)
            
            if 'category' in request.POST:
                new_category = request.POST.get('category')
                if new_category != video.category:
                    changes.append(f'category: {video.category} → {new_category}')
                    video.category = new_category
            
            if 'duration' in request.POST:
                new_duration = int(request.POST.get('duration', video.duration))
                if new_duration != video.duration:
                    changes.append(f'duration: {video.duration} → {new_duration}')
                    video.duration = new_duration
            
            if 'price' in request.POST:
                old_price = float(video.price)
                new_price = float(request.POST.get('price', old_price))
                if old_price != new_price:
                    changes.append(f'price: ${old_price} → ${new_price}')
                video.price = Decimal(str(new_price))
            
            if 'video_url' in request.POST:
                video.video_url = request.POST.get('video_url', video.video_url)
            
            if 'order' in request.POST:
                new_order = int(request.POST.get('order', video.order))
                if new_order != video.order:
                    changes.append(f'order: {video.order} → {new_order}')
                    video.order = new_order
            
            if 'is_active' in request.POST:
                is_active = request.POST.get('is_active', 'true') == 'true'
                old_status = 'active' if video.is_active else 'inactive'
                new_status = 'active' if is_active else 'inactive'
                if old_status != new_status:
                    changes.append(f'status: {old_status} → {new_status}')
                video.is_active = is_active
            
            if 'allow_download' in request.POST:
                allow_download = request.POST.get('allow_download') == 'true'
                video.allow_download = allow_download
            
            if 'disable_screenshots' in request.POST:
                disable_screenshots = request.POST.get('disable_screenshots') == 'true'
                video.disable_screenshots = disable_screenshots
            
            # Update course if provided
            if 'course_id' in request.POST:
                course_id = request.POST.get('course_id')
                old_course = video.course.title if video.course else 'None'
                if course_id:
                    try:
                        new_course = Course.objects.get(id=course_id)
                        video.course = new_course
                        changes.append(f'course: {old_course} → {new_course.title}')
                    except Course.DoesNotExist:
                        pass
                else:
                    video.course = None
                    changes.append(f'course: {old_course} → None')
            
            # Handle thumbnail upload if present
            if 'thumbnail' in request.FILES:
                thumbnail = request.FILES['thumbnail']
                thumbnail.name = sanitize_filename(thumbnail.name)
                video.thumbnail = thumbnail
                changes.append('thumbnail: updated')
                print(f"✅ Thumbnail uploaded: {thumbnail.name}")
            
            video.save()
            
        else:
            # Handle JSON data (for S3 updates)
            try:
                data = json.loads(request.body)
                print(f"🔵 Handling JSON update for video {video_id}")
                print(f"📦 Update data: {data}")
                
                changes_list = []
                
                # Handle S3 thumbnail key update
                if 'thumbnail_key' in data or 'thumbnail_s3_key' in data:
                    thumbnail_key = data.get('thumbnail_key') or data.get('thumbnail_s3_key')
                    if thumbnail_key and thumbnail_key != video.thumbnail_s3_key:
                        video.thumbnail_s3_key = thumbnail_key
                        changes_list.append('thumbnail_s3_key: updated')
                        print(f"✅ Thumbnail S3 key updated: {thumbnail_key}")
                
                # Handle S3 video key update
                if 's3_key' in data and data['s3_key'] != video.s3_key:
                    video.s3_key = data['s3_key']
                    video.video_url = data['s3_key']  # Keep in sync
                    changes_list.append(f's3_key: updated')
                    print(f"✅ Video S3 key updated: {data['s3_key']}")
                
                # Handle video_url update (could be S3 key or external URL)
                if 'video_url' in data and data['video_url'] != video.video_url:
                    old_url = video.video_url
                    video.video_url = data['video_url']
                    # If it looks like an S3 key, also update s3_key
                    if 's3.amazonaws.com' not in data['video_url'] and not data['video_url'].startswith('http'):
                        video.s3_key = data['video_url']
                    changes_list.append(f'video_url: {old_url} → {data["video_url"]}')
                
                # Handle other field updates
                if 'title' in data and data['title'] != video.title:
                    changes_list.append(f'title: {video.title} → {data["title"]}')
                    video.title = data['title']
                
                if 'description' in data:
                    video.description = data['description']
                    changes_list.append('description: updated')
                
                if 'category' in data and data['category'] != video.category:
                    changes_list.append(f'category: {video.category} → {data["category"]}')
                    video.category = data['category']
                
                if 'module' in data and data['module'] != getattr(video, 'module', ''):
                    video.module = data['module']
                    changes_list.append('module: updated')
                
                if 'duration' in data and int(data['duration']) != video.duration:
                    changes_list.append(f'duration: {video.duration} → {data["duration"]}')
                    video.duration = int(data['duration'])
                
                if 'price' in data:
                    old_price = float(video.price)
                    new_price = float(data['price'])
                    if old_price != new_price:
                        changes_list.append(f'price: ${old_price} → ${new_price}')
                    video.price = Decimal(str(data['price']))
                
                if 'order' in data and int(data['order']) != video.order:
                    changes_list.append(f'order: {video.order} → {data["order"]}')
                    video.order = int(data['order'])
                
                if 'is_active' in data and data['is_active'] != video.is_active:
                    old_status = 'active' if video.is_active else 'inactive'
                    new_status = 'active' if data['is_active'] else 'inactive'
                    changes_list.append(f'status: {old_status} → {new_status}')
                    video.is_active = data['is_active']
                
                if 'allow_download' in data:
                    video.allow_download = data['allow_download']
                
                if 'disable_screenshots' in data:
                    video.disable_screenshots = data['disable_screenshots']
                
                # Update course
                if 'course_id' in data:
                    old_course = video.course.title if video.course else 'None'
                    if data['course_id']:
                        try:
                            new_course = Course.objects.get(id=int(data['course_id']))
                            video.course = new_course
                            changes_list.append(f'course: {old_course} → {new_course.title}')
                        except (Course.DoesNotExist, ValueError) as e:
                            print(f"Course not found: {e}")
                    else:
                        video.course = None
                        changes_list.append(f'course: {old_course} → None')
                
                # Save with bypass_validation for S3 updates
                if changes_list:
                    video.save(bypass_validation=True)
                    changes.extend(changes_list)
                else:
                    print("ℹ️ No changes detected")
                
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
        
        # Log activity if there were changes
        admin_user = get_admin_user(request)
        if changes:
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Updated video: {video.title}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'video_id': video.id,
                    'changes': changes
                }
            )
            print(f"✅ Video updated with changes: {changes}")
        else:
            print("ℹ️ No changes to log")
        
        # Get thumbnail URL (prioritize S3 thumbnail if available)
        thumbnail_url = None
        if video.thumbnail_s3_key:
            from django.conf import settings
            thumbnail_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{video.thumbnail_s3_key}"
        elif video.thumbnail and hasattr(video.thumbnail, 'url'):
            try:
                thumbnail_url = video.thumbnail.url
            except:
                thumbnail_url = None
        
        # Get video URL
        video_url = None
        if video.s3_key:
            from django.conf import settings
            video_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{video.s3_key}"
        elif video.video_url:
            video_url = video.video_url
        elif video.video_file and hasattr(video.video_file, 'url'):
            try:
                video_url = video.video_file.url
            except:
                video_url = None
        
        return JsonResponse({
            'success': True,
            'message': 'Video updated successfully',
            'changes': changes,
            'video': {
                'id': video.id,
                'title': video.title,
                'thumbnail_url': thumbnail_url,
                'video_url': video_url,
                's3_key': video.s3_key,
                'thumbnail_s3_key': video.thumbnail_s3_key
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating video: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@require_POST
def admin_api_video_delete(request, video_id):
    """Delete video"""
    video = get_object_or_404(TrainingVideo, id=video_id)
    title = video.title
    
    # Get admin user before deletion
    admin_user = get_admin_user(request)
    
    video.delete()
    
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted video: {title}',
        metadata={
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({'success': True})


@admin_required
@csrf_exempt
def admin_api_video_upload(request):
    """Handle video file upload and thumbnail upload together"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.FILES.get('video_file'):
        return JsonResponse({'error': 'No video file uploaded'}, status=400)
    
    video_file = request.FILES['video_file']
    
    # Validate file type
    allowed_types = ['video/mp4', 'video/webm', 'video/ogg', 'video/quicktime']
    if video_file.content_type not in allowed_types:
        return JsonResponse({
            'error': 'Invalid file type. Please upload MP4, WebM, OGG, or MOV files.'
        }, status=400)
    
    # Validate file size (500MB max)
    if video_file.size > 500 * 1024 * 1024:
        return JsonResponse({
            'error': 'File too large. Maximum size is 500MB.'
        }, status=400)
    
    # Sanitize filename for video
    safe_filename = sanitize_filename(video_file.name)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    video_filename = f"videos/{timestamp}_{safe_filename}"
    
    # Save video file
    video_path = default_storage.save(video_filename, video_file)
    video_url = default_storage.url(video_path)
    
    # Handle thumbnail upload if present
    thumbnail_url = None
    thumbnail_path = None
    if request.FILES.get('thumbnail'):
        thumbnail_file = request.FILES['thumbnail']
        
        # Validate thumbnail file type
        image_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if thumbnail_file.content_type in image_types:
            # Sanitize filename for thumbnail
            thumb_safe_filename = sanitize_filename(thumbnail_file.name)
            thumb_filename = f"video_thumbnails/{timestamp}_{thumb_safe_filename}"
            
            # Save thumbnail file
            thumbnail_path = default_storage.save(thumb_filename, thumbnail_file)
            thumbnail_url = default_storage.url(thumbnail_path)
            print(f"✅ Thumbnail uploaded: {thumbnail_url}")
        else:
            print(f"⚠️ Invalid thumbnail type: {thumbnail_file.content_type}")
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Uploaded video file: {video_file.name}',
        metadata={
            'admin_username': request.session.get('admin_username'),
            'file_name': video_file.name,
            'file_size': video_file.size,
            'file_path': video_path,
            'thumbnail_path': thumbnail_path
        }
    )
    
    return JsonResponse({
        'success': True,
        'path': video_path,
        'url': video_url,
        'filename': video_file.name,
        'size': video_file.size,
        'thumbnail_path': thumbnail_path,
        'thumbnail_url': thumbnail_url
    })

# ==================== ADMIN API - PDF MANAGEMENT (FIXED) ====================

@admin_required
def admin_api_pdfs(request):
    """Get all PDFs with proper S3 URLs"""
    pdfs = PDF.objects.select_related('course').order_by('-created_at')
    
    data = []
    for pdf in pdfs:
        data.append({
            'id': pdf.id,
            'title': pdf.title,
            'description': pdf.description[:100] + '...' if len(pdf.description) > 100 else pdf.description,
            'course_id': pdf.course.id if pdf.course else None,
            'course_name': pdf.course.title if pdf.course else 'General',
            'category': pdf.category,
            'pages': pdf.pages,
            'file_size': pdf.file_size,
            'price': float(pdf.price),
            'is_free': pdf.is_free,
            'access_level': pdf.access_level,
            'views': pdf.views,
            'downloads': pdf.downloads,
            'cover_url': pdf.cover_url,
            'pdf_url': pdf.file_url,
            'is_active': pdf.is_active,
            'is_featured': pdf.is_featured,
            'created_at': pdf.created_at.strftime('%Y-%m-%d'),
            'uploaded': pdf.created_at.strftime('%Y-%m-%d'),
            'size': pdf.file_size,
            'has_file': bool(pdf.pdf_s3_key or pdf.pdf_file),
            'pdf_s3_key': pdf.pdf_s3_key,
            'cover_s3_key': pdf.cover_s3_key,
            'is_s3_pdf': pdf.is_s3_pdf,
        })
    
    return JsonResponse(data, safe=False)


@admin_required
def admin_api_pdf_detail(request, pdf_id):
    """Get single PDF details"""
    pdf = get_object_or_404(PDF, id=pdf_id)
    
    cover_url = None
    if pdf.cover_image:
        try:
            cover_url = pdf.cover_image.url
        except:
            cover_url = None
    
    pdf_url = None
    if pdf.pdf_file:
        try:
            pdf_url = pdf.pdf_file.url
        except:
            pdf_url = None
    
    data = {
        'id': pdf.id,
        'title': pdf.title,
        'description': pdf.description,
        'course_id': pdf.course.id if pdf.course else None,
        'course_name': pdf.course.title if pdf.course else None,
        'category': pdf.category,
        'pages': pdf.pages,
        'file_size': pdf.file_size,
        'price': float(pdf.price),
        'is_free': pdf.is_free,
        'access_level': pdf.access_level,
        'cover_image': cover_url,
        'pdf_url': pdf_url,
        'downloads': pdf.downloads,
        'views': pdf.views,
        'tags': pdf.tags,
        'is_active': pdf.is_active,
        'is_featured': pdf.is_featured,
        'order': pdf.order,
        'created_at': pdf.created_at.strftime('%Y-%m-%d %H:%M'),
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_pdf_create(request):
    """Create a new PDF record with S3 keys from direct upload"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        print(f"📝 Creating PDF with S3 data: {data}")
    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    
    # Validate required fields
    title = data.get('title')
    if not title:
        return JsonResponse({'error': 'Title is required'}, status=400)
    
    pdf_key = data.get('pdf_key') or data.get('pdf_s3_key')
    if not pdf_key:
        return JsonResponse({'error': 'PDF S3 key is required'}, status=400)
    
    try:
        # Create PDF instance
        pdf = PDF(
            title=title,
            description=data.get('description', ''),
            category=data.get('category', 'other'),
            pages=int(data.get('pages', 0)),
            price=Decimal(str(data.get('price', 0))),
            access_level=data.get('access_level', 'free'),
            tags=data.get('tags', ''),
            is_active=data.get('is_active', True),
            is_featured=data.get('is_featured', False),
            order=int(data.get('order', 0)),
        )
        
        # Set is_free based on price
        pdf.is_free = (pdf.price == 0)
        
        # CRITICAL: Store PDF S3 key
        pdf.pdf_s3_key = pdf_key
        if hasattr(pdf, 'pdf_file'):
            pdf.pdf_file.name = pdf_key  # For backward compatibility
        
        # Set file size if provided
        if data.get('size_display'):
            pdf.file_size = data.get('size_display')
        
        # Store cover image S3 key if provided
        cover_key = data.get('cover_key') or data.get('cover_s3_key')
        if cover_key:
            pdf.cover_s3_key = cover_key
            if hasattr(pdf, 'cover_image'):
                pdf.cover_image.name = cover_key  # For backward compatibility
        
        # Set course if provided
        course_id = data.get('course_id')
        if course_id and course_id != 'null' and course_id != '':
            try:
                pdf.course = Course.objects.get(id=int(course_id))
                print(f"📚 Linked to course: {pdf.course.title}")
            except (Course.DoesNotExist, ValueError) as e:
                print(f"⚠️ Course not found: {e}")
        
        # Save with bypass_validation=True for S3 uploads
        pdf.save(bypass_validation=True)
        print(f"✅ PDF created with ID: {pdf.id}")
        print(f"   PDF S3 key: {pdf_key}")
        print(f"   Cover S3 key: {cover_key if cover_key else 'None'}")
        
        # Log activity
        admin_user = get_admin_user(request)
        ActivityLog.objects.create(
            user=admin_user,
            action='ADMIN_ACTION',
            description=f'Created PDF: {pdf.title}',
            metadata={
                'admin_username': request.session.get('admin_username'),
                'pdf_id': pdf.id,
                'pdf_title': pdf.title,
                'pdf_key': pdf_key,
                'cover_key': cover_key
            }
        )
        
        # Generate URLs for response
        from django.conf import settings
        if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') and settings.AWS_S3_CUSTOM_DOMAIN:
            pdf_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{pdf_key}"
            cover_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{cover_key}" if cover_key else None
        else:
            pdf_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{pdf_key}"
            cover_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{cover_key}" if cover_key else None
        
        return JsonResponse({
            'success': True,
            'id': pdf.id,
            'title': pdf.title,
            'message': 'PDF created successfully',
            'pdf_url': pdf_url,
            'cover_url': cover_url,
            'pdf_key': pdf_key,
            'cover_key': cover_key
        })
        
    except Exception as e:
        print(f"❌ Error creating PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@csrf_exempt
def admin_api_pdf_update(request, pdf_id):
    """Update PDF - handles both JSON and multipart/form-data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    pdf = get_object_or_404(PDF, id=pdf_id)
    
    changes = []
    
    try:
        # Handle multipart/form-data (for file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            print(f"🔵 Handling multipart form data update for PDF {pdf_id}")
            
            if 'title' in request.POST:
                new_title = request.POST.get('title')
                if new_title != pdf.title:
                    changes.append(f'title: {pdf.title} → {new_title}')
                    pdf.title = new_title
            
            if 'description' in request.POST:
                pdf.description = request.POST.get('description', pdf.description)
            
            if 'category' in request.POST:
                new_category = request.POST.get('category')
                if new_category != pdf.category:
                    changes.append(f'category: {pdf.category} → {new_category}')
                    pdf.category = new_category
            
            if 'pages' in request.POST:
                new_pages = int(request.POST.get('pages', pdf.pages))
                if new_pages != pdf.pages:
                    changes.append(f'pages: {pdf.pages} → {new_pages}')
                    pdf.pages = new_pages
            
            if 'price' in request.POST:
                old_price = float(pdf.price)
                new_price = float(request.POST.get('price', old_price))
                if old_price != new_price:
                    changes.append(f'price: ${old_price} → ${new_price}')
                pdf.price = Decimal(str(new_price))
                pdf.is_free = (pdf.price == 0)
            
            if 'access_level' in request.POST:
                new_level = request.POST.get('access_level')
                if new_level != pdf.access_level:
                    changes.append(f'access_level: {pdf.access_level} → {new_level}')
                    pdf.access_level = new_level
            
            if 'tags' in request.POST:
                pdf.tags = request.POST.get('tags', pdf.tags)
            
            if 'is_active' in request.POST:
                is_active = request.POST.get('is_active') == 'true'
                old_status = 'active' if pdf.is_active else 'inactive'
                new_status = 'active' if is_active else 'inactive'
                if old_status != new_status:
                    changes.append(f'status: {old_status} → {new_status}')
                pdf.is_active = is_active
            
            if 'is_featured' in request.POST:
                is_featured = request.POST.get('is_featured') == 'true'
                old_featured = 'featured' if pdf.is_featured else 'not featured'
                new_featured = 'featured' if is_featured else 'not featured'
                if old_featured != new_featured:
                    changes.append(f'featured: {old_featured} → {new_featured}')
                pdf.is_featured = is_featured
            
            if 'order' in request.POST:
                new_order = int(request.POST.get('order', pdf.order))
                if new_order != pdf.order:
                    changes.append(f'order: {pdf.order} → {new_order}')
                    pdf.order = new_order
            
            # Handle cover image upload if present
            if 'cover_image' in request.FILES:
                cover = request.FILES['cover_image']
                cover.name = sanitize_filename(cover.name)
                pdf.cover_image = cover
                changes.append('cover_image: updated')
            
            # Update course
            if 'course_id' in request.POST:
                course_id = request.POST.get('course_id')
                old_course = pdf.course.title if pdf.course else 'None'
                if course_id:
                    try:
                        new_course = Course.objects.get(id=course_id)
                        pdf.course = new_course
                        changes.append(f'course: {old_course} → {new_course.title}')
                    except Course.DoesNotExist:
                        pass
                else:
                    pdf.course = None
                    changes.append(f'course: {old_course} → None')
            
            pdf.save()
            
        else:
            # Handle JSON data
            try:
                data = json.loads(request.body)
                print(f"🔵 Handling JSON update for PDF {pdf_id}")
                
                if 'title' in data and data['title'] != pdf.title:
                    changes.append(f'title: {pdf.title} → {data["title"]}')
                    pdf.title = data['title']
                
                if 'description' in data:
                    pdf.description = data['description']
                
                if 'category' in data and data['category'] != pdf.category:
                    changes.append(f'category: {pdf.category} → {data["category"]}')
                    pdf.category = data['category']
                
                if 'pages' in data and int(data['pages']) != pdf.pages:
                    changes.append(f'pages: {pdf.pages} → {data["pages"]}')
                    pdf.pages = int(data['pages'])
                
                if 'price' in data:
                    old_price = float(pdf.price)
                    new_price = float(data['price'])
                    if old_price != new_price:
                        changes.append(f'price: ${old_price} → ${new_price}')
                    pdf.price = Decimal(str(data['price']))
                    pdf.is_free = (pdf.price == 0)
                
                if 'access_level' in data and data['access_level'] != pdf.access_level:
                    changes.append(f'access_level: {pdf.access_level} → {data["access_level"]}')
                    pdf.access_level = data['access_level']
                
                if 'tags' in data:
                    pdf.tags = data['tags']
                
                if 'is_active' in data and data['is_active'] != pdf.is_active:
                    old_status = 'active' if pdf.is_active else 'inactive'
                    new_status = 'active' if data['is_active'] else 'inactive'
                    changes.append(f'status: {old_status} → {new_status}')
                    pdf.is_active = data['is_active']
                
                if 'is_featured' in data and data['is_featured'] != pdf.is_featured:
                    old_featured = 'featured' if pdf.is_featured else 'not featured'
                    new_featured = 'featured' if data['is_featured'] else 'not featured'
                    changes.append(f'featured: {old_featured} → {new_featured}')
                    pdf.is_featured = data['is_featured']
                
                if 'order' in data and int(data['order']) != pdf.order:
                    changes.append(f'order: {pdf.order} → {data["order"]}')
                    pdf.order = int(data['order'])
                
                # Update course
                if 'course_id' in data:
                    old_course = pdf.course.title if pdf.course else 'None'
                    if data['course_id']:
                        try:
                            new_course = Course.objects.get(id=data['course_id'])
                            pdf.course = new_course
                            changes.append(f'course: {old_course} → {new_course.title}')
                        except Course.DoesNotExist:
                            pass
                    else:
                        pdf.course = None
                        changes.append(f'course: {old_course} → None')
                
                pdf.save()
                
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
        
        # Log activity
        admin_user = get_admin_user(request)
        if changes:
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Updated PDF: {pdf.title}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'pdf_id': pdf.id,
                    'changes': changes
                }
            )
            print(f"✅ PDF updated with changes: {changes}")
        
        # Get cover image URL
        cover_url = None
        if pdf.cover_image:
            try:
                cover_url = pdf.cover_image.url
            except:
                cover_url = None
        
        return JsonResponse({
            'success': True,
            'message': 'PDF updated successfully',
            'changes': changes,
            'pdf': {
                'id': pdf.id,
                'title': pdf.title,
                'cover_image': cover_url
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@require_POST
def admin_api_pdf_delete(request, pdf_id):
    """Delete PDF"""
    pdf = get_object_or_404(PDF, id=pdf_id)
    title = pdf.title
    
    # Get admin user before deletion
    admin_user = get_admin_user(request)
    
    pdf.delete()
    
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted PDF: {title}',
        metadata={
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({'success': True})


@admin_required
@csrf_exempt
def admin_api_pdf_upload(request):
    """Handle PDF file upload and cover image upload together"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Get admin user first to ensure it exists before any database operations
    admin_user = get_admin_user(request)
    if not admin_user:
        return JsonResponse({'error': 'Admin user not found'}, status=401)
    
    if not request.FILES.get('pdf_file'):
        return JsonResponse({'error': 'No PDF file uploaded'}, status=400)
    
    pdf_file = request.FILES['pdf_file']
    
    # Validate file type
    if pdf_file.content_type != 'application/pdf':
        return JsonResponse({
            'error': 'Invalid file type. Please upload PDF files only.'
        }, status=400)
    
    # Validate file size (50MB max)
    max_size = 50 * 1024 * 1024  # 50MB
    if pdf_file.size > max_size:
        return JsonResponse({
            'error': f'File too large. Maximum size is 50MB. Your file is {pdf_file.size / (1024*1024):.1f}MB'
        }, status=400)
    
    try:
        # Create media/pdfs directory if it doesn't exist
        pdfs_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
        os.makedirs(pdfs_dir, exist_ok=True)
        
        # Sanitize filename for PDF
        safe_filename = sanitize_filename(pdf_file.name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"pdfs/{timestamp}_{safe_filename}"
        
        # Save PDF file using default_storage
        pdf_path = default_storage.save(pdf_filename, pdf_file)
        pdf_url = default_storage.url(pdf_path)
        
        # Get file size in human readable format
        file_size_bytes = pdf_file.size
        if file_size_bytes < 1024:
            file_size_display = f"{file_size_bytes} B"
        elif file_size_bytes < 1024 * 1024:
            file_size_display = f"{file_size_bytes / 1024:.1f} KB"
        else:
            file_size_display = f"{file_size_bytes / (1024 * 1024):.1f} MB"
        
        # Verify file was saved
        full_path = os.path.join(settings.MEDIA_ROOT, pdf_path)
        if os.path.exists(full_path):
            saved_size = os.path.getsize(full_path)
            print(f"✅ PDF saved successfully: {full_path}")
            print(f"📁 File size: {saved_size} bytes ({file_size_display})")
        else:
            print(f"⚠️ PDF file not found after save: {full_path}")
            # Try to check if storage exists method works
            if default_storage.exists(pdf_path):
                print(f"✅ File exists in storage: {pdf_path}")
            else:
                return JsonResponse({'error': 'File was not saved properly'}, status=500)
        
        # Handle cover image upload if present
        cover_url = None
        cover_path = None
        cover_size_display = None
        
        if request.FILES.get('cover_image'):
            cover_file = request.FILES['cover_image']
            
            # Validate cover image file type
            image_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
            if cover_file.content_type in image_types:
                # Validate cover image size (5MB max)
                if cover_file.size > 5 * 1024 * 1024:
                    return JsonResponse({
                        'warning': 'Cover image too large (max 5MB), but PDF was uploaded',
                        'path': pdf_path,
                        'url': pdf_url,
                        'filename': pdf_file.name,
                        'size': file_size_bytes,
                        'size_display': file_size_display
                    }, status=200)
                
                # Create pdf_covers directory if it doesn't exist
                covers_dir = os.path.join(settings.MEDIA_ROOT, 'pdf_covers')
                os.makedirs(covers_dir, exist_ok=True)
                
                # Sanitize filename for cover
                cover_safe_filename = sanitize_filename(cover_file.name)
                cover_filename = f"pdf_covers/{timestamp}_{cover_safe_filename}"
                
                # Save cover image file
                cover_path = default_storage.save(cover_filename, cover_file)
                cover_url = default_storage.url(cover_path)
                
                # Calculate cover size
                if cover_file.size < 1024:
                    cover_size_display = f"{cover_file.size} B"
                elif cover_file.size < 1024 * 1024:
                    cover_size_display = f"{cover_file.size / 1024:.1f} KB"
                else:
                    cover_size_display = f"{cover_file.size / (1024 * 1024):.1f} MB"
                
                print(f"✅ Cover image uploaded: {cover_url} ({cover_size_display})")
            else:
                print(f"⚠️ Invalid cover image type: {cover_file.content_type}")
                return JsonResponse({
                    'warning': 'Cover image skipped - invalid file type (use JPEG, PNG, GIF, or WebP)',
                    'path': pdf_path,
                    'url': pdf_url,
                    'filename': pdf_file.name,
                    'size': file_size_bytes,
                    'size_display': file_size_display
                }, status=200)
        
        # Log activity
        ActivityLog.objects.create(
            user=admin_user,
            action='ADMIN_ACTION',
            description=f'Uploaded PDF file: {pdf_file.name}',
            metadata={
                'admin_username': request.session.get('admin_username'),
                'file_name': pdf_file.name,
                'file_size': file_size_bytes,
                'file_size_display': file_size_display,
                'file_path': pdf_path,
                'cover_path': cover_path
            }
        )
        
        return JsonResponse({
            'success': True,
            'path': pdf_path,
            'url': pdf_url,
            'filename': pdf_file.name,
            'size': file_size_bytes,
            'size_display': file_size_display,
            'cover_path': cover_path,
            'cover_url': cover_url,
            'cover_size_display': cover_size_display
        })
        
    except Exception as e:
        print(f"❌ Error uploading PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Upload failed: {str(e)}'}, status=500)


@admin_required
@csrf_exempt
def admin_api_pdf_fix_file(request, pdf_id):
    """Create a placeholder file for a PDF that has no file"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.user.is_staff:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        from .models import PDF
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage
        import os
        from datetime import datetime
        
        pdf = PDF.objects.get(id=pdf_id)
        
        # Create media/pdfs directory if it doesn't exist
        pdfs_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
        os.makedirs(pdfs_dir, exist_ok=True)
        
        # Create placeholder content
        placeholder_content = f"""PLACEHOLDER PDF FILE
Title: {pdf.title}
ID: {pdf.id}
Created: {datetime.now()}
This is a placeholder file created because the original file was missing.
Please upload the actual PDF file through the admin interface.
"""
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"pdfs/placeholder_{pdf.id}_{timestamp}.pdf"
        
        # Save placeholder file
        file_path = default_storage.save(filename, ContentFile(placeholder_content.encode('utf-8')))
        
        # Update PDF record
        pdf.pdf_file.name = file_path
        pdf.file_size = f"{len(placeholder_content)} bytes"
        pdf.save()
        
        print(f"✅ Created placeholder file for PDF {pdf_id}: {file_path}")
        
        # Log activity
        admin_user = get_admin_user(request)
        ActivityLog.objects.create(
            user=admin_user,
            action='ADMIN_ACTION',
            description=f'Created placeholder file for PDF: {pdf.title}',
            metadata={
                'admin_username': request.session.get('admin_username'),
                'pdf_id': pdf.id
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Placeholder file created',
            'file_path': file_path,
            'file_url': default_storage.url(file_path)
        })
        
    except PDF.DoesNotExist:
        return JsonResponse({'error': 'PDF not found'}, status=404)
    except Exception as e:
        print(f"❌ Error fixing PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    

@admin_required
def admin_api_pdf_debug(request, pdf_id):
    """Debug endpoint to check PDF file status"""
    try:
        pdf = PDF.objects.get(id=pdf_id)
        
        result = {
            'id': pdf.id,
            'title': pdf.title,
            'has_pdf_file': pdf.pdf_file is not None,
            'pdf_file_name': pdf.pdf_file.name if pdf.pdf_file else None,
            'pdf_file_url': pdf.pdf_file.url if pdf.pdf_file else None,
        }
        
        if pdf.pdf_file:
            try:
                # Check if file exists on disk
                file_path = pdf.pdf_file.path
                result['file_path'] = file_path
                result['file_exists_on_disk'] = os.path.exists(file_path)
                
                if os.path.exists(file_path):
                    result['file_size'] = os.path.getsize(file_path)
                    result['file_modified'] = os.path.getmtime(file_path)
                else:
                    # Try storage exists method
                    try:
                        result['storage_exists'] = pdf.pdf_file.storage.exists(pdf.pdf_file.name)
                    except:
                        result['storage_exists'] = False
                        
            except Exception as e:
                result['file_error'] = str(e)
        
        return JsonResponse(result)
        
    except PDF.DoesNotExist:
        return JsonResponse({'error': 'PDF not found'}, status=404)    

# ==================== ADMIN API - PACKAGE MANAGEMENT ====================

@admin_required
def admin_api_packages(request):
    """API endpoint for packages in admin"""
    if request.method == 'GET':
        packages = Package.objects.all().order_by('-id')
        
        data = []
        for pkg in packages:
            # Get image URL properly
            image_url = None
            if pkg.image:
                try:
                    # This will return the full URL including /media/
                    image_url = pkg.image.url
                    print(f"Package {pkg.id} image URL: {image_url}")  # DEBUG
                except Exception as e:
                    print(f"Error getting image for package {pkg.id}: {e}")
                    image_url = None
            
            # Get features as list
            features = []
            if pkg.features:
                if isinstance(pkg.features, list):
                    features = pkg.features
                elif isinstance(pkg.features, str):
                    # Try to parse JSON first
                    try:
                        import json
                        features = json.loads(pkg.features)
                    except:
                        # Split by new lines
                        features = [f.strip() for f in pkg.features.split('\n') if f.strip()]
            
            data.append({
                'id': pkg.id,
                'name': pkg.name,
                'description': pkg.full_description or pkg.short_description,
                'price': float(pkg.price) if pkg.price else 0,
                'package_type': pkg.package_type,
                'features': features,
                'image': image_url,  # ← THIS MUST BE THE FULL URL
                'status': 'active' if pkg.is_active else 'inactive',
                'sales': pkg.total_sales,
                'revenue': float(pkg.total_revenue) if pkg.total_revenue else 0,
                'is_recurring': pkg.is_recurring,
                'duration': f"{pkg.duration_days} days",
                'created_at': pkg.created_at.strftime('%Y-%m-%d %H:%M') if pkg.created_at else None
            })
        
        return JsonResponse(data, safe=False)
    

@admin_required
@csrf_exempt
def admin_api_package_detail(request, package_id):
    """Get single package details"""
    package = get_object_or_404(Package, id=package_id)
    
    # Get image URL properly
    image_url = None
    if package.image:
        try:
            image_url = package.image.url
        except Exception as e:
            print(f"Error getting image for package {package.id}: {e}")
            image_url = None
    
    # Get features as list
    features = []
    if package.features:
        if isinstance(package.features, list):
            features = package.features
        elif isinstance(package.features, str):
            try:
                import json
                features = json.loads(package.features)
            except:
                features = [f.strip() for f in package.features.split('\n') if f.strip()]
    
    # Get benefits as list
    benefits = []
    if package.benefits:
        if isinstance(package.benefits, list):
            benefits = package.benefits
        elif isinstance(package.benefits, str):
            try:
                import json
                benefits = json.loads(package.benefits)
            except:
                benefits = [f.strip() for f in package.benefits.split('\n') if f.strip()]
    
    data = {
        'id': package.id,
        'name': package.name,
        'short_description': package.short_description,
        'full_description': package.full_description,
        'price': float(package.price) if package.price else 0,
        'original_price': float(package.original_price) if package.original_price else None,
        'discount_percentage': package.discount_percentage,
        'package_type': package.package_type,
        'duration_days': package.duration_days,
        'is_recurring': package.is_recurring,
        'is_featured': package.is_featured,
        'is_popular': package.is_popular,
        'is_active': package.is_active,
        'order': package.order,
        'features': features,
        'benefits': benefits,
        'image': image_url,
        'total_sales': package.total_sales,
        'total_revenue': float(package.total_revenue) if package.total_revenue else 0,
        'created_at': package.created_at.strftime('%Y-%m-%d %H:%M') if package.created_at else None,
        'updated_at': package.updated_at.strftime('%Y-%m-%d %H:%M') if package.updated_at else None,
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_package_create(request):
    """Create a new package"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    package = Package(
        name=data.get('name'),
        package_type=data.get('package_type'),
        short_description=data.get('short_description', ''),
        full_description=data.get('full_description', ''),
        price=Decimal(str(data.get('price', 0))),
        duration_days=int(data.get('duration_days', 30)),
        is_recurring=data.get('is_recurring', False),
        is_featured=data.get('is_featured', False),
        is_popular=data.get('is_popular', False),
        is_active=data.get('is_active', True),
        order=data.get('order', 0),
        features=data.get('features', []),
        benefits=data.get('benefits', []),
    )
    
    if data.get('original_price'):
        package.original_price = Decimal(str(data['original_price']))
        if package.original_price > 0:
            package.discount_percentage = int(((package.original_price - package.price) / package.original_price) * 100)
    
    package.save()
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Created package: {package.name}',
        metadata={
            'admin_username': request.session.get('admin_username'),
            'package_id': package.id
        }
    )
    
    return JsonResponse({'success': True, 'id': package.id})


@admin_required
@csrf_exempt
def admin_api_package_update(request, package_id):
    """Update package - handles both JSON and multipart/form-data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    package = get_object_or_404(Package, id=package_id)
    
    changes = []
    
    try:
        # Handle multipart/form-data (for file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            print(f"🔵 Handling multipart form data update for package {package_id}")
            
            if 'name' in request.POST:
                new_name = request.POST.get('name')
                if new_name != package.name:
                    changes.append(f'name: {package.name} → {new_name}')
                    package.name = new_name
            
            if 'short_description' in request.POST:
                package.short_description = request.POST.get('short_description', package.short_description)
            
            if 'full_description' in request.POST:
                package.full_description = request.POST.get('full_description', package.full_description)
            
            if 'price' in request.POST:
                old_price = float(package.price)
                new_price = float(request.POST.get('price', old_price))
                if old_price != new_price:
                    changes.append(f'price: ${old_price} → ${new_price}')
                    # Update discount percentage if original_price exists
                    if package.original_price and package.original_price > 0:
                        package.discount_percentage = int(((package.original_price - Decimal(str(new_price))) / package.original_price) * 100)
                package.price = Decimal(str(new_price))
            
            if 'duration_days' in request.POST:
                new_duration = int(request.POST.get('duration_days', package.duration_days))
                if new_duration != package.duration_days:
                    changes.append(f'duration: {package.duration_days} → {new_duration}')
                    package.duration_days = new_duration
            
            if 'is_recurring' in request.POST:
                is_recurring = request.POST.get('is_recurring') == 'true'
                old_recurring = 'recurring' if package.is_recurring else 'one-time'
                new_recurring = 'recurring' if is_recurring else 'one-time'
                if old_recurring != new_recurring:
                    changes.append(f'billing: {old_recurring} → {new_recurring}')
                package.is_recurring = is_recurring
            
            if 'is_featured' in request.POST:
                is_featured = request.POST.get('is_featured') == 'true'
                old_featured = 'featured' if package.is_featured else 'not featured'
                new_featured = 'featured' if is_featured else 'not featured'
                if old_featured != new_featured:
                    changes.append(f'featured: {old_featured} → {new_featured}')
                package.is_featured = is_featured
            
            if 'is_popular' in request.POST:
                is_popular = request.POST.get('is_popular') == 'true'
                old_popular = 'popular' if package.is_popular else 'not popular'
                new_popular = 'popular' if is_popular else 'not popular'
                if old_popular != new_popular:
                    changes.append(f'popular: {old_popular} → {new_popular}')
                package.is_popular = is_popular
            
            if 'is_active' in request.POST:
                is_active = request.POST.get('is_active') == 'true'
                old_status = 'active' if package.is_active else 'inactive'
                new_status = 'active' if is_active else 'inactive'
                if old_status != new_status:
                    changes.append(f'status: {old_status} → {new_status}')
                package.is_active = is_active
            
            if 'order' in request.POST:
                new_order = int(request.POST.get('order', package.order))
                if new_order != package.order:
                    changes.append(f'order: {package.order} → {new_order}')
                    package.order = new_order
            
            if 'features' in request.POST:
                try:
                    new_features = json.loads(request.POST.get('features'))
                except:
                    new_features = request.POST.get('features', '').split('\n')
                package.features = new_features
                changes.append('features: updated')
            
            if 'benefits' in request.POST:
                try:
                    new_benefits = json.loads(request.POST.get('benefits'))
                except:
                    new_benefits = request.POST.get('benefits', '').split('\n')
                package.benefits = new_benefits
                changes.append('benefits: updated')
            
            # Handle image upload if present
            if 'image' in request.FILES:
                image = request.FILES['image']
                image.name = sanitize_filename(image.name)
                package.image = image
                changes.append('image: updated')
            
            package.save()
            
        else:
            # Handle JSON data
            try:
                data = json.loads(request.body)
                print(f"🔵 Handling JSON update for package {package_id}")
                
                if 'name' in data and data['name'] != package.name:
                    changes.append(f'name: {package.name} → {data["name"]}')
                    package.name = data['name']
                
                if 'short_description' in data:
                    package.short_description = data['short_description']
                
                if 'full_description' in data:
                    package.full_description = data['full_description']
                
                if 'price' in data:
                    old_price = float(package.price)
                    new_price = float(data['price'])
                    if old_price != new_price:
                        changes.append(f'price: ${old_price} → ${new_price}')
                        # Update discount percentage if original_price exists
                        if package.original_price and package.original_price > 0:
                            package.discount_percentage = int(((package.original_price - Decimal(str(new_price))) / package.original_price) * 100)
                    package.price = Decimal(str(new_price))
                
                if 'original_price' in data:
                    package.original_price = Decimal(str(data['original_price']))
                    if package.original_price > 0 and package.price > 0:
                        package.discount_percentage = int(((package.original_price - package.price) / package.original_price) * 100)
                
                if 'duration_days' in data and int(data['duration_days']) != package.duration_days:
                    changes.append(f'duration: {package.duration_days} → {data["duration_days"]}')
                    package.duration_days = int(data['duration_days'])
                
                if 'is_recurring' in data and data['is_recurring'] != package.is_recurring:
                    old_recurring = 'recurring' if package.is_recurring else 'one-time'
                    new_recurring = 'recurring' if data['is_recurring'] else 'one-time'
                    changes.append(f'billing: {old_recurring} → {new_recurring}')
                    package.is_recurring = data['is_recurring']
                
                if 'is_featured' in data and data['is_featured'] != package.is_featured:
                    old_featured = 'featured' if package.is_featured else 'not featured'
                    new_featured = 'featured' if data['is_featured'] else 'not featured'
                    changes.append(f'featured: {old_featured} → {new_featured}')
                    package.is_featured = data['is_featured']
                
                if 'is_popular' in data and data['is_popular'] != package.is_popular:
                    old_popular = 'popular' if package.is_popular else 'not popular'
                    new_popular = 'popular' if data['is_popular'] else 'not popular'
                    changes.append(f'popular: {old_popular} → {new_popular}')
                    package.is_popular = data['is_popular']
                
                if 'is_active' in data and data['is_active'] != package.is_active:
                    old_status = 'active' if package.is_active else 'inactive'
                    new_status = 'active' if data['is_active'] else 'inactive'
                    changes.append(f'status: {old_status} → {new_status}')
                    package.is_active = data['is_active']
                
                if 'order' in data and int(data['order']) != package.order:
                    changes.append(f'order: {package.order} → {data["order"]}')
                    package.order = int(data['order'])
                
                if 'features' in data:
                    package.features = data['features']
                    changes.append('features: updated')
                
                if 'benefits' in data:
                    package.benefits = data['benefits']
                    changes.append('benefits: updated')
                
                package.save()
                
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
        
        # Log activity
        admin_user = get_admin_user(request)
        if changes:
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Updated package: {package.name}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'package_id': package.id,
                    'changes': changes
                }
            )
            print(f"✅ Package updated with changes: {changes}")
        
        # Get image URL
        image_url = None
        if package.image:
            try:
                image_url = package.image.url
            except:
                image_url = None
        
        return JsonResponse({
            'success': True,
            'message': 'Package updated successfully',
            'changes': changes,
            'package': {
                'id': package.id,
                'name': package.name,
                'price': float(package.price),
                'image': image_url
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating package: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@require_POST
def admin_api_package_delete(request, package_id):
    """Delete package"""
    package = get_object_or_404(Package, id=package_id)
    name = package.name
    
    # Check if package has orders
    orders_count = PaymentTransaction.objects.filter(package_name=name).count()
    if orders_count > 0:
        return JsonResponse({
            'error': f'Cannot delete package with {orders_count} orders'
        }, status=400)
    
    # Get admin user
    admin_user = get_admin_user(request)
    
    package.delete()
    
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted package: {name}',
        metadata={
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({'success': True})


@admin_required
@csrf_exempt
def admin_api_package_toggle_popular(request, package_id):
    """Toggle popular status of a package"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    package = get_object_or_404(Package, id=package_id)
    
    # Toggle popular status
    package.is_popular = not package.is_popular
    package.save()
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f"{'Set' if package.is_popular else 'Removed'} popular status for package: {package.name}",
        metadata={
            'admin_username': request.session.get('admin_username'),
            'package_id': package.id,
            'is_popular': package.is_popular
        }
    )
    
    return JsonResponse({
        'success': True,
        'is_popular': package.is_popular
    })


# ==================== ADMIN API - ORDER MANAGEMENT ====================

@admin_required
def admin_api_orders(request):
    """Get all orders/transactions"""
    orders = PaymentTransaction.objects.select_related('user', 'course').order_by('-created_at')[:100]
    
    data = []
    for order in orders:
        # Determine item name
        item_name = 'Unknown'
        if order.course:
            item_name = order.course.title
        elif order.package_name:
            item_name = order.package_name
        elif order.program_name:
            item_name = order.program_name
        elif order.partnership_name:
            item_name = order.partnership_name
        
        data.append({
            'id': f"#{order.reference[-8:]}",
            'reference': order.reference,
            'customer': order.user.get_full_name() or order.user.username if order.user else 'Unknown',
            'package': order.get_payment_type_display(),
            'item': item_name,
            'amount': f"${float(order.amount):,.2f}",
            'date': order.created_at.strftime('%Y-%m-%d'),
            'payment': order.payment_method.title(),
            'status': order.status,
        })
    
    return JsonResponse(data, safe=False)


@admin_required
def admin_api_order_detail(request, order_id):
    """Get single order details"""
    # Try to find by ID or reference
    try:
        if order_id.isdigit():
            order = get_object_or_404(PaymentTransaction, id=int(order_id))
        else:
            order = get_object_or_404(PaymentTransaction, reference=order_id)
    except:
        return JsonResponse({'error': 'Order not found'}, status=404)
    
    # Determine item name
    item_name = 'Unknown'
    if order.course:
        item_name = order.course.title
    elif order.package_name:
        item_name = order.package_name
    elif order.program_name:
        item_name = order.program_name
    elif order.partnership_name:
        item_name = order.partnership_name
    
    data = {
        'id': order.id,
        'reference': order.reference,
        'user_id': order.user.id if order.user else None,
        'user_name': order.user.get_full_name() or order.user.username if order.user else 'Guest',
        'user_email': order.user.email if order.user else order.email or 'Unknown',
        'user_phone': order.user.phone if order.user else order.phone or 'Unknown',
        'amount': float(order.amount),
        'amount_formatted': f"${float(order.amount):,.2f}",
        'payment_type': order.get_payment_type_display(),
        'payment_method': order.payment_method,
        'status': order.status,
        'item_name': item_name,
        'course_id': order.course.id if order.course else None,
        'created_at': order.created_at.strftime('%Y-%m-%d %H:%M'),
        'paid_at': order.paid_at.strftime('%Y-%m-%d %H:%M') if order.paid_at else None,
        'metadata': order.metadata,
    }
    
    return JsonResponse(data)


# ==================== ADMIN API - PARTNERSHIP MANAGEMENT ====================

@admin_required
def admin_api_partnerships(request):
    """Get all partnership applications"""
    partnerships = UserPartnership.objects.select_related('user', 'program').order_by('-created_at')
    
    data = []
    for p in partnerships:
        data.append({
            'id': p.id,
            'company': p.user.get_full_name() or p.user.username,
            'contact': p.user.email,
            'tier': p.program.get_tier_display() if p.program else 'N/A',
            'amount': f"${float(p.investment_amount):,.2f}",
            'nda': 'Signed' if p.contract_signed else 'Pending',
            'status': p.status,
        })
    
    return JsonResponse(data, safe=False)


@admin_required
@csrf_exempt
def admin_api_update_partnership_status(request, partnership_id):
    """Update partnership status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    partnership = get_object_or_404(UserPartnership, id=partnership_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    old_status = partnership.status
    partnership.status = data.get('status', partnership.status)
    partnership.manager_notes = data.get('notes', partnership.manager_notes)
    
    if data.get('contract_signed'):
        partnership.contract_signed = True
        partnership.contract_signed_at = timezone.now()
    
    partnership.save()
    
    # Notify user
    Notification.objects.create(
        user=partnership.user,
        title='Partnership Status Updated',
        message=f'Your {partnership.program.name} partnership status is now: {partnership.status}',
        notification_type='INFO',
    )
    
    # Log activity
    admin_user = get_admin_user(request)
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Updated partnership status for {partnership.user.username}: {old_status} → {partnership.status}',
        metadata={
            'admin_username': request.session.get('admin_username'),
            'partnership_id': partnership.id
        }
    )
    
    return JsonResponse({'success': True})

# ==================== ADMIN API - SUPPORT TICKETS ====================

@admin_required
def admin_api_support_tickets(request):
    """Get all support tickets"""
    tickets = SupportTicket.objects.select_related('user', 'assigned_to').order_by('-created_at')
    
    data = []
    for ticket in tickets:
        data.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'subject': ticket.subject[:50] + '...' if len(ticket.subject) > 50 else ticket.subject,
            'user': ticket.user.username,
            'user_email': ticket.user.email,
            'category': ticket.get_category_display(),
            'priority': ticket.priority,
            'status': ticket.status,
            'reply_count': ticket.reply_count,
            'created': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            'assigned_to': ticket.assigned_to.username if ticket.assigned_to else None,
        })
    
    return JsonResponse(data, safe=False)


@admin_required
def admin_api_ticket_detail(request, ticket_id):
    """Get ticket details with replies"""
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    replies = ticket.replies.select_related('user').order_by('created_at')
    
    reply_data = []
    for reply in replies:
        reply_data.append({
            'id': reply.id,
            'user': reply.user.username,
            'user_email': reply.user.email,
            'message': reply.message,
            'is_internal': reply.is_internal,
            'created_at': reply.created_at.strftime('%Y-%m-%d %H:%M'),
            'time_ago': timesince(reply.created_at),
        })
    
    data = {
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
        'subject': ticket.subject,
        'message': ticket.message,
        'user': ticket.user.username,
        'user_email': ticket.user.email,
        'category': ticket.get_category_display(),
        'priority': ticket.priority,
        'status': ticket.status,
        'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
        'assigned_to': ticket.assigned_to.username if ticket.assigned_to else None,
        'replies': reply_data,
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_ticket_reply(request, ticket_id):
    """Reply to ticket"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    # Get admin user
    admin_user = get_admin_user(request)
    
    reply = TicketReply.objects.create(
        ticket=ticket,
        user=admin_user,
        message=data.get('message'),
        is_internal=data.get('is_internal', False),
    )
    
    # Update ticket
    ticket.reply_count += 1
    ticket.last_reply_at = timezone.now()
    
    if not data.get('is_internal'):
        ticket.status = 'waiting_reply'
    
    ticket.save()
    
    # Notify user if not internal
    if not data.get('is_internal'):
        Notification.objects.create(
            user=ticket.user,
            title=f'Ticket #{ticket.ticket_number} Updated',
            message=f'Your ticket has received a reply: {data.get("message")[:100]}...',
            notification_type='INFO',
            related_object_type='ticket',
            related_object_id=ticket.id,
        )
    
    return JsonResponse({
        'success': True,
        'id': reply.id,
        'created_at': reply.created_at.strftime('%Y-%m-%d %H:%M'),
    })


@admin_required
@csrf_exempt
def admin_api_ticket_update_status(request, ticket_id):
    """Update ticket status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    old_status = ticket.status
    ticket.status = data.get('status', ticket.status)
    
    if ticket.status == 'resolved':
        ticket.resolved_at = timezone.now()
        admin_user = get_admin_user(request)
        ticket.resolved_by = admin_user
    elif ticket.status == 'closed':
        ticket.closed_at = timezone.now()
    
    ticket.save()
    
    # Notify user
    Notification.objects.create(
        user=ticket.user,
        title=f'Ticket #{ticket.ticket_number} Status Changed',
        message=f'Your ticket status changed from {old_status} to {ticket.status}',
        notification_type='INFO',
    )
    
    return JsonResponse({'success': True})

# ==================== ADMIN API - BLOG MANAGEMENT ====================

@admin_required
def admin_api_blogs(request):
    """Get all blogs with featured images"""
    blogs = Blog.objects.all().select_related('author').order_by('-created_at')
    
    data = []
    for blog in blogs:
        featured_image_url = None
        if blog.featured_image:
            try:
                featured_image_url = blog.featured_image.url
            except:
                featured_image_url = None
        
        data.append({
            'id': blog.id,
            'title': blog.title,
            'slug': blog.slug,
            'author': blog.author.username if blog.author else 'Admin',
            'category': blog.get_category_display(),
            'views': blog.views,
            'published': blog.published_at.strftime('%Y-%m-%d') if blog.published_at else 'Not published',
            'status': blog.status,
            'is_featured': blog.is_featured,
            'featured_image': featured_image_url,
        })
    
    return JsonResponse(data, safe=False)


@admin_required
@csrf_exempt
def admin_api_blog_detail(request, blog_id):
    """Get single blog details"""
    blog = get_object_or_404(Blog, id=blog_id)
    
    featured_image_url = None
    if blog.featured_image:
        try:
            featured_image_url = blog.featured_image.url
        except:
            featured_image_url = None
    
    data = {
        'id': blog.id,
        'title': blog.title,
        'slug': blog.slug,
        'content': blog.content,
        'excerpt': blog.excerpt,
        'author_id': blog.author.id if blog.author else None,
        'author_name': blog.author.username if blog.author else 'Admin',
        'category': blog.category,
        'category_display': blog.get_category_display(),
        'tags': blog.tags,
        'views': blog.views,
        'status': blog.status,
        'is_featured': blog.is_featured,
        'featured_image': featured_image_url,
        'meta_title': blog.meta_title,
        'meta_description': blog.meta_description,
        'meta_keywords': blog.meta_keywords,
        'published_at': blog.published_at.strftime('%Y-%m-%d %H:%M') if blog.published_at else None,
        'created_at': blog.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': blog.updated_at.strftime('%Y-%m-%d %H:%M') if blog.updated_at else None,
    }
    
    return JsonResponse(data)


@admin_required
@csrf_exempt
def admin_api_blog_create(request):
    """Create a new blog post"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Handle FormData (if file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            title = request.POST.get('title')
            content = request.POST.get('content')
            excerpt = request.POST.get('excerpt', '')
            category = request.POST.get('category', 'general')
            tags = request.POST.get('tags', '')
            status = request.POST.get('status', 'draft')
            is_featured = request.POST.get('is_featured') == 'true'
            
            # Get admin user as author
            admin_user = get_admin_user(request)
            
            blog = Blog(
                title=title,
                content=content,
                excerpt=excerpt,
                category=category,
                tags=tags,
                status=status,
                is_featured=is_featured,
                author=admin_user,
            )
            
            if status == 'published':
                blog.published_at = timezone.now()
            
            # Handle featured image if present
            if 'featured_image' in request.FILES:
                featured_image = request.FILES['featured_image']
                featured_image.name = sanitize_filename(featured_image.name)
                blog.featured_image = featured_image
            
            blog.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created blog: {blog.title}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'blog_id': blog.id
                }
            )
            
            return JsonResponse({'success': True, 'id': blog.id})
        
        # Handle JSON data
        else:
            data = json.loads(request.body)
            
            admin_user = get_admin_user(request)
            
            blog = Blog(
                title=data.get('title'),
                content=data.get('content'),
                excerpt=data.get('excerpt', ''),
                category=data.get('category', 'general'),
                tags=data.get('tags', ''),
                status=data.get('status', 'draft'),
                is_featured=data.get('is_featured', False),
                meta_title=data.get('meta_title', ''),
                meta_description=data.get('meta_description', ''),
                meta_keywords=data.get('meta_keywords', ''),
                author=admin_user,
            )
            
            if blog.status == 'published':
                blog.published_at = timezone.now()
            
            blog.save()
            
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Created blog: {blog.title}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'blog_id': blog.id
                }
            )
            
            return JsonResponse({'success': True, 'id': blog.id})
            
    except Exception as e:
        print(f"Error creating blog: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@csrf_exempt
def admin_api_blog_update(request, blog_id):
    """Update blog post - handles both JSON and multipart/form-data"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    blog = get_object_or_404(Blog, id=blog_id)
    
    changes = []
    
    try:
        # Handle multipart/form-data (for file uploads)
        if request.content_type and 'multipart/form-data' in request.content_type:
            print(f"🔵 Handling multipart form data update for blog {blog_id}")
            
            if 'title' in request.POST:
                new_title = request.POST.get('title')
                if new_title != blog.title:
                    changes.append(f'title: {blog.title} → {new_title}')
                    blog.title = new_title
            
            if 'content' in request.POST:
                blog.content = request.POST.get('content', blog.content)
            
            if 'excerpt' in request.POST:
                blog.excerpt = request.POST.get('excerpt', blog.excerpt)
            
            if 'category' in request.POST:
                new_category = request.POST.get('category')
                if new_category != blog.category:
                    changes.append(f'category: {blog.category} → {new_category}')
                    blog.category = new_category
            
            if 'tags' in request.POST:
                blog.tags = request.POST.get('tags', blog.tags)
            
            if 'status' in request.POST:
                new_status = request.POST.get('status')
                if new_status != blog.status:
                    changes.append(f'status: {blog.status} → {new_status}')
                    blog.status = new_status
                    if blog.status == 'published' and not blog.published_at:
                        blog.published_at = timezone.now()
            
            if 'is_featured' in request.POST:
                is_featured = request.POST.get('is_featured') == 'true'
                old_featured = 'featured' if blog.is_featured else 'not featured'
                new_featured = 'featured' if is_featured else 'not featured'
                if old_featured != new_featured:
                    changes.append(f'featured: {old_featured} → {new_featured}')
                blog.is_featured = is_featured
            
            # Handle featured image upload if present
            if 'featured_image' in request.FILES:
                featured_image = request.FILES['featured_image']
                featured_image.name = sanitize_filename(featured_image.name)
                blog.featured_image = featured_image
                changes.append('featured_image: updated')
            
            blog.save()
            
        else:
            # Handle JSON data
            try:
                data = json.loads(request.body)
                print(f"🔵 Handling JSON update for blog {blog_id}")
                
                if 'title' in data and data['title'] != blog.title:
                    changes.append(f'title: {blog.title} → {data["title"]}')
                    blog.title = data['title']
                
                if 'content' in data:
                    blog.content = data['content']
                
                if 'excerpt' in data:
                    blog.excerpt = data['excerpt']
                
                if 'category' in data and data['category'] != blog.category:
                    changes.append(f'category: {blog.category} → {data["category"]}')
                    blog.category = data['category']
                
                if 'tags' in data:
                    blog.tags = data['tags']
                
                if 'status' in data and data['status'] != blog.status:
                    changes.append(f'status: {blog.status} → {data["status"]}')
                    blog.status = data['status']
                    if blog.status == 'published' and not blog.published_at:
                        blog.published_at = timezone.now()
                
                if 'is_featured' in data and data['is_featured'] != blog.is_featured:
                    old_featured = 'featured' if blog.is_featured else 'not featured'
                    new_featured = 'featured' if data['is_featured'] else 'not featured'
                    changes.append(f'featured: {old_featured} → {new_featured}')
                    blog.is_featured = data['is_featured']
                
                if 'meta_title' in data:
                    blog.meta_title = data['meta_title']
                
                if 'meta_description' in data:
                    blog.meta_description = data['meta_description']
                
                if 'meta_keywords' in data:
                    blog.meta_keywords = data['meta_keywords']
                
                blog.save()
                
            except json.JSONDecodeError as e:
                return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
        
        # Log activity
        admin_user = get_admin_user(request)
        if changes:
            ActivityLog.objects.create(
                user=admin_user,
                action='ADMIN_ACTION',
                description=f'Updated blog: {blog.title}',
                metadata={
                    'admin_username': request.session.get('admin_username'),
                    'blog_id': blog.id,
                    'changes': changes
                }
            )
            print(f"✅ Blog updated with changes: {changes}")
        
        # Get featured image URL
        featured_image_url = None
        if blog.featured_image:
            try:
                featured_image_url = blog.featured_image.url
            except:
                featured_image_url = None
        
        return JsonResponse({
            'success': True,
            'message': 'Blog updated successfully',
            'changes': changes,
            'blog': {
                'id': blog.id,
                'title': blog.title,
                'featured_image': featured_image_url
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating blog: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@admin_required
@require_POST
def admin_api_blog_delete(request, blog_id):
    """Delete blog post"""
    blog = get_object_or_404(Blog, id=blog_id)
    title = blog.title
    
    admin_user = get_admin_user(request)
    
    blog.delete()
    
    ActivityLog.objects.create(
        user=admin_user,
        action='ADMIN_ACTION',
        description=f'Deleted blog: {title}',
        metadata={
            'admin_username': request.session.get('admin_username')
        }
    )
    
    return JsonResponse({'success': True})


# ==================== ADMIN API - REPORTS ====================

@admin_required
def admin_api_revenue_report(request):
    """Get revenue report data"""
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    
    # Daily revenue for last 30 days
    daily = []
    for i in range(30):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        revenue = PaymentTransaction.objects.filter(
            status='completed',
            paid_at__range=[day_start, day_end]
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        daily.append({
            'date': day.strftime('%Y-%m-%d'),
            'revenue': float(revenue)
        })
    
    # Revenue by package type
    by_package = []
    package_revenues = PaymentTransaction.objects.filter(
        status='completed'
    ).values('payment_type').annotate(total=Sum('amount'))
    
    payment_types = dict(PaymentTransaction.PAYMENT_TYPES)
    for item in package_revenues:
        by_package.append({
            'name': payment_types.get(item['payment_type'], item['payment_type']),
            'total': float(item['total'])
        })
    
    # Revenue by course
    by_course = []
    course_revenues = PaymentTransaction.objects.filter(
        status='completed',
        course__isnull=False
    ).values('course__title').annotate(total=Sum('amount'))
    
    for item in course_revenues:
        by_course.append({
            'name': item['course__title'] or 'Unknown Course',
            'total': float(item['total'])
        })
    
    # Totals
    total_30d = PaymentTransaction.objects.filter(
        status='completed',
        paid_at__gte=thirty_days_ago
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    total_all = PaymentTransaction.objects.filter(
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    return JsonResponse({
        'daily': daily,
        'by_package': by_package,
        'by_course': by_course,
        'total_30d': float(total_30d),
        'total_all': float(total_all)
    })


@admin_required
def admin_api_users_report(request):
    """Get users report data"""
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    
    # Daily new users for last 30 days
    daily = []
    for i in range(30):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        count = MfalmeUsers.objects.filter(
            date_joined__range=[day_start, day_end]
        ).count()
        
        daily.append({
            'date': day.strftime('%Y-%m-%d'),
            'new_users': count
        })
    
    # Users by rank
    by_rank = []
    rank_counts = MfalmeUsers.objects.values('elite_rank').annotate(count=Count('id'))
    for item in rank_counts:
        by_rank.append({
            'tier': item['elite_rank'],
            'count': item['count']
        })
    
    # Users by status
    by_status = []
    status_counts = MfalmeUsers.objects.values('account_status').annotate(count=Count('id'))
    for item in status_counts:
        by_status.append({
            'status': item['account_status'],
            'count': item['count']
        })
    
    # Course enrollment stats
    course_enrollments = []
    enrollments = UserCourse.objects.values('course__title').annotate(
        count=Count('id'),
        revenue=Sum('payment__amount')
    ).order_by('-count')[:10]
    
    for item in enrollments:
        course_enrollments.append({
            'course': item['course__title'] or 'Unknown',
            'students': item['count'],
            'revenue': float(item['revenue'] or 0)
        })
    
    # Totals
    total = MfalmeUsers.objects.count()
    active = MfalmeUsers.objects.filter(account_status='active').count()
    pending = MfalmeUsers.objects.filter(account_status='pending').count()
    verified = MfalmeUsers.objects.filter(email_verified=True).count()
    
    return JsonResponse({
        'daily': daily,
        'by_rank': by_rank,
        'by_status': by_status,
        'course_enrollments': course_enrollments,
        'total': total,
        'active': active,
        'pending': pending,
        'verified': verified,
    })

# ==================== ADMIN API - DELETE ITEM ====================

@admin_required
@require_POST
def admin_api_delete_item(request, item_type, item_id):
    """Generic delete function"""
    models_map = {
        'user': MfalmeUsers,
        'video': TrainingVideo,
        'pdf': PDF,
        'course': Course,
        'package': Package,
        'blog': Blog,
        'ticket': SupportTicket,
        'partnership': UserPartnership,
    }
    
    model_class = models_map.get(item_type)
    if not model_class:
        return JsonResponse({'error': 'Invalid item type'}, status=400)
    
    try:
        item = model_class.objects.get(id=item_id)
        
        # Get admin user
        admin_user = get_admin_user(request)
        
        # Don't allow deleting yourself
        if item_type == 'user' and admin_user and item.id == admin_user.id:
            return JsonResponse({'error': 'Cannot delete yourself'}, status=400)
        
        # Get identifier for logging
        identifier = str(item)
        if hasattr(item, 'title'):
            identifier = item.title
        elif hasattr(item, 'name'):
            identifier = item.name
        elif hasattr(item, 'username'):
            identifier = item.username
        elif hasattr(item, 'email'):
            identifier = item.email
        
        item.delete()
        
        ActivityLog.objects.create(
            user=admin_user,
            action='ADMIN_ACTION',
            description=f'Deleted {item_type}: {identifier}',
            metadata={
                'admin_username': request.session.get('admin_username'),
                'item_type': item_type,
                'item_id': item_id
            }
        )
        
        return JsonResponse({'success': True})
        
    except model_class.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)

# ==================== ADMIN API - EXPORT FUNCTIONS ====================

@admin_required
def admin_api_users_export(request):
    """Export users to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get users
    users = MfalmeUsers.objects.all().order_by('-date_joined')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email', 'Phone', 'Full Name', 'Rank', 'Status', 'Date Joined'])
        
        for user in users:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.phone,
                user.get_full_name(),
                user.elite_rank,
                user.account_status,
                user.date_joined.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users Export"
        
        # Headers
        headers = ['ID', 'Username', 'Email', 'Phone', 'Full Name', 'Rank', 'Status', 'Date Joined']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.username)
            ws.cell(row=row_num, column=3, value=user.email)
            ws.cell(row=row_num, column=4, value=user.phone)
            ws.cell(row=row_num, column=5, value=user.get_full_name())
            ws.cell(row=row_num, column=6, value=user.elite_rank)
            ws.cell(row=row_num, column=7, value=user.account_status)
            ws.cell(row=row_num, column=8, value=user.date_joined.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_orders_export(request):
    """Export orders to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get orders
    orders = PaymentTransaction.objects.select_related('user').order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="orders_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Reference', 'Customer', 'Email', 'Amount', 'Status', 'Payment Method', 'Date'])
        
        for order in orders:
            writer.writerow([
                order.id,
                order.reference,
                order.user.get_full_name() or order.user.username if order.user else 'Guest',
                order.user.email if order.user else order.email or 'N/A',
                float(order.amount),
                order.status,
                order.payment_method,
                order.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Export"
        
        # Headers
        headers = ['ID', 'Reference', 'Customer', 'Email', 'Amount', 'Status', 'Payment Method', 'Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.reference)
            ws.cell(row=row_num, column=3, value=order.user.get_full_name() or order.user.username if order.user else 'Guest')
            ws.cell(row=row_num, column=4, value=order.user.email if order.user else order.email or 'N/A')
            ws.cell(row=row_num, column=5, value=float(order.amount))
            ws.cell(row=row_num, column=6, value=order.status)
            ws.cell(row=row_num, column=7, value=order.payment_method)
            ws.cell(row=row_num, column=8, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orders_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_courses_export(request):
    """Export courses to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get courses
    courses = Course.objects.all().order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Price', 'Duration (weeks)', 'Videos', 'PDFs', 'Status', 'Created'])
        
        for course in courses:
            writer.writerow([
                course.id,
                course.title,
                float(course.price),
                course.duration_weeks,
                course.video_count(),
                course.pdf_count(),
                'Active' if course.is_active else 'Inactive',
                course.created_at.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses Export"
        
        # Headers
        headers = ['ID', 'Title', 'Price', 'Duration (weeks)', 'Videos', 'PDFs', 'Status', 'Created']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, course in enumerate(courses, 2):
            ws.cell(row=row_num, column=1, value=course.id)
            ws.cell(row=row_num, column=2, value=course.title)
            ws.cell(row=row_num, column=3, value=float(course.price))
            ws.cell(row=row_num, column=4, value=course.duration_weeks)
            ws.cell(row=row_num, column=5, value=course.video_count())
            ws.cell(row=row_num, column=6, value=course.pdf_count())
            ws.cell(row=row_num, column=7, value='Active' if course.is_active else 'Inactive')
            ws.cell(row=row_num, column=8, value=course.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="courses_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_videos_export(request):
    """Export videos to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get videos
    videos = TrainingVideo.objects.select_related('course').order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="videos_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Course', 'Duration (min)', 'Price', 'Views', 'Status', 'Created'])
        
        for video in videos:
            writer.writerow([
                video.id,
                video.title,
                video.course.title if video.course else 'Uncategorized',
                video.duration,
                float(video.price),
                video.view_count,
                'Active' if video.is_active else 'Inactive',
                video.created_at.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Videos Export"
        
        # Headers
        headers = ['ID', 'Title', 'Course', 'Duration (min)', 'Price', 'Views', 'Status', 'Created']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, video in enumerate(videos, 2):
            ws.cell(row=row_num, column=1, value=video.id)
            ws.cell(row=row_num, column=2, value=video.title)
            ws.cell(row=row_num, column=3, value=video.course.title if video.course else 'Uncategorized')
            ws.cell(row=row_num, column=4, value=video.duration)
            ws.cell(row=row_num, column=5, value=float(video.price))
            ws.cell(row=row_num, column=6, value=video.view_count)
            ws.cell(row=row_num, column=7, value='Active' if video.is_active else 'Inactive')
            ws.cell(row=row_num, column=8, value=video.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="videos_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_pdfs_export(request):
    """Export PDFs to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get PDFs
    pdfs = PDF.objects.select_related('course').order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="pdfs_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Course', 'Pages', 'Price', 'Downloads', 'Status', 'Created'])
        
        for pdf in pdfs:
            writer.writerow([
                pdf.id,
                pdf.title,
                pdf.course.title if pdf.course else 'General',
                pdf.pages,
                float(pdf.price),
                pdf.downloads,
                'Active' if pdf.is_active else 'Inactive',
                pdf.created_at.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PDFs Export"
        
        # Headers
        headers = ['ID', 'Title', 'Course', 'Pages', 'Price', 'Downloads', 'Status', 'Created']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, pdf in enumerate(pdfs, 2):
            ws.cell(row=row_num, column=1, value=pdf.id)
            ws.cell(row=row_num, column=2, value=pdf.title)
            ws.cell(row=row_num, column=3, value=pdf.course.title if pdf.course else 'General')
            ws.cell(row=row_num, column=4, value=pdf.pages)
            ws.cell(row=row_num, column=5, value=float(pdf.price))
            ws.cell(row=row_num, column=6, value=pdf.downloads)
            ws.cell(row=row_num, column=7, value='Active' if pdf.is_active else 'Inactive')
            ws.cell(row=row_num, column=8, value=pdf.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pdfs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_packages_export(request):
    """Export packages to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get packages
    packages = Package.objects.all().order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="packages_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Type', 'Price', 'Duration (days)', 'Recurring', 'Sales', 'Revenue', 'Status'])
        
        for pkg in packages:
            writer.writerow([
                pkg.id,
                pkg.name,
                pkg.package_type,
                float(pkg.price),
                pkg.duration_days,
                'Yes' if pkg.is_recurring else 'No',
                pkg.total_sales,
                float(pkg.total_revenue) if pkg.total_revenue else 0,
                'Active' if pkg.is_active else 'Inactive'
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Packages Export"
        
        # Headers
        headers = ['ID', 'Name', 'Type', 'Price', 'Duration (days)', 'Recurring', 'Sales', 'Revenue', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, pkg in enumerate(packages, 2):
            ws.cell(row=row_num, column=1, value=pkg.id)
            ws.cell(row=row_num, column=2, value=pkg.name)
            ws.cell(row=row_num, column=3, value=pkg.package_type)
            ws.cell(row=row_num, column=4, value=float(pkg.price))
            ws.cell(row=row_num, column=5, value=pkg.duration_days)
            ws.cell(row=row_num, column=6, value='Yes' if pkg.is_recurring else 'No')
            ws.cell(row=row_num, column=7, value=pkg.total_sales)
            ws.cell(row=row_num, column=8, value=float(pkg.total_revenue) if pkg.total_revenue else 0)
            ws.cell(row=row_num, column=9, value='Active' if pkg.is_active else 'Inactive')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="packages_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_partnerships_export(request):
    """Export partnerships to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get partnerships
    partnerships = UserPartnership.objects.select_related('user', 'program').order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="partnerships_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'User', 'Email', 'Program', 'Tier', 'Investment', 'Status', 'Created'])
        
        for p in partnerships:
            writer.writerow([
                p.id,
                p.user.get_full_name() or p.user.username,
                p.user.email,
                p.program.name if p.program else 'N/A',
                p.program.get_tier_display() if p.program else 'N/A',
                float(p.investment_amount),
                p.status,
                p.created_at.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Partnerships Export"
        
        # Headers
        headers = ['ID', 'User', 'Email', 'Program', 'Tier', 'Investment', 'Status', 'Created']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, p in enumerate(partnerships, 2):
            ws.cell(row=row_num, column=1, value=p.id)
            ws.cell(row=row_num, column=2, value=p.user.get_full_name() or p.user.username)
            ws.cell(row=row_num, column=3, value=p.user.email)
            ws.cell(row=row_num, column=4, value=p.program.name if p.program else 'N/A')
            ws.cell(row=row_num, column=5, value=p.program.get_tier_display() if p.program else 'N/A')
            ws.cell(row=row_num, column=6, value=float(p.investment_amount))
            ws.cell(row=row_num, column=7, value=p.status)
            ws.cell(row=row_num, column=8, value=p.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="partnerships_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_blogs_export(request):
    """Export blogs to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    
    # Get blogs
    blogs = Blog.objects.all().order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="blogs_export_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Author', 'Category', 'Views', 'Status', 'Published', 'Created'])
        
        for blog in blogs:
            writer.writerow([
                blog.id,
                blog.title,
                blog.author.username if blog.author else 'Admin',
                blog.get_category_display(),
                blog.views,
                blog.status,
                blog.published_at.strftime('%Y-%m-%d') if blog.published_at else 'Not published',
                blog.created_at.strftime('%Y-%m-%d')
            ])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blogs Export"
        
        # Headers
        headers = ['ID', 'Title', 'Author', 'Category', 'Views', 'Status', 'Published', 'Created']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, blog in enumerate(blogs, 2):
            ws.cell(row=row_num, column=1, value=blog.id)
            ws.cell(row=row_num, column=2, value=blog.title)
            ws.cell(row=row_num, column=3, value=blog.author.username if blog.author else 'Admin')
            ws.cell(row=row_num, column=4, value=blog.get_category_display())
            ws.cell(row=row_num, column=5, value=blog.views)
            ws.cell(row=row_num, column=6, value=blog.status)
            ws.cell(row=row_num, column=7, value=blog.published_at.strftime('%Y-%m-%d') if blog.published_at else 'Not published')
            ws.cell(row=row_num, column=8, value=blog.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="blogs_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


@admin_required
def admin_api_revenue_export(request):
    """Export revenue report to Excel/CSV"""
    import csv
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    format_type = request.GET.get('format', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Parse dates
    if start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start = timezone.now() - timedelta(days=30)
    
    if end_date:
        end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    else:
        end = timezone.now()
    
    # Get transactions in date range
    transactions = PaymentTransaction.objects.filter(
        status='completed',
        created_at__range=[start, end]
    ).order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="revenue_report_{start_date}_to_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date', 'Reference', 'Customer', 'Item', 'Amount', 'Payment Method'])
        
        total = 0
        for t in transactions:
            item_name = t.course.title if t.course else t.package_name or t.program_name or 'N/A'
            writer.writerow([
                t.created_at.strftime('%Y-%m-%d %H:%M'),
                t.reference,
                t.user.get_full_name() or t.user.username if t.user else 'Guest',
                item_name,
                float(t.amount),
                t.payment_method
            ])
            total += float(t.amount)
        
        writer.writerow([])
        writer.writerow(['TOTAL', '', '', '', total, ''])
        
        return response
    
    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Revenue Report: {start_date} to {end_date}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Date', 'Reference', 'Customer', 'Item', 'Amount', 'Payment Method']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        total = 0
        for row_num, t in enumerate(transactions, 3):
            item_name = t.course.title if t.course else t.package_name or t.program_name or 'N/A'
            ws.cell(row=row_num, column=1, value=t.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=2, value=t.reference)
            ws.cell(row=row_num, column=3, value=t.user.get_full_name() or t.user.username if t.user else 'Guest')
            ws.cell(row=row_num, column=4, value=item_name)
            ws.cell(row=row_num, column=5, value=float(t.amount))
            ws.cell(row=row_num, column=6, value=t.payment_method)
            total += float(t.amount)
        
        # Total row
        total_row = len(transactions) + 3
        ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total).font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="revenue_report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response


# ==================== DEBUG MEDIA VIEW ====================

def debug_media(request):
    """Debug view to check media files"""
    from django.http import HttpResponse
    import os
    from django.conf import settings
    
    html = "<h1>Media Files Debug</h1>"
    
    # Check media root
    html += f"<h2>Media Root: {settings.MEDIA_ROOT}</h2>"
    html += f"<p>Exists: {os.path.exists(settings.MEDIA_ROOT)}</p>"
    
    # List media directory contents
    if os.path.exists(settings.MEDIA_ROOT):
        html += "<h3>Media Directory Contents:</h3><ul>"
        for root, dirs, files in os.walk(settings.MEDIA_ROOT):
            rel_path = os.path.relpath(root, settings.MEDIA_ROOT)
            if rel_path == '.':
                rel_path = ''
            for file in files:
                full_path = os.path.join(root, file)
                html += f"<li>{rel_path}/{file} - {os.path.getsize(full_path)} bytes</li>"
        html += "</ul>"
    else:
        html += "<p>❌ Media directory does not exist!</p>"
    
    # Check video thumbnails
    from .models import TrainingVideo
    html += "<h3>Video Thumbnails:</h3><ul>"
    for video in TrainingVideo.objects.all():
        if video.thumbnail:
            try:
                url = video.thumbnail.url
                path = video.thumbnail.path
                exists = os.path.exists(path)
                html += f"<li>Video {video.id}: {video.title}<br>URL: {url}<br>Path: {path}<br>Exists: {exists}</li>"
            except Exception as e:
                html += f"<li>Video {video.id}: ERROR - {str(e)}</li>"
        else:
            html += f"<li>Video {video.id}: {video.title} - ❌ No thumbnail</li>"
    html += "</ul>"
    
    # Check PDF covers
    from .models import PDF
    html += "<h3>PDF Covers:</h3><ul>"
    for pdf in PDF.objects.all():
        if pdf.cover_image:
            try:
                url = pdf.cover_image.url
                path = pdf.cover_image.path
                exists = os.path.exists(path)
                html += f"<li>PDF {pdf.id}: {pdf.title}<br>URL: {url}<br>Path: {path}<br>Exists: {exists}</li>"
            except Exception as e:
                html += f"<li>PDF {pdf.id}: ERROR - {str(e)}</li>"
        else:
            html += f"<li>PDF {pdf.id}: {pdf.title} - ❌ No cover</li>"
    html += "</ul>"
    
    return HttpResponse(html)

# ==================== TEST COURSE SIMPLE ====================

@csrf_exempt
def test_course_simple(request):
    """Simple test to create a course - CSRF exempt for testing"""
    from django.http import HttpResponse
    from decimal import Decimal
    
    if request.method == 'POST':
        try:
            course = Course(
                title=request.POST.get('title', 'Test Course'),
                description="Test Description",
                price=Decimal(request.POST.get('price', '99.99')),
                duration_weeks=4,
                is_active=True
            )
            course.save()
            
            return HttpResponse(f"""
            <h1>✅ Course Created Successfully!</h1>
            <p>ID: {course.id}</p>
            <p>Title: {course.title}</p>
            <p>Price: ${course.price}</p>
            <p><a href="/admin/debug-courses-api/">View All Courses</a></p>
            <p><a href="/admin/">Back to Admin</a></p>
            """)
        except Exception as e:
            return HttpResponse(f"""
            <h1>❌ Error</h1>
            <pre>{e}</pre>
            <p><a href="/admin/test-course-simple/">Try Again</a></p>
            """)
    
    return HttpResponse("""
    <html>
    <body>
        <h1>Test Course Creation</h1>
        <form method="post">
            <input type="text" name="title" placeholder="Course Title" value="Test Course"><br>
            <input type="text" name="price" placeholder="Price" value="99.99"><br>
            <button type="submit">Create Test Course</button>
        </form>
        <p><a href="/admin/">Back to Admin</a></p>
    </body>
    </html>
    """)

# ==================== HELPER FUNCTIONS ====================

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def timesince(dt, default="just now"):
    """Return human readable time since"""
    if not dt:
        return default
    now = timezone.now()
    diff = now - dt
    
    seconds = diff.total_seconds()
    if seconds < 60:
        return "just now"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds // 3600)
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif seconds < 604800:
        days = int(seconds // 86400)
        return f"{days} day{'s' if days != 1 else ''} ago"
    elif seconds < 2592000:
        weeks = int(seconds // 604800)
        return f"{weeks} week{'s' if weeks != 1 else ''} ago"
    elif seconds < 31536000:
        months = int(seconds // 2592000)
        return f"{months} month{'s' if months != 1 else ''} ago"
    else:
        years = int(seconds // 31536000)
        return f"{years} year{'s' if years != 1 else ''} ago"



# Add to your admin_views.py - SIMPLE VERSION (NO @admin_required)

# ==================== SIMPLE MERCHANDISE API ====================

def admin_api_merchandise(request):
    """Get all merchandise - simple"""
    items = Merchandise.objects.filter(is_active=True).order_by('-created_at')
    data = [{'id': i.id, 'name': i.name, 'price_usd': float(i.price_usd), 'image': i.image.url if i.image else None} for i in items]
    return JsonResponse(data, safe=False)


@csrf_exempt
def admin_api_merchandise_create(request):
    """Create merchandise - simple"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    name = request.POST.get('name')
    price_usd = request.POST.get('price_usd')
    image = request.FILES.get('image')
    
    if not name or not price_usd:
        return JsonResponse({'error': 'Name and price required'}, status=400)
    
    item = Merchandise(name=name, price_usd=Decimal(price_usd))
    if image:
        item.image = image
    item.save()
    
    return JsonResponse({'success': True, 'id': item.id})


@require_POST
def admin_api_merchandise_delete(request, item_id):
    """Delete merchandise"""
    item = get_object_or_404(Merchandise, id=item_id)
    item.delete()
    return JsonResponse({'success': True})


# ==================== SIMPLE ORDERS API ====================

def admin_api_orders(request):
    """Get all merchandise orders"""
    orders = MerchandiseOrder.objects.all().order_by('-created_at')
    data = [{
        'id': o.id,
        'order_number': o.order_number,
        'customer_name': o.customer_name,
        'customer_email': o.customer_email,
        'total_usd': float(o.total_usd),
        'payment_status': o.payment_status,
        'order_status': o.order_status,
        'items': o.items,
        'created_at': o.created_at.strftime('%Y-%m-%d %H:%M'),
        'tracking_number': o.tracking_number
    } for o in orders]
    return JsonResponse(data, safe=False)


@csrf_exempt
def admin_api_order_update(request, order_id):
    """Update order status and tracking"""
    order = get_object_or_404(MerchandiseOrder, id=order_id)
    data = json.loads(request.body)
    
    if 'order_status' in data:
        order.order_status = data['order_status']
    if 'tracking_number' in data:
        order.tracking_number = data['tracking_number']
    order.save()
    
    return JsonResponse({'success': True})


# ==================== SIMPLE TICKET API ====================

def admin_api_ticket_types(request):
    """Get all ticket types"""
    types = TicketType.objects.filter(is_active=True)
    data = [{'id': t.id, 'name': t.name, 'price_usd': float(t.price_usd)} for t in types]
    return JsonResponse(data, safe=False)


@csrf_exempt
def admin_api_ticket_type_create(request):
    """Create ticket type"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    data = json.loads(request.body)
    ticket_type = TicketType(name=data['name'], price_usd=Decimal(data['price_usd']))
    ticket_type.save()
    return JsonResponse({'success': True, 'id': ticket_type.id})


def admin_api_ticket_purchases(request):
    """Get all ticket purchases"""
    tickets = TicketPurchase.objects.all().order_by('-created_at')
    data = [{
        'id': t.id,
        'ticket_number': t.ticket_number,
        'customer_name': t.customer_name,
        'customer_email': t.customer_email,
        'ticket_type': t.ticket_type.name if t.ticket_type else 'N/A',
        'price_usd': float(t.price_usd),
        'payment_status': t.payment_status,
        'checked_in': t.checked_in,
        'created_at': t.created_at.strftime('%Y-%m-%d %H:%M')
    } for t in tickets]
    return JsonResponse(data, safe=False)


@csrf_exempt
def admin_api_ticket_checkin(request, ticket_id):
    """Check in a ticket"""
    ticket = get_object_or_404(TicketPurchase, id=ticket_id)
    if ticket.checked_in:
        return JsonResponse({'error': 'Already checked in'}, status=400)
    
    ticket.checked_in = True
    ticket.checked_in_at = timezone.now()
    ticket.save()
    
    return JsonResponse({'success': True, 'checked_in_at': ticket.checked_in_at.strftime('%Y-%m-%d %H:%M')})       


   
# ========== HELPER FUNCTIONS ==========
def send_ticket_email(ticket):
    """Send ticket email with downloadable ticket"""
    subject = f"Your Ticket for {ticket.event.title} - {ticket.ticket_number}"
    
    # Generate QR code URL (using QR Server API)
    qr_data = f"{ticket.ticket_number}|{ticket.attendee_name}|{ticket.event.title}|{ticket.event.date}"
    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={qr_data}"
    
    html_message = f"""
    <!DOCTYPE html>
    <html>
    <head><style>
        body {{ font-family: Arial, sans-serif; }}
        .ticket {{ border: 2px solid #FFD700; border-radius: 10px; padding: 20px; max-width: 500px; margin: 0 auto; }}
        .header {{ text-align: center; border-bottom: 1px solid #ccc; padding-bottom: 10px; }}
        .details {{ margin: 20px 0; }}
        .row {{ display: flex; justify-content: space-between; margin: 10px 0; }}
        .qr {{ text-align: center; margin-top: 20px; }}
        .footer {{ text-align: center; font-size: 12px; color: #666; margin-top: 20px; }}
    </style></head>
    <body>
        <div class="ticket">
            <div class="header">
                <h2>MFALME BETTERDAYS CAPITAL</h2>
                <p>Elite Trading & Investment</p>
            </div>
            <div class="details">
                <div class="row"><strong>Ticket Number:</strong> <span>{ticket.ticket_number}</span></div>
                <div class="row"><strong>Attendee:</strong> <span>{ticket.attendee_name}</span></div>
                <div class="row"><strong>Phone:</strong> <span>{ticket.attendee_phone}</span></div>
                <div class="row"><strong>Email:</strong> <span>{ticket.attendee_email}</span></div>
                <div class="row"><strong>Event:</strong> <span>{ticket.event.title}</span></div>
                <div class="row"><strong>Date:</strong> <span>{ticket.event.date.strftime('%B %d, %Y at %I:%M %p')}</span></div>
                <div class="row"><strong>Venue:</strong> <span>{ticket.event.venue}</span></div>
                <div class="row"><strong>Quantity:</strong> <span>{ticket.quantity}</span></div>
                <div class="row"><strong>Total Paid:</strong> <span>KES {ticket.total_amount_kes:,.2f}</span></div>
            </div>
            <div class="qr">
                <img src="{qr_url}" alt="QR Code" width="150">
                <p>Scan this QR code at the entrance</p>
            </div>
            <div class="footer">
                <p>Present this ticket at registration desk | For inquiries: +254706286667</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    send_mail(
        subject=subject,
        message=f"Your ticket for {ticket.event.title}\n\nTicket Number: {ticket.ticket_number}\nTotal: KES {ticket.total_amount_kes:,.2f}\n\nPresent this ticket at the entrance.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[ticket.attendee_email],
        fail_silently=False,
        html_message=html_message
    )
    
    # Send admin notification
    send_mail(
        subject=f"New Ticket Purchase - {ticket.ticket_number}",
        message=f"New ticket purchased!\n\nName: {ticket.attendee_name}\nPhone: {ticket.attendee_phone}\nEmail: {ticket.attendee_email}\nQuantity: {ticket.quantity}\nTotal: KES {ticket.total_amount_kes:,.2f}",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[settings.ADMIN_EMAIL],
        fail_silently=True,
    )


def send_merchandise_order_email(order):
    """Send merchandise order confirmation email"""
    subject = f"Order Confirmation - {order.order_number}"
    
    items_html = ""
    for item in order.items:
        items_html += f"<tr><td>{item['name']}</td><td>{item['quantity']}</td><td>KES {item['price']:,.2f}</td><td>KES {item['price'] * item['quantity']:,.2f}</td></tr>"
    
    html_message = f"""
    <!DOCTYPE html>
    <html>
    <head><style>
        body {{ font-family: Arial, sans-serif; }}
        .order {{ max-width: 600px; margin: 0 auto; }}
        .header {{ text-align: center; border-bottom: 2px solid #FFD700; padding-bottom: 10px; }}
        .details {{ margin: 20px 0; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background: #FFD700; color: #0a1520; }}
        .total {{ text-align: right; margin-top: 20px; font-size: 18px; }}
    </style></head>
    <body>
        <div class="order">
            <div class="header">
                <h2>MFALME BETTERDAYS CAPITAL</h2>
                <p>Order Confirmation</p>
            </div>
            <div class="details">
                <p><strong>Order Number:</strong> {order.order_number}</p>
                <p><strong>Customer:</strong> {order.customer_name}</p>
                <p><strong>Phone:</strong> {order.customer_phone}</p>
                <p><strong>Delivery Address:</strong> {order.delivery_address}</p>
            </div>
            <table>
                <tr><th>Item</th><th>Quantity</th><th>Unit Price</th><th>Total</th></tr>
                {items_html}
            </table>
            <div class="total">
                <p><strong>Subtotal: KES {order.subtotal:,.2f}</strong></p>
                <p><strong>Shipping: KES {order.shipping_cost:,.2f}</strong></p>
                <p><strong>Total: KES {order.total:,.2f}</strong></p>
            </div>
            <p>Your order will be processed and shipped within 3-5 business days.</p>
            <p>For inquiries, contact +254706286667</p>
        </div>
    </body>
    </html>
    """
    
    send_mail(
        subject=subject,
        message=f"Your order {order.order_number} has been confirmed. Total: KES {order.total:,.2f}",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[order.customer_email],
        fail_silently=False,
        html_message=html_message
    )
    
    # Send admin notification
    send_mail(
        subject=f"New Merchandise Order - {order.order_number}",
        message=f"New order received!\n\nCustomer: {order.customer_name}\nPhone: {order.customer_phone}\nTotal: KES {order.total:,.2f}\n\nView in admin panel.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[settings.ADMIN_EMAIL],
        fail_silently=True,
    )


# ========== MERCHANDISE CRUD ==========
@require_http_methods(["GET"])
def get_merchandise(request):
    items = Merchandise.objects.filter(status='active')
    data = [{
        'id': i.id, 'name': i.name, 'category': i.category,
        'description': i.description, 'price': float(i.price),
        'image': i.image, 'stock': i.stock, 'status': i.status
    } for i in items]
    return JsonResponse({'items': data})


@require_http_methods(["POST"])
def create_merchandise(request):
    try:
        data = json.loads(request.body)
        merch = Merchandise.objects.create(
            name=data['name'],
            category=data.get('category', 'apparel'),
            description=data.get('description', ''),
            price=Decimal(data['price']),
            stock=data.get('stock', 0),
            image=data.get('image', ''),
            image_key=data.get('image_key', '')
        )
        return JsonResponse({'success': True, 'id': merch.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def update_merchandise(request, id):
    try:
        merch = Merchandise.objects.get(id=id)
        data = json.loads(request.body)
        merch.name = data.get('name', merch.name)
        merch.category = data.get('category', merch.category)
        merch.description = data.get('description', merch.description)
        merch.price = Decimal(data.get('price', merch.price))
        merch.stock = data.get('stock', merch.stock)
        merch.status = data.get('status', merch.status)
        if data.get('image'):
            merch.image = data['image']
        merch.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def delete_merchandise(request, id):
    try:
        Merchandise.objects.get(id=id).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== MERCHANDISE ORDERS ==========
@require_http_methods(["GET"])
def get_merchandise_orders(request):
    orders = MerchandiseOrder.objects.all().order_by('-created_at')
    data = [{
        'id': o.id, 'order_number': o.order_number, 'customer_name': o.customer_name,
        'customer_phone': o.customer_phone, 'customer_email': o.customer_email,
        'total': float(o.total), 'status': o.status, 'created_at': o.created_at.strftime('%Y-%m-%d %H:%M')
    } for o in orders]
    return JsonResponse({'orders': data})


@require_http_methods(["POST"])
def update_merchandise_order_status(request, id):
    try:
        data = json.loads(request.body)
        order = MerchandiseOrder.objects.get(id=id)
        order.status = data.get('status', order.status)
        order.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== EVENTS ==========
@require_http_methods(["GET"])
def get_events(request):
    events = Event.objects.filter(is_active=True)
    data = [{
        'id': e.id, 'title': e.title, 'description': e.description,
        'date': e.date.isoformat(), 'venue': e.venue,
        'ticket_price_usd': float(e.ticket_price_usd),
        'poster_image': e.poster_image,
        'max_attendees': e.max_attendees,
        'current_bookings': e.current_bookings,
        'seats_remaining': e.seats_remaining,
        'is_sold_out': e.is_sold_out
    } for e in events]
    return JsonResponse({'events': data})


@require_http_methods(["GET"])
def get_event_detail(request, id):
    try:
        event = Event.objects.get(id=id)
        data = {
            'id': event.id, 'title': event.title, 'description': event.description,
            'date': event.date.isoformat(), 'venue': event.venue,
            'ticket_price_usd': float(event.ticket_price_usd),
            'poster_image': event.poster_image,
            'seats_remaining': event.seats_remaining
        }
        return JsonResponse(data)
    except Event.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)


@require_http_methods(["POST"])
def update_event(request, id):
    try:
        event = Event.objects.get(id=id)
        data = json.loads(request.body)
        event.title = data.get('title', event.title)
        event.description = data.get('description', event.description)
        event.date = data.get('date', event.date)
        event.venue = data.get('venue', event.venue)
        event.ticket_price_usd = Decimal(data.get('ticket_price_usd', event.ticket_price_usd))
        event.max_attendees = data.get('max_attendees', event.max_attendees)
        event.is_active = data.get('is_active', event.is_active)
        if data.get('poster_image'):
            event.poster_image = data['poster_image']
        event.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== TICKETS ==========
@require_http_methods(["GET"])
def get_tickets(request):
    tickets = EventTicket.objects.all().order_by('-created_at')
    data = [{
        'id': t.id, 'ticket_number': t.ticket_number, 'attendee_name': t.attendee_name,
        'attendee_phone': t.attendee_phone, 'attendee_email': t.attendee_email,
        'quantity': t.quantity, 'total_amount_kes': float(t.total_amount_kes),
        'event_title': t.event.title, 'status': t.status,
        'checked_in': t.checked_in, 'created_at': t.created_at.strftime('%Y-%m-%d %H:%M')
    } for t in tickets]
    return JsonResponse({'tickets': data})


@require_http_methods(["GET"])
def get_ticket_detail(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        data = {
            'id': ticket.id, 'ticket_number': ticket.ticket_number,
            'attendee_name': ticket.attendee_name, 'attendee_phone': ticket.attendee_phone,
            'attendee_email': ticket.attendee_email, 'quantity': ticket.quantity,
            'total_amount_kes': float(ticket.total_amount_kes),
            'event_title': ticket.event.title, 'event_date': ticket.event.date.isoformat(),
            'event_venue': ticket.event.venue, 'status': ticket.status,
            'checked_in': ticket.checked_in, 'checked_in_at': ticket.checked_in_at
        }
        return JsonResponse(data)
    except EventTicket.DoesNotExist:
        return JsonResponse({'error': 'Ticket not found'}, status=404)


@require_http_methods(["POST"])
def resend_ticket_email(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        send_ticket_email(ticket)
        return JsonResponse({'success': True, 'message': 'Ticket email resent successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_http_methods(["POST"])
def mark_ticket_checked_in(request, id):
    try:
        ticket = EventTicket.objects.get(id=id)
        ticket.checked_in = True
        ticket.checked_in_at = datetime.now()
        ticket.status = 'attended'
        ticket.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== ORDER CREATION ==========
@require_http_methods(["POST"])
def create_order(request):
    try:
        data = json.loads(request.body)
        
        order = Order.objects.create(
            customer_name=data['name'],
            customer_email=data.get('email', ''),
            customer_phone=data['phone'],
            item_type=data.get('type', 'ticket'),
            items=data.get('items', []),
            amount=Decimal(data['amount']),
            metadata={
                'address': data.get('address', ''),
                'quantity': data.get('quantity', 1)
            }
        )
        
        return JsonResponse({
            'success': True,
            'reference': order.reference,
            'order_id': order.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ========== SASAPAY INTEGRATION ==========
# Import SasaPay configuration from settings (Production)
SASAPAY_API_URL = getattr(settings, 'SASAPAY_API_URL', 'https://api.sasapay.app/api/v1')
SASAPAY_CLIENT_ID = getattr(settings, 'SASAPAY_CLIENT_ID', None)
SASAPAY_CLIENT_SECRET = getattr(settings, 'SASAPAY_CLIENT_SECRET', None)
SASAPAY_MERCHANT_CODE = getattr(settings, 'SASAPAY_MERCHANT_CODE', '600980')

# Aliases for compatibility with existing code
SASAPAY_API_KEY = SASAPAY_CLIENT_ID
SASAPAY_API_SECRET = SASAPAY_CLIENT_SECRET  
SASAPAY_SHORTCODE = SASAPAY_MERCHANT_CODE

# Verify production configuration
if not all([SASAPAY_CLIENT_ID, SASAPAY_CLIENT_SECRET]):
    print("⚠️ WARNING: SasaPay production credentials not configured!")

def generate_sasapay_signature(data):
    """Generate HMAC SHA256 signature for SasaPay"""
    secret = SASAPAY_API_SECRET
    sorted_data = {k: data[k] for k in sorted(data.keys())}
    sign_string = ""
    for k, v in sorted_data.items():
        sign_string += f"{k}{v}"
    signature = hmac.new(secret.encode(), sign_string.encode(), hashlib.sha256).hexdigest()
    return signature


@require_http_methods(["POST"])
def sasapay_stk_push(request):
    try:
        data = json.loads(request.body)
        phone = data['phone'].strip()
        amount = str(int(float(data['amount'])))
        reference = data['reference']
        description = data.get('description', 'MBC Payment')
        
        # Format phone number (remove 0 or +254 prefix)
        if phone.startswith('0'):
            phone = '254' + phone[1:]
        elif phone.startswith('+'):
            phone = phone[1:]
        
        payload = {
            'shortcode': SASAPAY_SHORTCODE,
            'amount': amount,
            'phone': phone,
            'reference': reference,
            'description': description,
            'callback_url': f"{settings.SITE_URL}/api/sasapay/callback/"
        }
        
        payload['signature'] = generate_sasapay_signature(payload)
        
        response = requests.post(
            f"{SASAPAY_API_URL}/stkpush",
            json=payload,
            headers={'Content-Type': 'application/json', 'Api-Key': SASAPAY_API_KEY},
            timeout=30
        )
        
        result = response.json()
        
        if result.get('success'):
            # Update order with checkout_request_id
            Order.objects.filter(reference=reference).update(
                checkout_request_id=result.get('checkout_request_id'),
                payment_reference=result.get('checkout_request_id')
            )
            return JsonResponse({
                'success': True,
                'checkout_request_id': result.get('checkout_request_id'),
                'message': 'STK Push sent successfully'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('message', 'STK Push failed')
            }, status=400)
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def sasapay_check_status(request, checkout_id):
    try:
        response = requests.get(
            f"{SASAPAY_API_URL}/status/{checkout_id}",
            headers={'Content-Type': 'application/json', 'Api-Key': SASAPAY_API_KEY},
            timeout=30
        )
        result = response.json()
        
        return JsonResponse({
            'status': result.get('status', 'pending'),
            'message': result.get('message', ''),
            'data': result
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def sasapay_callback(request):
    """Handle SasaPay callback after payment"""
    try:
        data = json.loads(request.body)
        checkout_request_id = data.get('checkout_request_id')
        status = data.get('status')
        reference = data.get('reference')
        
        if status == 'completed':
            # Update order
            order = Order.objects.filter(checkout_request_id=checkout_request_id).first()
            if order:
                order.status = 'completed'
                order.save()
                
                if order.item_type == 'ticket':
                    # Create ticket record
                    event = Event.objects.filter(is_active=True).first()
                    if event:
                        ticket = EventTicket.objects.create(
                            event=event,
                            attendee_name=order.customer_name,
                            attendee_phone=order.customer_phone,
                            attendee_email=order.customer_email,
                            quantity=order.metadata.get('quantity', 1),
                            unit_price_usd=249,
                            unit_price_kes=249 * 128,
                            order_reference=order.reference,
                            payment_reference=checkout_request_id,
                            status='confirmed'
                        )
                        event.current_bookings += ticket.quantity
                        event.save()
                        send_ticket_email(ticket)
                        
                elif order.item_type == 'merchandise':
                    # Create merchandise order record
                    merch_order = MerchandiseOrder.objects.create(
                        customer_name=order.customer_name,
                        customer_phone=order.customer_phone,
                        customer_email=order.customer_email,
                        delivery_address=order.metadata.get('address', ''),
                        items=order.items,
                        subtotal=order.amount,
                        total=order.amount,
                        payment_reference=checkout_request_id,
                        order_reference=order.reference,
                        status='paid'
                    )
                    send_merchandise_order_email(merch_order)
                    
                    # Update stock
                    for item in order.items:
                        try:
                            product = Merchandise.objects.get(id=item['id'])
                            product.stock -= item['quantity']
                            product.save()
                        except:
                            pass
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})